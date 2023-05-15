'''
Change History

'''
from datetime import datetime
import json
import numpy as np
import traceback
from test_mongo import MongoExplorer
import pandas as pd
from pandas import json_normalize
from test_yahoo import QuoteExplorer
from test_g20_v2 import StockScore
from test_stockstats import StockStats
from openpyxl import load_workbook
from pathlib import Path
from copy import copy
from typing import Union, Optional
#import numpy as np
#import pandas as pd
import openpyxl
#from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class PyExcel:
    def copy_excel_cell_range(
            self,
            src_ws: openpyxl.worksheet.worksheet.Worksheet,
            min_row: int = None,
            max_row: int = None,
            min_col: int = None,
            max_col: int = None,
            tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
            tgt_min_row: int = 1,
            tgt_min_col: int = 1,
            with_style: bool = True
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        copies all cells from the source worksheet [src_ws] starting from [min_row] row
        and [min_col] column up to [max_row] row and [max_col] column
        to target worksheet [tgt_ws] starting from [tgt_min_row] row
        and [tgt_min_col] column.

        @param src_ws:  source worksheet
        @param min_row: smallest row index in the source worksheet (1-based index)
        @param max_row: largest row index in the source worksheet (1-based index)
        @param min_col: smallest column index in the source worksheet (1-based index)
        @param max_col: largest column index in the source worksheet (1-based index)
        @param tgt_ws:  target worksheet.
                        If None, then the copy will be done to the same (source) worksheet.
        @param tgt_min_row: target row index (1-based index)
        @param tgt_min_col: target column index (1-based index)
        @param with_style:  whether to copy cell style. Default: True

        @return: target worksheet object
        """
        if tgt_ws is None:
            tgt_ws = src_ws

        # https://stackoverflow.com/a/34838233/5741205
        for row in src_ws.iter_rows(min_row=min_row, max_row=max_row,
                                    min_col=min_col, max_col=max_col):
            for cell in row:
                tgt_cell = tgt_ws.cell(
                    row=cell.row + tgt_min_row - 1,
                    column=cell.col_idx + tgt_min_col - 1,
                    value=cell.value
                )
                if with_style and cell.has_style:
                    # tgt_cell._style = copy(cell._style)
                    tgt_cell.font = copy(cell.font)
                    tgt_cell.border = copy(cell.border)
                    tgt_cell.fill = copy(cell.fill)
                    tgt_cell.number_format = copy(cell.number_format)
                    tgt_cell.protection = copy(cell.protection)
                    tgt_cell.alignment = copy(cell.alignment)
        return tgt_ws

    def append_df_to_excel(
            self,
            filename: Union[str, Path],
            df: pd.DataFrame,
            sheet_name: str = 'Sheet1',
            startrow: Optional[int] = None,
            max_col_width: int = 30,
            autofilter: bool = False,
            fmt_int: str = "#,##0",
            fmt_float: str = "#,##0.00",
            fmt_date: str = "yyyy-mm-dd",
            fmt_datetime: str = "yyyy-mm-dd hh:mm",
            truncate_sheet: bool = False,
            storage_options: Optional[dict] = None,
            **to_excel_kwargs
    ) -> None:
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        @param filename: File path or existing ExcelWriter
                         (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                           (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                         Per default (startrow=None) calculate the last row
                         in the existing DF and write to the next row...
        @param max_col_width: maximum column width in Excel. Default: 40
        @param autofilter: boolean - whether add Excel autofilter or not. Default: False
        @param fmt_int: Excel format for integer numbers
        @param fmt_float: Excel format for float numbers
        @param fmt_date: Excel format for dates
        @param fmt_datetime: Excel format for datetime's
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                               before writing DataFrame to Excel file
        @param storage_options: dict, optional
            Extra options that make sense for a particular storage connection, e.g. host, port,
            username, password, etc., if using a URL that will be parsed by fsspec, e.g.,
            starting “s3://”, “gcs://”.
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None

        Usage examples:

        >>> append_df_to_excel('/tmp/test.xlsx', df, autofilter=True,
                               freeze_panes=(1,0))

        >>> append_df_to_excel('/tmp/test.xlsx', df, header=None, index=False)

        >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                               index=False)

        >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                               index=False, startrow=25)

        >>> append_df_to_excel('/tmp/test.xlsx', df, index=False,
                               fmt_datetime="dd.mm.yyyy hh:mm")

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        def set_column_format(ws, column_letter, fmt):
            for cell in ws[column_letter]:
                cell.number_format = fmt
        filename = Path(filename)
        file_exists = filename.is_file()
        # process parameters
        # calculate first column number
        # if the DF will be written using `index=True`, then `first_col = 2`, else `first_col = 1`
        first_col = int(to_excel_kwargs.get("index", True)) + 1
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')
        # save content of existing sheets
        if file_exists:
            wb = load_workbook(filename)
            sheet_names = wb.sheetnames
            sheet_exists = sheet_name in sheet_names
            sheets = {ws.title: ws for ws in wb.worksheets}

        with pd.ExcelWriter(
            filename.with_suffix(".xlsx"),
            engine="openpyxl",
            mode="a" if file_exists else "w",
            if_sheet_exists="new" if file_exists else None,
            date_format=fmt_date,
            datetime_format=fmt_datetime,
            storage_options=storage_options
        ) as writer:
            if file_exists:
                # try to open an existing workbook
                writer.book = wb
                # get the last row in the existing Excel sheet
                # if it was not specified explicitly
                if startrow is None and sheet_name in writer.book.sheetnames:
                    startrow = writer.book[sheet_name].max_row
                # truncate sheet
                if truncate_sheet and sheet_name in writer.book.sheetnames:
                    # index of [sheet_name] sheet
                    idx = writer.book.sheetnames.index(sheet_name)
                    # remove [sheet_name]
                    writer.book.remove(writer.book.worksheets[idx])
                    # create an empty sheet [sheet_name] using old index
                    writer.book.create_sheet(sheet_name, idx)
                # copy existing sheets
                writer.sheets = sheets
            else:
                # file doesn't exist, we are creating a new one
                startrow = 0

            # write out the DataFrame to an ExcelWriter
            df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
            worksheet = writer.sheets[sheet_name]

            if autofilter:
                worksheet.auto_filter.ref = worksheet.dimensions

            for xl_col_no, dtyp in enumerate(df.dtypes, first_col):
                col_no = xl_col_no - first_col
                width = max(df.iloc[:, col_no].astype(str).str.len().max(),
                            len(df.columns[col_no]) + 6)
                width = min(max_col_width, width)
                column_letter = get_column_letter(xl_col_no)
                worksheet.column_dimensions[column_letter].width = width
                if np.issubdtype(dtyp, np.integer):
                    set_column_format(worksheet, column_letter, fmt_int)
                if np.issubdtype(dtyp, np.floating):
                    set_column_format(worksheet, column_letter, fmt_float)

        if file_exists and sheet_exists:
            # move (append) rows from new worksheet to the `sheet_name` worksheet
            wb = load_workbook(filename)
            # retrieve generated worksheet name
            new_sheet_name = set(wb.sheetnames) - set(sheet_names)
            if new_sheet_name:
                new_sheet_name = list(new_sheet_name)[0]
            # copy rows written by `df.to_excel(...)` to
            self.copy_excel_cell_range(
                src_ws=wb[new_sheet_name],
                tgt_ws=wb[sheet_name],
                tgt_min_row=startrow + 1,
                with_style=True
            )
            # remove new (generated by Pandas) worksheet
            del wb[new_sheet_name]
            wb.save(filename)
            wb.close()

    def to_excel(self, df, excel_file_name = '', ):
        self.append_df_to_excel(excel_file_name, df,
                startcol=0, header=False,
                columns=[
                    'symbol',
                    'Reason',
                    'Recommendation',
                    'model_run_date',
                    'start_date',
                    'end_date',
                    'duration',
                    'model_gain',
                    'model_perf',
                    'buy_and_hold_gain',
                    'buy_and_hold_perf',
                    'model_score',
                    'predict_date',
                    'predict_macd_accum',
                    'predict_macd_len',
                    'predict_action',
                    'predict_vol',
                    'AAOD',
                    'close',
                    'Score',
                    'CAP',
                    'CapScore',
                    'PEScore',
                    'AveragePE',
                    'PE',
                    'EPS',
                    'IndustryRank',
                    'Rev Growth',
                    'Gross Margin',
                    'G20',
                    'SR20',
                    'G20Close',
                    'G20Date',
                    'G20Total',
                    'G10',
                    'SR10',
                    'G10Close',
                    'G10Date',
                    'G10Total',
                    'G5',
                    'SR5',
                    'G5Close',
                    'G5Date',
                    'G5Total',
                    'G1',
                    'SR1',
                    'G1Close',
                    'G1Date',
                    'G1Total',
                    'runDate',
                    'macd_sign',
                    'peak',
                    'peak_date',
                    'accum',
                    'len',
                    'r',
                    'pre_macd_sign',
                    'pre_peak',
                    'pre_peak_date',
                    'pre_accum',
                    'pre_len'
                ])

if __name__ == '__main__':
    d = PyExcel()
    d.to_excel()


'''
TickHistory Ticker Coverage
===========================

Below is the listing of all tickers currently covered by TickHistory





Ticker	Name
-----------------------------------------------------
A	Agilent Technologies
AA	ALCOA CORPORATION
AAAU	Perth Mint Physical Gold ETF
AAL	American Airlines
AAN	Aaron's
AAOI	Applied Optoelectronics
AAON	Aaon
AAP	Advance Auto Parts
AAPL	Apple
AAT	American Assets Trust
AAWW	Atlas Air Worldwide
AAXJ	ISHARES MSCI ALL COUNTRY ASIA
AB	AllianceBernstein L.P.
ABBV	AbbVie
ABC	AmerisourceBergen
ABCB	Ameris Bancorp
ABCL	AbCellera Biologics
ABCM	Abcam
ABEO	Abeona Therapeutics
ABG	Asbury Automotive
ABM	ABM Industries
ABMD	ABIOMED
ABNB	Airbnb
ABOS	Acumen Pharmaceuticals
ABR	Arbor Realty Trust
ABSI	Absci
ABST	Absolute Software
ABT	Abbott Laboratories
ABTX	Allegiance Bancshares
AC	Associated Capital
ACA	Arcosa
ACAD	ACADIA Pharmaceuticals
ACB	Aurora Cannabis
ACC	American Campus Communities
ACCD	Accolade
ACCO	Acco Brands
ACEL	Accel Entertainment
ACES	ALPS Clean Energy ETF
ACGL	Arch Capital
ACHC	Acadia Healthcare
ACHL	ACHILLES THERAPEUTICS
ACHR	Archer Aviation
ACI	Albertsons Companies
ACIU	AC Immune SA
ACIW	ACI Worldwide
ACLS	Axcelis Technologies
ACM	AECOM
ACMR	ACM Research
ACN	Accenture
ACNB	Acnb
ACQR	Independence
ACRE	Ares Commercial Real
ACRS	Aclaris Therapeutics
ACRX	Acelrx Pharmaceuticals
ACTG	Acacia Research
ACV	Alberto-Culver
ACVA	ACV Auctions
ACWI	ISHARES MSCI ACWI INDEX FUND
ACWX	iShares MSCI ACWI ex U.S. ETF
ACXP	Acurx Pharmaceuticals
ADAG	Adagene
ADBE	Adobe Systems
ADC	Agree Realty
ADES	Advanced Emissions Solutions
ADGI	Adagio Therapeutics
ADI	Analog Devices
ADM	Archer-Daniels-Midland Co
ADMA	Adma Biologics Cmn
ADNT	Adient
ADP	Automatic Data Processing
ADPT	Adaptive Biotechnologies
ADRE	Invesco BLDRS Emerging Markets 50 ADR Index Fund
ADS	Alliance Data Systems
ADSK	Autodesk
ADT	ADT
ADTN	Adtran
ADTX	Aditx Therapeutics
ADUS	Addus Homecare
ADV	Advantage Solutions
ADVM	Adverum Biotechnologies
AE	Adams Resources
AEE	Ameren
AEG	Aegon N.V.
AEIS	Advanced Energy Inds
AEL	American Equity Investment Life
AEM	Agnico Eagle Mines
AEO	American Eagle Outfitters
AEP	American Electric Power
AER	AerCap N.V.
AERI	Aerie Pharmaceuticals Co
AES	AES
AEVA	Aeva Technologies
AFCG	AFC Gamma
AFG	American Financial
AFIB	Acutus Medical
AFL	AFLAC
AFMD	Affimed N.V.
AFRM	Affirm
AFYA	Afya
AG	First Majestic Silver
AGCB	Altimeter Growth 2
AGCO	AGCO
AGEN	Agenus
AGFS	Agrofresh Solutions
AGFY	Agrify
AGG	ISHARES BARCLAYS AGGREGATE BOND
AGGY	WisdomTree Barclays Yield Enhanced U.S. Aggregate Bond Fund
AGI	Alamos Gold
AGIO	Agios Pharmaceuticals
AGL	agilon health
AGLE	Aeglea Biotherapeutics
AGM	Federal Agricultural Mortgage
AGNC	AGNC Investment
AGO	Assured Guaranty
AGQ	PROSHARES ULTRA SILVER
AGR	Avangrid
AGRI	AgriFORCE Growing Systems
AGRO	Adecoagro S.A.
AGRX	Agile Therapeutics Commo
AGS	Playags
AGTC	Applied Genetic Technologies c
AGTI	Agiliti
AGX	Argan
AGYS	Agilysys
AHCO	Adapthealth
AHH	Armada Hoffler Properties
AIA	iShares S&P Asia ETF
AIG	American International
AIMC	Altra Industrial Motion
AIN	Albany International
AINV	Apollo Investment
AIQ	Global X Future Analytics Tech ETF
AIR	Aar
AIRC	Taxable Apartment ome REIT
AIRR	First Trust RBA American Industrial Renaissance ETF
AIT	Applied Industrial Technologies
AIV	Apartment Investment & Management
AIZ	Assurant
AJG	Arthur J. Gallagher &
AJRD	Aerojet Rocketdyne
AJX	Great Ajax
AKAM	Akamai Technologies
AKBA	Akebia Therapeutics Comm
AKR	Acadia Realty Trust
AKRO	Akero Therapeutics
AKTS	Akoustis Technologies
AKUS	Akouos
AKYA	Akoya Biosciences
AL	Air Lease
ALB	Albemarle
ALBO	Albireo Pharma
ALC	Alcon
ALCO	Alico
ALDX	Aldeyra Therapeutics
ALE	ALLETE
ALEC	Alector
ALEX	Alexander & Baldwin
ALF	ALFI
ALG	Alamo
ALGM	ALLEGRO MICROSYSTEMS
ALGN	Align Technology
ALGS	Aligos Therapeutics
ALGT	Allegiant Travel
ALHC	Alignment Healthcare
ALIT	Alight
ALK	Alaska Air
ALKS	Alkermes
ALKT	Alkami Technology
ALL	Allstate
ALLE	Allegion
ALLK	Allakos
ALLO	Allogene Therapeutics
ALLY	Ally Financial
ALNY	Alnylam Pharmaceuticals
ALRM	Alarm.com
ALRS	Alerus Finl
ALSN	Allison Transmission
ALTG	Alta Equipment
ALTO	Alto Ingredients
ALTR	ALTAIR ENGINEERING
ALTS	ProShares Morningstar Alternatives Solution ETF
ALV	Autoliv
ALVR	AlloVir
ALX	Alexander's
ALXO	ALX Oncology
ALZN	Alzamend Neuro
AM	Antero Midstream
AMAL	Amalgamated Bank
AMAM	Ambrx Biopharma
AMAT	Applied Materials
AMBA	Ambarella
AMBC	Ambac Financial
AMC	Amc Entertainment
AMCR	AMCOR
AMCX	Amc Networks
AMD	Advanced Micro Devices
AME	AMETEK
AMED	Amedisys
AMEH	Apollo Medical
AMG	Affiliated Managers
AMGN	Amgen
AMH	American Homes 4 Rent
AMJ	JPMORGAN CHASE CAPITAL XVI JP M
AMK	Assetmark Financial
AMKR	Amkor Technology
AMLP	ALERIAN MLP ETF
AMN	AMN Healthcare Services
AMNB	American National
AMOT	Allied Motion
AMP	Ameriprise Financial
AMPH	Amphastar Pharmaceuticals
AMPL	Amplitude
AMR	Alpha Metallurgical Resources
AMRC	Ameresco
AMRK	A-mark Precious Metals c
AMRN	Amarin
AMRS	Amyris
AMRX	Amneal Pharmaceuticals
AMSC	American Superconductor or
AMSF	Amerisafe
AMST	Amesite
AMSWA	American Software
AMT	American Tower
AMTB	Amerant Bancorp
AMTI	Applied Molecular Transport
AMWD	American Woodmark
AMWL	Amwell Health
AMZN	Amazon.com
AN	AUTONATION
ANAB	Anaptysbio
ANAT	American National
ANDE	Andersons
ANEB	Anebulo Pharmaceuticals
ANET	Arista Networks
ANF	ABERCROMBIE & FITCH
ANGI	ANGI Homeservices
ANGL	VanEck Vectors Fallen Angel High Yield Bond ETF
ANGN	Angion Biomedica
ANGO	Angiodynamics
ANIK	Anika Therapeutics
ANIP	Ani Pharmaceuticals
ANNX	Annexon Biosciences
ANPC	AnPac Bio-Medical Science
ANSS	ANSYS
ANTM	Anthem
ANVS	Annovis Bio
ANY	Sphere 3D . Common Shares
AOA	iShares Core Aggressive Allocation ETF
AOK	iShares Core Conservative Allocation ETF
AOM	iShares Moderate Allocation ETF
AOMR	Angel Oak Mortgage
AON	Aon
AOR	iShares Core Growth Allocation ETF
AOS	A.O. Smith
AOSL	Alpha & Omega
AOUT	American Outdoor Brands
APA	Apache
APAM	Artisan Partners Asset Management
APD	Air Products & Chemicals
APEI	American Public Education
APH	Amphenol
API	Agora
APLE	Apple Hospitality REIT
APLS	Apellis Pharmaceuticals
APLT	Applied Therapeutics
APO	Apollo Global Management
APOG	Apogee Enterprises
APP	Applovin
APPF	Appfolio
APPH	AppHarvest
APPN	Appian
APPS	Digital Turbine
APR	Apria
APRE	Aprea Therapeutics
APT	Alpha Pro Tech
APTS	Preferred Apartment
APTV	Aptiv Plc
APTX	Aptinyx
APVO	Aptevo Therapeutics
APYX	Apyx Medical
AQB	AquaBounty Technologies
AQN	Algonquin Power & Utilities
AQST	Aquestive Therapeutics
AQUA	Evoqua Water Technologies
AR	Antero Resources
ARAV	Aravive
ARAY	Accuray
ARB	AltShares Merger Arbitrage ETF
ARCB	Arcbest o
ARCC	Ares Capital
ARCE	Arco Platform
ARCH	Arch Resources
ARCO	Arcos Dorados s
ARCT	Arcturus Therapeutics
ARDX	Ardelyx
ARE	Alexandria Real Estate Equities
ARES	Ares Management
ARGO	Argo International
ARGT	Global X MSCI Argentina
ARGX	argenx SE
ARHS	Arhaus
ARI	Apollo Commercial
ARKF	ARK Fintech Innovation ETF
ARKG	ARK Genomic Revolution ETF
ARKK	ARK Innovation ETF
ARKO	Arko
ARKQ	ARK Industrial Innovation ETF
ARKW	ARK Next Generation Internet ETF
ARL	American Realty Investors
ARLO	Arlo Technologies
ARLP	Alliance Resource Partners
ARMK	Aramark
ARMR	Armor US Equity Index ETF
ARNA	Arena Pharmaceuticals
ARNC	Arconic
AROC	Archrock
AROW	Arrow Financial
ARQT	Arcutis Biotherapeutics
ARR	Armour Residential Reit
ARRY	Array Technologies
ARTNA	Artesian Resource
ARVL	Arrival Ordinary Shares
ARVN	Arvinas
ARW	Arrow Electronics
ARWR	Arrowhead Pharmaceuticals
ASAN	Asana
ASB	Associated Banc-Corp
ASC	Ardmore Shipping
ASGN	ASGN
ASH	ASHLAND GLOBAL HOLDINGS
ASHR	Xtrackers Harvest CSI 300 China A-Shares
ASHX	Xtrackers MSCI China A lusion Equity ETF
ASIX	Advansix
ASLE	AerSale
ASMB	Assembly Biosciences Com
ASML	ASML
ASO	AmSouth Ban
ASPN	Aspen Aerogels
ASPS	Altisource Portfolio
ASPU	Aspen
ASTE	Astec Industries
ASTL	Algoma Steel Common Shares
ASTR	Astra Space
ASTS	AST SpaceMobile
ASUR	Asure Software
ASX	ASE Technology
ATAI	ATAI Life Sciences
ATC	Atotech
ATCO	Atlas
ATEC	Alphatec
ATEN	A10 Networks
ATEX	Anterix
ATGE	DeVry
ATHA	Athira Pharma
ATHM	Autohome
ATHX	Athersys
ATI	ALLEGHENY TECHNOLOGIES INCORPORATED
ATIP	ATI Physical Therapy
ATKR	Atkore Intl
ATLC	Atlanticus
ATLO	Ames National
ATNI	Atn International
ATNX	Athenex
ATO	Atmos Energy
ATOM	Atomera
ATOS	Atossa Therapeutics
ATR	Aptargroup
ATRA	Atara Biotherapeutics
ATRC	Atricure
ATRI	Atrion
ATRO	Astronics
ATRS	Antares Pharma
ATSG	Air Transport Srvcs
ATUS	Altice USA
ATVI	Activision Blizzard
ATXI	Avenue Therapeutics
ATY	AcuityAds s
AU	AngloGold Ashanti
AUB	Atlantic Union Bankshares
AUBN	Auburn National Bancorp
AUDC	AudioCodes
AUPH	Aurinia Pharmaceuticals
AUR	Aurora Innovation
AUUD	Auddia
AUVI	Applied UV
AUY	Yamana Gold
AVA	Avista
AVAH	Aveanna Healthcare
AVAL	Grupo Aval Acciones y Valores S.A.
AVAV	Aerovironment
AVB	AvalonBay Communities
AVCO	Avalon Globocare .
AVD	American Vanguard
AVDX	AvidXchange
AVEM	Avantis Emerging Markets Equity ETF
AVEO	Aveo Pharmaceuticals
AVGO	Broadcom
AVID	Avid Technology
AVIR	Atea Pharmaceuticals
AVLR	Avalara
AVNS	Avanos Medical
AVNT	Avient
AVO	Mission Produce
AVPT	AvePoint
AVRO	Avrobio
AVT	Avnet
AVTE	Aerovate Therapeutics
AVTR	Avantor
AVUV	Avantis U.S. Small Cap Value ETF
AVXL	Anavex Life Sciences .
AVY	Avery Dennison
AVYA	Avaya
AWAY	ETFMG Travel Tech ETF
AWH	Aspira Women's Health
AWI	Armstrong World Industries
AWK	American Water Works
AWR	American States Water
AX	Axos Financial
AXDX	Accelerate Diagnostics c
AXGN	Axogen
AXL	American Axle & Mfg
AXLA	Axcella Health
AXNX	Axonics Modulation Technologies
AXON	Axon Enterprise
AXP	American Express Co
AXS	AXIS Capital
AXSM	Axsome Therapeutics
AXTA	Axalta Coating Systems
AXTI	Axt
AY	Atlantica Sustainable Infrastructure
AYI	ACUITY BRANDS
AYLA	Ayala Pharmaceuticals
AYTU	Aytu Bioscience
AYX	Alteryx
AZEK	AZEK
AZO	AutoZone
AZPN	Aspen Technology
AZRE	Azure Power Global
AZYO	Aziyo Biologics
AZZ	Azz
B	Barnes
BA	Boeing
BAB	Invesco Taxable Municipal Bond ETF
BAC	Bank of America
BAH	Booz Allen Hamilton
BALY	Bally's
BAM	Brookfield Asset Management
BAMR	BAM Reinsurance
BANC	Banc Of California Commo
BAND	Bandwidth
BANF	Bancfirst
BANR	Banner
BAOS	Baosheng Media
BAP	Credicorp
BAPR	Innovator S&P 500 Buffer ETF
BAR	GraniteShares Gold Trust
BARK	The Original BARK
BASE	Couchbase
BATRA	Liberty Media Series A Lbty Braves
BATRK	Liberty Media Series C Lbty Braves
BAX	Baxter International
BB	BlackBerry
BBAX	JPMorgan BetaBuilders Developed Asia ex-Japan ETF
BBBY	BED BATH & BEYOND
BBCA	JPMorgan BetaBuilders Canada ETF
BBCP	Concrete Pumping
BBD	Banco Bradesco S.A.
BBDC	Barings BDC
BBEU	JPMorgan BetaBuilders Europe ETF
BBH	VanEck Vectors Biotech ETF
BBIO	BridgeBio Pharma
BBJP	JPMorgan BetaBuilders Japan ETF
BBMC	JPMorgan BetaBuilders U.S. Mid Cap Equity ETF
BBRE	JPMorgan BetaBuilders MSCI US REIT ETF
BBSI	Barrett Business Services
BBU	Brookfield Business Partners
BBVA	Banco Bilbao Vizcaya Argentaria S.A.
BBWI	Bath & Body Works
BBY	Best Buy
BC	BRUNSWICK CORPORATION
BCAB	BioAtla
BCAT	Blackrock Capital Allocation Trust
BCBP	Bcb Bancorp
BCC	Boise Cascade
BCE	BCE
BCEL	Atreca
BCI	Aberdeen Standard Bloomberg All Commodity Strategy K-1 Free ETF
BCLI	Brainstorm Cell Therapeutics i
BCML	Baycom
BCO	The Brink's
BCOR	Blucora
BCOV	Brightcove
BCPC	Balchem
BCRX	Biocryst Pharmaceuticals
BCS	Barclays
BCTX	Briacell Therapeutics . Common Shares
BDC	Belden
BDN	Brandywine Realty Trust
BDSI	Biodelivery Sciences
BDSX	BIODESIX INC
BDTX	Black Diamond Therapeutics
BDX	Becton Dickinson
BE	Bloom Energy
BEAM	Beam Therapeutics
BECN	Beacon Roofing Supply
BEKE	KE
BELFB	Bel Fuse
BEN	Franklin Resources
BEP	Brookfield Renewable Partners L.P.
BEPC	Brookfield Renewable
BERY	Berry Global
BEST	BEST
BETZ	Roundhill Sports Betting & iGaming ETF
BF.A	Brown-forman
BF.B	Brown-Forman
BFAM	Bright Horizons Family Solutions
BFC	Bank First
BFEB	Innovator S&P 500 Buffer ETF February Series
BFIN	Bankfinancial
BFS	Saul Centers
BFST	Business First Bancshares
BG	Bunge
BGCP	Bgc Partners
BGFV	Big 5 Sporting Goods
BGS	B&g Foods
BGSF	Bg Staffing
BH	Biglari
BH.A	Biglari
BHB	Bar Harbor Bankshares
BHC	Bausch Health Companies
BHE	Benchmark Electronics
BHG	Bright Health
BHLB	Berkshire Hills Bancorp
BHSE	Bull Horn
BHVN	Biohaven Pharmaceutical
BIDU	Bidu
BIG	BIG LOTS
BIGC	BigCommerce
BIIB	Biogen
BIL	SPDR LEHMAN 1-3 MONTH T-BILL ETF
BILI	Bilibili
BILL	Bill.com
BIO	Bio-Rad Laboratories
BIOX	Bioceres Crop Solutions
BIP	Brookfield Infrastructure Partners L.P.
BIPC	Brookfield Infrastructure
BIRD	Allbirds
BITF	Bitfarms
BITO	ProShares Bitcoin Strategy ETF
BIV	VANGUARD IT BOND ETF
BIZD	VanEck Vectors BDC ome ETF
BJ	BJ's Wholesale Club
BJAN	Innovator S&P 500 Buffer ETF
BJDX	Bluejay Diagnostics
BJK	VanEck Vectors Gaming ETF
BJRI	Bj's Restaurants
BJUL	Innovator S&P 500 Buffer ETF
BK	The Bank of New York Mellon
BKCH	Global X Blockchain ETF
BKD	Brookdale Senior Living
BKE	Buckle
BKH	Black Hills
BKI	Black Knight
BKKT	Bakkt
BKLN	POWERSHARES EXCHANGE-TRADED FUND
BKNG	Booking
BKR	Baker Hughes
BKU	BankUnited
BL	BlackLine
BLBD	Perseon
BLBX	Blackboxstocks
BLCN	Reality Shares Nasdaq NexGen Economy ETF
BLCT	BlueCity
BLD	TopBuild
BLDE	Blade Air Mobility
BLDP	Ballard Power Systems
BLDR	Builders FirstSource
BLFS	Biolife Solutions
BLI	Berkeley Lights
BLK	BlackRock
BLKB	Blackbaud
BLL	Ball
BLMN	Bloomin Brands
BLND	Blend Labs
BLNK	Blink Charging
BLOK	Amplify Transformational Data Sharing ETF
BLPH	Bellerophon Therapeutics
BLUE	bluebird bio
BLV	Vanguard Long-Term Bond ETF
BLX	Banco Latinoamericano
BMBL	Bumble
BMEA	Biomea Fusion
BMI	Badger Meter
BMO	Bank of Montreal
BMRC	Bank Of Marin
BMRN	BioMarin Pharmaceutical
BMY	Bristol-Myers Squibb
BND	VANGUARD TOTAL BOND MARKET ETF
BNDX	Vanguard Total International Bond ETF
BNFT	Benefitfocus
BNGO	Bionano Genomics
BNL	Broadstone Net Lease
BNO	United States Brent Oil Fund
BNR	Burning Rock Biotech
BNS	The Bank of Nova Scotia
BNTX	BioNTech SE
BOCT	Innovator S&P 500 Buffer ETF
BODY	The Beachbody
BOH	Bank of Hawaii
BOKF	Bok Financial
BOLT	Bolt Biotherapeutics
BON	Bon Natural Life
BOND	PIMCO Total Return ETF
BOOM	Dmc Global
BOOT	Boot Barn
BOTZ	Global X Funds Global X Robotics & Artificial Intelligence ETF
BOX	Box
BPMC	Blueprint Medicines
BPMP	BP Midstream Partners LP
BPOP	Popular
BPRN	The Bank Of Princeton
BPTS	Biophytis
BQ	Boqii
BR	Broadridge Financial Solutions
BRBR	Bellring Brands
BRC	Brady
BRFS	BRF S.A.
BRG	Bluerock Residential Growth
BRID	Bridgford Foods
BRK.A	Berkshire Hathaway
BRK.B	BERKSHIRE HATHAWAY
BRKL	Brookline Bancorp
BRKR	Bruker
BRMK	Broadmark Realty Capital
BRO	Brown & Brown
BROS	Dutch Bros
BRP	Brp
BRT	Brt Apartments
BRX	Brixmor Property
BRY	Berry
BRZU	Direxion Daily Brazil Bull 2X Shares
BSAC	Banco Santander-Chile
BSBK	Bogota Financial
BSBR	Banco Santander (Brasil) S.A.
BSCM	Invesco BulletShares 2022 orate Bond ETF
BSCN	Invesco BulletShares 2023 orate Bond ETF
BSCO	Invesco BulletShares 2024 orate Bond ETF
BSGM	Biosig Technologies
BSIG	Brightsphere Investment
BSJM	Invesco BulletShares 2022 High Yield orate Bond ETF
BSJN	Invesco BulletShares 2023 High Yield orate Bond ETF
BSJO	Invesco BulletShares 2024 High Yield orate Bond ETF
BSM	Black Stone Minerals
BSMX	Banco Santander Mexico S.A. Institucion de Banca Multiple Grupo Financiero Santander Mexico
BSRR	Sierra
BSV	VANGUARD ST BOND ETF
BSVN	Bank7 .
BSX	Boston Scientific
BSY	BENTLEY SYSTEMS INC
BTAI	Bioxcel Therapeutics
BTAL	AGFiQ US Market Neutral Anti-Beta Fund
BTB	Bit Brother Ordinary Shares
BTF	Valkyrie Bitcoin Strategy ETF
BTG	B2Gold
BTRS	BTRS s
BTU	PEABODY ENERGY CORPORATION
BTWN	Bridgetown
BTX	Brooklyn ImmunoTherapeutics
BUD	ANHEUSER-BUSCH INBEV SA
BUFF	Laddered Fund of S&P 500 Power Buffer ETFs
BURL	Burlington Stores
BUSE	First Busey
BV	Brightview
BVH	Bluegreen Vacations
BVN	Compania de Minas Buenaventura S.A.A.
BVS	Bioventus
BW	Babcock & Wilcox Enterprises
BWA	BorgWarner
BWB	Bridgewater Bancshares
BWFG	Bankwell Financial
BWMN	Bowman Consulting
BWX	SPDR LEHMAN INTL TREASURY BD
BWXT	BWX Technologies
BX	Blackstone
BXMT	Blackstone Mortgage Trust
BXP	Boston Properties
BY	Byline Bancorp
BYD	Boyd Gaming
BYND	Beyond Meat
BYSI	Beyondspring
BZ	Kanzhun
BZH	Beazer Homes Usa
BZUN	Baozun
C	Citigroup
CAAP	oraciÃ³n AmÃ©rica Airports S.A.
CABA	Cabaletta Bio
CABO	Cable One
CAC	Camden National
CACC	Credit Acceptance
CACI	Caci International
CADE	Cadence Bancorporation
CADL	Candel Therapeutics
CAE	CAE
CAG	Conagra Brands
CAH	Cardinal Health
CAJ	Canon
CAKE	Cheesecake Factory
CAL	Caleres
CALA	Calithera Biosciences Co
CALB	California Bancorp
CALM	Cal-maine Foods
CALT	Calliditas Therapeutics
CALX	Calix
CAMP	Calamp
CANO	Cano Health
CAPL	CrossAmerica Partners LP
CAR	Avis Budget
CARA	Cara Therapeutics Common
CARE	Carter Bankshares
CARG	CarGurus
CARR	Carrier Global
CARS	Cars Com
CARZ	First Trust NASDAQ Global Auto Index Fund
CASA	Casa Systems
CASH	Meta Financial
CASI	Casi Pharmaceuticals
CASS	Cass Information Systems
CASY	Casey's General Stores
CAT	Caterpillar
CATC	Cambridge Bancorp
CATO	Cato
CATY	Cathay General Bancorp
CB	Chubb
CBAN	Colony Bankcorp
CBAY	Cymabay Therapeutics Comm
CBD	Companhia Brasileira de Distribuicao
CBFV	Cb Financial Svc Cmn
CBH	ALLIANZGI CONVERTIBLE & INCOME 2024 TARGET TERM FUND
CBIO	Catalyst Biosciences
CBNK	Capital Bancorp
CBOE	Cboe Global Markets
CBRE	CBRE
CBRL	Cracker Barrel Old Country Store
CBSH	Commerce Bancshares
CBT	Cabot
CBTX	Cbtx
CBU	Community Bank System
CBZ	Cbiz
CC	CHEMOURS (THE)
CCB	Coastal Financial
CCBG	Capital City Bank
CCCC	C4 Therapeutics
CCCS	CCC Intelligent Solutions s
CCEP	Coca-Cola European Partners
CCF	Chase
CCI	Crown Castle International
CCJ	Cameco
CCK	CROWN HOLDINGS
CCL	Carnival
CCMP	Cmc Materials
CCNE	Cnb Financial
CCO	Clear Channel Outdoor s
CCOI	Cogent Communications
CCRN	Cross Country Healthcare
CCS	Century Communities
CCSI	Consensus Cloud Solutions
CCU	COMPANIA CERVECERIAS UNIDAS S.A.
CCXI	ChemoCentryx
CD	Chindata
CDAK	Codiak BioSciences
CDAY	Ceridian HCM
CDE	Couer Mining
CDEV	Centennial Resource Development
CDK	CDK Global
CDLX	Cardlytics
CDMO	Avid Bioservices
CDNA	Caredx
CDNS	Cadence Design Systems
CDTX	Cidara Therapeutics
CDW	CDW CORPORATION
CDXC	Chromadex
CDXS	Codexis
CDZI	Cadiz
CE	Celanese
CECE	Ceco Environmental
CEF	Sprott Physical Gold and Silver Trust
CEIX	Consol Energy
CELH	Celsius
CEMI	Chembio Diagnostics
CENT	Central Garden & Pet
CENTA	Central Garden & Pet
CENX	Century Aluminum
CEQP	Crestwood Equity Partners LP
CERE	Cerevel Therapeutics s
CERN	Cerner
CERS	Cerus
CERT	Certara
CEVA	Ceva
CF	CF Industries
CFB	Crossfirst Bankshares
CFFI	C&f Financial
CFFN	Capitol Federal Financial
CFG	Citizens Financial
CFLT	Confluent
CFR	Cullen/Frost Bankers
CFRX	Contrafect Common
CFX	Colfax
CG	The Carlyle
CGBD	TCG BDC
CGC	Canopy Growth
CGEM	Cullinan Oncology
CGNT	Cognyte Software
CGNX	Cognex
CGW	Invesco S&P Global Water Index ETF
CHCO	City
CHCT	Community Healthcare Trust
CHD	Church & Dwight
CHDN	Churchill Downs
CHE	Chemed
CHEF	Chefs Warehouse
CHGG	Chegg
CHH	Choice Hotels International
CHIK	Global X MSCI China Information Technology ETF
CHIQ	Global X MSCI China Consumer Discretionary ETF
CHIR	Chiron
CHIS	Global X MSCI Consumer Staples ETF
CHK	CHESAPEAKE ENERGY CORPORATION
CHKP	Check Point Software Technologies
CHMG	Chemung Financial
CHMI	Cherry Hill Mortgage Investmen
CHNG	Change Healthcare
CHPT	ChargePoint s
CHRS	Coherus Biosciences Comm
CHRW	C. H. Robinson Worldwide
CHS	Chico's Fas
CHTR	Charter Communications
CHUY	Chuy's
CHWY	Chewy
CHX	ChampionX
CI	CIGNA
CIA	Citizens
CIBR	First Trust NASDAQ CEA Cybersecurity ETF
CIEN	CIENA CORPORATION
CIG	Companhia Energetica de Minas Gerais
CIGI	Colliers International
CIM	Chimera Investment
CINF	Cincinnati Financial
CINT	CI&T
CIO	City Office Reit
CIR	Circor International
CIVB	Civista Bancshares
CIVI	Civitas Solutions
CIX	Compx International
CIXX	CI Financial
CIZN	Citizens
CKPT	Checkpoint Therapeutics
CL	Colgate-Palmolive
CLAR	Clarus
CLAS	Class Acceleration
CLB	Core Laboratories N.V.
CLBK	Columbia Financial
CLDT	Chatham Lodging
CLDX	Celldex Therapeutics
CLEU	China Liberal Education
CLF	CLEVELAND-CLIFFS
CLFD	Clearfield
CLH	Clean Harbors
CLIX	ProShares Long Online/Short Stores ETF
CLMT	Calumet Specialty Products Partners
CLNE	Clean Energy Fuels
CLOU	Global X Cloud Computing ETF
CLOV	Clover Health Investments
CLPR	Clipper Realty
CLR	Continental Resources
CLS	Celestica
CLSK	CleanSpark
CLVS	Clovis Oncology
CLVT	Clarivate Plc
CLW	Clearwater
CLX	The Clorox
CLXT	Calyxt
CM	Canadian Imperial Bank of Commerce
CMA	Comerica
CMBM	Cambium Networks
CMC	Commercial Metals
CMCL	Caledonia Mining
CMCO	Columbus Mckinnon
CMCSA	Comcast
CMCT	Cim Commercial Trust
CME	CME
CMF	iShares California Muni Bond ETF
CMG	Chipotle Mexican Grill
CMI	Cummins
CMP	Compass Minerals
CMPI	Checkmate Pharmaceuticals
CMPR	Cimpress
CMPS	COMPASS Pathways
CMRE	Costamare
CMRX	Chimerix
CMS	CMS Energy
CMTL	Comtech Telecommunications
CNA	CNA Financial
CNC	Centene
CNCE	Concert Pharmaceuticals
CNDT	Conduent
CNEY	CN Energy
CNHI	CNH Industrial N.V.
CNI	Canadian National Railway
CNK	Cinemark
CNM	Core & Main
CNMD	Conmed
CNNE	Cannae
CNOB	Connectone Bancorp
CNP	CenterPoint Energy
CNQ	Canadian Natural Resources
CNR	Cornerstone Building Brands
CNRG	SPDR S&P Kensho Clean Power ETF
CNS	Cohen & Steers
CNSL	Consolidated Comm
CNTA	Centessa Pharmaceuticals
CNTB	Connect Biopharma
CNTG	Centogene Bv
CNTY	Century Casinos
CNVY	Convey Parent
CNX	CNX RESOURCES CORPORATION
CNXC	Concentrix
CNXN	Pc Connection
CODI	Compass Diversified
CODX	Co-diagnostics
COF	Capital One Financial
COFS	Choiceone Financial
COHR	Coherent
COHU	Cohu
COIN	Coinbase
COKE	Coca-cola Bottling
COLB	Columbia Banking System
COLD	Americold Realty Trust
COLI	Colicity
COLL	Collegium Pharmaceutical
COLM	Columbia Sportswear
COMM	CommScope
COMP	Compass
COMT	iShares Commodities Select Strategy ETF
CONE	CyrusOne
CONN	Conn's
CONX	CONX .
COO	The Cooper Companies
COOK	Traeger
COOP	Mr. Cooper
COP	ConocoPhillips
COPX	Global X Copper Miners ETF
CORR	Corenergy Infrastructure
CORS	Corsair Partnering
CORT	Corcept Therapeutics
COST	Costco Wholesale
COTY	Coty
COUP	Coupa Software
COUR	Coursera
COWN	Cowen
CP	Canadian Pacific Railway
CPA	Copa S.A.
CPB	Campbell Soup
CPE	Callon Petroleum
CPF	Central Pacific Financial
CPG	Crescent Point Energy
CPI	IQ Real Return ETF
CPK	Chesapeake Utilities
CPLG	Corepoint Lodging
CPNG	Coupang
CPOP	Pop Culture
CPRI	Capri
CPRT	Copart
CPRX	Catalyst Pharmaceutical
CPS	Cooper Standard
CPSI	Computer Programs And Sys
CPT	Camden Property Trust
CQP	Cheniere Energy Partners L.P.
CQQQ	Invesco China Technology ETF
CR	Crane
CRAI	Cra International
CRBP	Corbus Pharmaceuticals
CRBU	Caribou Biosciences
CRC	California Resources
CRCT	Cricut
CRD.A	Crawford &
CRH	CRH
CRHC	Cohn Robbins
CRI	Carter's
CRIS	Curis
CRK	Comstock Resources
CRL	Charles River Laboratories International
CRM	Salesforce.com
CRMD	Cormedix
CRMT	America's Car-mart
CRNC	Cerence
CRNX	Crinetics Pharmaceuticals
CRON	Cronos
CRS	Carpenter Technology
CRSP	CRISPR Therapeutics AG
CRSR	Corsair Gaming
CRTX	Cortexyme
CRUS	Cirrus Logic
CRVL	Corvel
CRWD	CrowdStrike
CS	Credit Suisse AG
CSBR	Champions Oncology
CSCO	Cisco Systems
CSGP	Costar
CSGS	Csg Systems International
CSII	Cardiovascular Systems
CSIQ	Canadian Solar Common Shares (BC)
CSL	Carlisle Companies
CSML	IQ Chaikin U.S. Small Cap ETF
CSQ	Calamos Strategic Total Return Fund
CSR	Centerspace
CSTE	Caesarstone
CSTL	Castle Biosciences
CSTM	Constellium SE
CSTR	Capstar Financial
CSV	Carriage Services
CSWC	Capital Southwest
CSWI	Csw Industrials
CSX	CSX
CTAS	Cintas
CTBI	Community Trust Bancorp
CTKB	Cytek Biosciences
CTLP	Cantaloupe
CTLT	Catalent
CTMX	Cytomx Therapeutics
CTO	Cto Realty Growth
CTOS	Custom Truck One Source
CTRA	Coterra Energy
CTRE	Caretrust Reit
CTRN	Citi Trends
CTS	Cts
CTSH	Cognizant Technology Solutions
CTSO	Cytosorbents
CTT	Catchmark Timber Trust
CTVA	CORTEVA
CTXS	Citrix Systems
CUBE	CubeSmart
CUBI	Customers Bancorp Common
CUE	Cue Biopharma
CUK	Carnival & Plc
CURI	CuriosityStream
CURO	Curo
CURV	Torrid
CUTR	Cutera
CUZ	Cousins Properties
CVAC	CureVac N.V.
CVBF	CVB Financial
CVCO	Cavco Industries
CVCY	Central Valley Comm Bank
CVE	Cenovus Energy
CVET	Covetrus
CVGW	Calavo Growers
CVI	CVR Energy
CVLT	Commvault Systems
CVLY	Codorus Valley Bancorp
CVM	Cel-sci
CVNA	Carvana
CVRX	CVRx
CVS	CVS Health
CVX	Chevron
CW	Curtiss-Wright
CWB	SPDR Bloomberg Barclays Convertible Securities
CWBR	Cohbar
CWCO	Consolidated Water
CWEB	Direxion Daily CSI China Internet Index Bull 2X Shares
CWEN	Clearway Energy
CWEN.A	Clearway Energy
CWI	SPDR MSCI ACWI ex-US ETF
CWK	Cushman & Wakefield
CWST	Casella Waste Systems
CWT	California Water Service
CX	CEMEX S.A.B. de C.V.
CXM	Sprinklr
CXSE	WisdomTree China ex-State-Owned Enterprises Fund
CXW	Corecivic
CYBE	Cyberoptics
CYBR	CyberArk Software
CYCN	Cyclerion Therapeutics
CYH	Community Health Systems
CYRX	Cryoport
CYT	Cyteir Therapeutics
CYTK	Cytokinetics Com
CZNC	Citizens And Northern
CZR	Caesars Entertainment
D	Dominion Energy
DAC	Danaos
DADA	Dada Nexus
DAKT	Daktronics
DAL	Delta Air Lines
DAN	Dana
DAO	Youdao
DAR	Darling Ingredients
DATS	DatChat
DAWN	Day One Biopharmaceuticals
DB	Deutsche Bank Aktiengesellschaft
DBA	POWERSHARES DB AGRICULTURE FUND
DBC	POWERSHARES DB COMMODITY INDEX
DBD	Diebold
DBEF	Xtrackers MSCI EAFE Hedged Equity ETF
DBGI	Digital Brands
DBI	Designer Brands
DBMF	iM DBi Managed Futures Strategy ETF
DBO	Invesco DB Oil Fund
DBRG	DigitalBridge
DBTX	Decibel Therapeutics
DBVT	DBV Technologies S.A.
DBX	Dropbox
DCBO	Docebo
DCGO	DocGo
DCI	Donaldson
DCO	Ducommun
DCOM	Dime Community Bancshares
DCP	DCP Midstream LP
DCPH	Deciphera Pharmaceuticals
DCT	Duck Creek Technologies
DD	DUPONT DE NEMOURS
DDD	3d Systems
DDL	Dingdong
DDM	PROSHARES ULTRA DOW30
DDOG	Datadog
DDS	DILLARD'S
DDWM	WisdomTree Dynamic Currency Hedged International Equity Fund
DE	Deere &
DEA	Easterly Government Properties
DECK	Deckers Outdoor
DEI	Douglas Emmett
DELL	DELL TECHNOLOGIES
DEM	WISDOMTREE EMERGING MARKETS HIG
DEN	Denbury
DENN	Denny's
DES	WisdomTree US SmallCap Dividend ETF
DESP	Despegar.com
DFAC	Dimensional U.S. Core Equity 2 ETF
DFEN	Direxion Daily Aerospace & Defense Bull 3X Shares
DFH	Dream Finders Homes
DFIN	Donnelley Financial Solutions
DFND	Reality Shares DIVCON Dividend Defender ETF
DFNL	Davis Select Financial ETF
DFS	Discover Financial Services
DG	Dollar General
DGICA	Donegal
DGII	Digi International
DGRO	iShares Core Dividend Growth ETF
DGRS	WisdomTree U.S. SmallCap Dividend Growth Fund
DGRW	WisdomTree U.S. Dividend Growth Fund
DGS	WisdomTree Emerging Markets SmallCap Divdend Fund
DGX	Quest Diagnostics
DH	Definitive Healthcare . Class A
DHC	Diversified Healthcare Trust Common Shares Of Bene
DHHC	DiamondHead
DHI	D. R. Horton
DHIL	Diamond Hill Investment
DHR	Danaher
DHT	Dht
DHX	Dhi
DIA	SPDR DOW JONES INDUSTRIAL AVERAGE
DIBS	1stdibs.com
DIDI	Xiaoju Kuaizhi
DIN	Dine Brands Global
DIOD	Diodes
DIS	The Walt Disney
DISCA	Discovery Class A
DISCB	Discovery
DISCK	Discovery Class C
DISH	Dish Network
DJCO	Daily Journal
DJP	IPATH DOW JONES UBS COMMODITY
DK	Delek Us
DKL	Delek Logistics Partners
DKNG	DraftKings
DKS	DICK'S Sporting Goods
DLB	Dolby Laboratories
DLN	WISDOMTREE LARGECAP DIVIDEND FUND
DLO	DLocal
DLR	Digital Realty Trust
DLS	WisdomTree International SmallCap Dividend
DLTH	Duluth
DLTR	Dollar Tree
DLX	Deluxe
DM	Desktop Metal
DMLP	Dorchester Minerals
DMRC	Digimarc
DMRL	DeltaShares S&P 500 Managed Risk ETF
DMTK	Dermtech
DNA	Ginkgo Bioworks
DNAY	Codex DNA
DNB	DUN & BRADSTREET CORPORATION (T
DNL	WisdomTree Global ex-U.S. Quality Dividend Growth Fund
DNLI	Denali Therapeutics
DNMR	Danimer Scientific
DNN	Denison Mines
DNOW	Now
DNP	DNP Select ome Fund
DNUT	Krispy Kreme
DOC	Physicians Realty Trust
DOCN	DigitalOcean
DOCS	Doximity
DOCU	DocuSign
DOG	ProShares Short Dow30
DOLE	Dole
DOMA	Doma s
DOMO	Domo
DON	WisdomTree US MidCap Dividend ETF
DOOO	BRP
DOOR	Masonite International
DORM	Dorman Products
DOV	Dover
DOW	DOW
DOX	Amdocs
DOYU	DouYu International
DPST	Direxion Daily Regional Banks Bull 3X Shares
DPZ	Domino's Pizza
DRE	Duke Realty
DRH	Diamondrock Hospitality
DRI	Darden Restaurants
DRIP	Direxion Daily S&P Oil & Gas Exp. & Prod. Bear 2X Shares
DRIV	Global X Autonomous & Electric Vehicles ETF
DRMA	Dermata Therapeutics
DRQ	Dril-quip
DRRX	Durect
DRSK	Aptus Defined Risk ETF
DRVN	Driven Brands
DSEP	FT Cboe Vest U.S. Equity Deep Buffer ETF - September
DSEY	Diversey
DSGN	Design Therapeutics
DSGX	The Descartes Systems
DSI	iShares MSCI KLD 400 Social ETF
DSKE	Daseke
DSP	Viant Technology
DT	Dynatrace
DTC	Solo Brands
DTE	DTE Energy
DTEC	ALPS Disruptive Technologies ETF
DTIL	Precision Biosciences
DTM	DT
DUK	Duke Energy
DUOL	Duolingo
DUST	Direxion Daily Gold Miners Index Bear 2X
DVA	DaVita
DVAX	Dynavax Technologies
DVN	Devon Energy
DVY	ISHARES DOW JONES SELECT
DVYE	iShares Emerging Markets Dividend ETF
DWAS	Invesco DWA SmallCap Momentum ETF
DWM	WisdomTree DEFA
DX	Dynex Capital
DXC	DXC Technology
DXCM	Dexcom
DXD	PROSHARES ULTRASHORT DOW30
DXJ	WISDOMTREE JAPAN TOTAL DIVIDEND
DXPE	Dxp Enterprises
DY	Dycom Industries
DYAI	Dyadic International
DYN	Dyne Therapeutics
DYNS	Dynamics Special Purpose
DZSI	Dzs
E	Eni S.p.A.
EA	Electronic Arts
EAR	Eargo
EARN	Ellington Residential Mortgage
EAT	Brinker International
EB	Eventbrite
EBAY	eBay
EBC	Eastern Bankshares
EBET	Esports Technologies
EBF	Ennis
EBIX	Ebix
EBIZ	Global X E-commerce ETF
EBMT	Eagle Bancorp Montana
EBND	SPDR Barclays Capital Emerging Markets Local Bond ETF
EBON	Ebang International Class A Ordinary Shares
EBR	Centrais Eletricas Brasileiras S.A. - Eletrobras
EBS	Emergent BioSolutions
EBTC	Enterprise Bancorp
EC	Ecopetrol S.A.
ECH	iShares MSCI Chile ETF
ECL	Ecolab
ECOL	Us Ecology
ECOM	Channeladvisor
ECON	Columbia Emerging Markets Consumer ETF
ECPG	Encore Capital
ECVT	Ecovyst
ED	Consolidated Edison
EDC	DIREXION EMERGING MARKETS BULL
EDIT	Editas Medicine
EDOC	Global X Telemedicine & Digital Health ETF
EDR	Endeavor
EDRY	EurDry drybulk
EDTK	Skillful Craftsman
EDU	New Oriental Education & Technology
EDV	Vanguard Ext Duration Treasury ETF
EEFT	Euronet Worldwide
EEIQ	Elite Education
EEM	ISHARES MSCI EMERGING MARKETS INDEX
EEMA	iShares MSCI Emerging Markets Asia ETF
EEMV	iShares Edge MSCI Min Vol Emerging Markets ETF
EEX	Emerald
EFA	ISHARES MSCI EAFE INDEX
EFAV	iShares Edge MSCI Min Vol EAFE ETF
EFC	Ellington Financial
EFG	iShares MSCI EAFE Growth ETF
EFSC	Enterprise Financial Svcs
EFTR	eFFECTOR Therapeutics
EFV	ISHARES MSCI VALUE INDEX FUND
EFX	Equifax
EGAN	Egain
EGBN	Eagle Bancorp
EGHT	8 X 8
EGLE	Eagle Bulk Shipping Commo
EGO	Eldorado Gold
EGP	EastGroup Properties
EGRX	Eagle Pharmaceuticals Co
EHC	Encompass Health
EHTH	eHealth
EIDO	iShares MSCI Indonesia ETF
EIG	Employers
EIGR	Eiger Biopharmaceuticals
EIX	Edison Int'l
EL	Estee Lauder Cos.
ELA	Envela
ELAN	Elanco Animal Health
ELEV	Elevation Oncology
ELF	E.L.F. Beauty
ELMD	Electromed
ELMS	Electric Last Mile Solutions
ELOX	Eloxx Pharmaceuticals
ELP	Companhia Paranaense de Energia - COPEL
ELS	Equity LifeStyle Properties
ELY	Callaway Golf
ELYM	Eliem Therapeutics
EM	Smart Share Global
EMB	ISHARES JP MORGAN EM BOND FD
EME	EMCOR
EMGF	iShares Edge MSCI Multifactor Emerging Markets ETF
EML	Eastern
EMLC	MARKET VECTORS EMERGING MARKETS
EMLP	First Trust North American Energy Infrastructure Fund
EMN	Eastman Chemical
EMP	Entergy Mississippi
EMQQ	Emerging Markets Internet & Ecommerce ETF
EMR	Emerson Electric
EMTY	ProShares Decline of the Retail Store ETF
EMXC	iShares MSCI Emerging Markets ex China ETF
ENB	Enbridge
ENDP	ENDO INTERNATIONAL
ENIA	Enel Americas S.A.
ENIC	Enel Chile S.A.
ENLC	EnLink Midstream
ENOB	Enochian Biosciences Common Shares
ENPC	Executive Network Partnering
ENPH	Enphase Energy
ENR	Energizer
ENS	EnerSys
ENSG	Ensign
ENTA	Enanta Pharmaceuticals c
ENTG	Entegris
ENV	Envestnet
ENVA	Enova International
ENVX	Enovix
ENZ	Enzo Biochem
EOG	EOG Resources
EOLS	Evolus
EOSE	Eos Energy Enterprises
EPAC	Enerpac Tool
EPAM	EPAM Systems
EPAY	Bottomline Technologies
EPC	Edgewell Personal Care Co
EPD	Enterprise Products Partners L.P.
EPI	WISDOMTREE INDIA EARNINGS FUND
EPIX	ESSA Pharma
EPM	Evolution Petroleum
EPOL	iShares MSCI Poland ETF
EPP	ISHARES MSCI PACIFIC EX-JAPAN I
EPR	EPR Properties
EPRF	Innovator S&P High Quality Preferred ETF
EPRT	Essential Properties Realty Trust
EPU	iShares MSCI Peru ETF
EPZM	Epizyme
EQ	EQUILLIUM
EQBK	Equity Bancshares
EQC	Equity Commonwealth
EQH	Equitable
EQIX	Equinix
EQOS	EQONEX LIMITED Ordinary Shares
EQR	Equity Residential
EQT	EQT CORPORATION
EQX	Equinox Gold
ERAS	Erasca
ERF	Enerplus
ERIC	Telefonaktiebolaget LM Ericsson (publ)
ERIE	Erie Indemnity
ERII	Energy Recovery
ERUS	iShares MSCI Russia ETF
ERX	DIREXION DAILY ENERGY BULL 3X SHARES
ERY	Direxion Daily Energy Bear 2X Shares
ES	Eversource Energy
ESBA	Empire State Realty OP
ESCA	Escalade
ESE	Esco Technologies
ESGD	iShares Trust iShares ESG Aware MSCI EAFE ETF
ESGE	iShares ESG MSCI EM ETF
ESGR	Enstar
ESGU	iShares ESG MSCI USA ETF
ESGV	Vanguard ESG U.S. Stock ETF
ESI	Element Solutions
ESLT	Elbit Systems
ESMT	EngageSmart
ESNT	Essent
ESPO	VanEck Vectors Video Gaming and eSports ETF
ESPR	Esperion Therapeutics Co
ESQ	Esquire Financial
ESRT	Empire State Realty Trust
ESS	Essex Property Trust
ESSA	Essa Bancorp
ESTA	Establishment Labs s
ESTC	Elastic N.V.
ESTE	Earthstone Energy
ET	Energy Transfer LP
ETD	Ethan Allen Interiors
ETN	Eaton
ETNB	89bio
ETON	Eton Pharmaceutcials
ETR	Entergy
ETRN	Equitrans Midstream
ETSY	Etsy
ETWO	E2open Parent s
EUFN	iShares MSCI Europe Financials ETF
EUO	PROSHARES ULTRASHORT EURO
EURN	Euronav NV
EVA	Enviva Partners
EVAX	Evaxion Biotech
EVBG	Everbridge
EVBN	Evans Bancorp
EVC	Entravision Communication
EVCM	EverCommerce
EVER	Everquote
EVFM	Evofem Biosciences
EVGO	EVgo Class A
EVH	Evolent Health
EVI	Evi Industries
EVLO	Evelo Biosciences
EVOP	Evo Payments
EVR	Evercore
EVRG	Evergy
EVRI	Everi
EVTC	EVERTEC
EW	Edwards Lifesciences
EWA	ISHARES MSCI AUSTRALIA INDEX
EWBC	East West Bancorp
EWC	ISHARES MSCI CANADA INDEX FUND
EWCO	Invesco S&P 500 Equal Weight Communication Services ETF
EWCZ	European Wax Center
EWG	ISHARES MSCI GERMANY INDEX FUND
EWH	ISHARES MSCI HONG KONG INDEX
EWI	ISHARES MSCI ITALY INDEX FUND
EWJ	ISHARES MSCI JAPAN INDEX FUND
EWL	iShares MSCI Switzerland ETF
EWM	ISHARES MSCI MALAYSIA INDEX FUND
EWP	ISHARES MSCI SPAIN INDEX FUND
EWQ	ISHARES MSCI FRANCE INDEX FUND
EWS	ISHARES MSCI SINGAPORE INDEX
EWT	ISHARES MSCI TAIWAN INDEX FUND
EWTX	Edgewise Therapeutics
EWU	ISHARES MSCI UNITED KINGDOM IND
EWW	ISHARES MSCI MEXICO INDEX FUND
EWX	SPDR S&P Emerging Small Cap ETF
EWY	ISHARES MSCI SOUTH KOREA INDEX
EWZ	ISHARES MSCI BRAZIL INDEX
EXAS	Exact Sciences
EXC	Exelon
EXEL	Exelixis
EXFY	Expensify
EXG	Eaton Vance Tax-Managed Global Diversified Equity ome Fund
EXI	iShares S&P Global Industrials ETF
EXK	Endeavour Silver
EXLS	Exlservice
EXP	Eagle Materials
EXPD	Expeditors
EXPE	Expedia
EXPI	Exp World
EXPO	Exponent
EXPR	Express
EXR	Extra Space Storage
EXTN	Exterran
EXTR	Extreme Networks
EYE	National Vision
EZA	ISHARES MSCI SOUTH AFRICA INDEX
EZGO	EZGO Technologies
EZPW	Ezcorp
EZU	ISHARES MSCI EMU INDEX FUND
F	Ford Motor
FA	First Advantage
FAF	First American Financial
FALN	iShares U.S. Fallen Angels USD Bond ETF
FAMI	Farmmi
FAN	First Trust ISE Global Wind Energy Index Fund
FANG	Diamondback Energy
FARM	Farmer Bros
FARO	Faro Technologies
FAS	DIREXION DAILY FINANCIAL BULL 3X SHARES
FAST	Fastenal Co
FATE	Fate Therapeutics
FAZ	DIREXION DAILY FINANCIAL BEAR 3X SHARES
FB	Facebook
FBC	Flagstar Bancorp
FBHS	Fortune Brands Home & Security
FBIO	Fortress Biotech
FBIZ	First Business Financial
FBK	Fb Financial
FBMS	First Bancshares
FBNC	First Bancorp
FBND	Fidelity Total Bond ETF
FBP	First Bancorp
FBT	First Trust Amex Biotechnology Index
FC	Franklin Covey
FCAP	First Capital
FCBC	First Community
FCCO	First Community
FCEL	Fuelcell Energy
FCF	First Commonwealth
FCFS	FirstCash
FCG	First Trust Natural Gas ETF
FCN	FTI Consulting
FCNCA	First Citizens Bancshares
FCOM	Fidelity MSCI Communication Services Index ETF
FCPT	Four Corners Property Trust
FCX	Freeport-McMoRan
FDBC	Fidelity D & D Bancorp
FDIS	Fidelity MSCI Consumer Discretionary Index ETF
FDL	First Trust Morningstar Dividend Leaders
FDM	First Trust Dow Jones Select MicroCap Index Fund
FDMT	4D Molecular Therapeutics
FDN	First Trust Dow Jones Internet Index
FDNI	First Trust Dow Jones International Internet ETF
FDP	Fresh Del Monte Produce
FDS	Factset Research Systems
FDX	FedEx
FE	FirstEnergy
FELE	Franklin Electric
FEM	First Trust Emerging Markets AlphaDEX Fund
FEMY	Femasys
FENC	Fennec Pharmaceuticals
FENY	Fidelity MSCI Energy Index ETF
FEZ	SPDR DJ EURO STOXX 50 ETF
FF	Futurefuel
FFBC	First Financial Bancorp
FFIC	Flushing Financial
FFIE	Faraday Future Intelligent Electric
FFIN	First Financial Bankshares
FFIV	F5 Networks
FFWM	First Foundation
FGBI	First Guaranty Bancshares
FGEN	FibroGen
FHB	First Hawaiian
FHI	Federated Hermes
FHLC	Fidelity MSCI Health Care Index ETF
FHN	FIRST HORIZON NATIONAL CORPORATION
FHS	First High-School Education
FHTX	Foghorn Therapeutics
FIBK	First Interstate Banc
FICO	Fair Isaac
FIDU	Fidelity MSCI Industrials Index ETF
FIGS	FIGS
FINM	Marlin Technology
FINX	Global X FinTech Thematic ETF
FIS	Fidelity National Information Services
FISI	Financial Institutions
FISK	Empire State Realty OP
FISV	Fiserv
FITB	Fifth Third Bancorp
FIVE	Five Below
FIVG	Defiance Next Gen Connectivity ETF
FIVN	Five9
FIW	First Trust ISE Water Index Fund
FIX	Comfort Systems
FIXD	First Trust TCW Opportunistic Fixed ome ETF
FIXX	Homology Medicines
FIZZ	National Beverage
FJUL	FT Cboe Vest U.S. Equity Buffer ETF - July Series
FL	Foot Locker
FLCB	Franklin Liberty U.S. Core Bond ETF
FLDM	Fluidigm
FLEX	Flex
FLGT	Fulgent Genetics
FLIC	First Of Long Island
FLJP	Franklin FTSE Japan ETF
FLMN	Falcon Minerals
FLNC	Fluence Energy
FLNG	FLEX LNG
FLNT	Fluent
FLO	Flowers Foods
FLOT	iShares Floating Rate Bond ETF
FLOW	Spx Flow
FLQL	Franklin LibertyQ U.S. Equity ETF
FLR	Fluor
FLRU	Franklin FTSE Russia ETF
FLS	Flowserve
FLSP	Franklin Liberty Systematic Style Premia ETF
FLT	FleetCor Technologies
FLWS	1-800 Flowers.Com
FLYW	Flywire
FM	iShares MSCI Frontier 100 ETF
FMAO	Farmers & Merchants Bancorp
FMB	First Trust Managed Municipal ETF
FMBH	First Mid Bancshares
FMC	FMC
FMF	First Trust Morningstar Managed Futures Strategy Fund
FMNB	Farmers National Banc
FMS	Fresenius Medical Care AG & KGaA
FMTX	Forma Therapeutics
FN	Fabrinet
FNB	F.N.B.
FNCB	Fncb Bancorp
FNCH	Finch Therapeutics
FNCL	Fidelity MSCI Financials Index ETF
FND	Floor & Decor
FNDA	Schwab Fundamental U.S. Small Index ETF
FNDC	Schwab Fundamental International Small Index ETF
FNDE	Schwab Fundamental Emerging Markets Large Index ETF
FNDF	Schwab Fundamental International Large Index ETF
FNDX	Schwab Fundamental U.S. Large Index ETF
FNF	Fidelity National Financial
FNGU	MicroSectors FANG+™ Index 3X Leveraged ETN
FNHC	Federated National
FNKO	Funko
FNLC	First Bancorp
FNV	Franco-Nevada
FNWB	First Northwest Bancorp
FOCS	Focus Financial Partners
FOE	Ferro
FOLD	Amicus Therapeutics
FONR	Fonar
FOR	Forestar
FORG	ForgeRock
FORM	FormFactor
FORR	Forrester Research
FOSL	FOSSIL GROUP
FOUR	Shift4 Payments
FOX	Twenty-First Century Fox Class B
FOXA	Twenty-First Century Fox Class A
FOXF	Fox Factory
FPE	First Trust Preferred Securities and ome ETF
FPEI	First Trust Institutional Preferred Securities and ome ETF
FPH	Five Point s
FPI	Farmland Partners
FPL	FIRST TRUST NEW OPPORTUNITIES MLP & ENERGY FUND
FPX	First Trust US Equity Opportunities ETF
FPXI	First Trust International IPO ETF
FR	First Industrial Realty Trust
FRAF	Franklin Finl Svcs
FRBA	First Bank Williamstown
FRBK	Republic First Bancorp
FRC	First Republic Bank
FREL	Fidelity MSCI Real Estate Index ETF
FREQ	Frequency Therapeutics
FREY	FREYR Battery
FRG	Franchise
FRGI	Fiesta Restaurant
FRHC	Freedom
FRLN	Freeline Therapeutics
FRME	First Merchants
FRO	Frontline
FROG	JFrog
FRPH	Frp
FRPT	Freshpet
FRSH	Freshworks
FRT	Federal Realty Investment Trust
FRTA	Forterra
FSBC	Five Star Bancorp
FSBW	Fs Bancorp
FSEP	FT Cboe Vest U.S. Equity Buffer ETF - September Series
FSFG	First Savings Financial
FSLR	FIRST SOLAR
FSLY	Fastly
FSM	Fortuna Silver Mines
FSP	Franklin Street Pptys
FSR	Fisker
FSS	Federal Signal
FSSI	Fortistar Sustainable Solutions
FSTA	Fidelity MSCI Consumer Staples Index ETF
FSTR	Lb Foster
FSV	FirstService
FTAI	Fortress Transportation and Infrastructure Investors LLC
FTCH	Farfetch
FTCI	FTC Solar
FTCS	First Trust Capital Strength ETF
FTDR	frontdoor
FTEC	Fidelity MSCI Information Technology Index ETF
FTGC	First Trust Global Tactical Commodity Strategy Fund
FTHM	Fathom
FTI	TechnipFMC
FTLS	First Trust Long/Short Equity Fund
FTNT	Fortinet
FTS	Fortis
FTSL	First Trust Senior Loan Exchange-Traded Fund
FTSM	First Trust Enhanced Short Maturity ETF
FTV	Fortive
FTXG	First Trust Nasdaq Food & Beverage ETF
FTXR	First Trust Nasdaq Transportation ETF
FUBO	fuboTV /FL
FUL	H.B. Fuller
FULC	Fulcrum Therapeutics
FULT	Fulton Financial
FUN	Cedar Fair L.P.
FUNC	First United
FUSN	Fusion Pharmaceuticals
FUTU	Futu
FUV	Arcimoto
FVCB	Fvcbankcorp
FVD	First Trust Value Line Dividend Index
FVRR	Fiverr International
FWONA	Formula One
FWONK	Formula One
FWRD	Forward Air
FXA	RYDEX CURRENCY SHARES AUSTRALIAN
FXD	First Trust Consumer Discretionary AlphaDEX Fund
FXE	CURRENCYSHARES EURO TRUST
FXG	First Trust Consumer Staples AlphaDEX Fund
FXH	First Trust Health Care AlphaDEX Fund
FXI	ISHARES FTSE CHINA 25 INDEX FUND
FXL	First Trust Technology AlphaDEX Fund
FXLV	F45 Training
FXO	First Trust Financials AlphaDEX Fund
FXR	First Trust Industrials/Producer Durables AlphaDEX Fund
FXY	RYDEX CURRENCYSHARES JAPANESE
FYBR	Frontier Communications Parent
FYC	First Trust Small Cap Growth AlphaDEX Fund
FYT	First Trust Small Cap Value AlphaDEX Fund
FYX	First Trust Small Cap Core AlphaDEX Fund
G	Genpact
GABC	German American Bancorp
GAIA	Gaiam
GAL	SPDR SSgA Global Allocation ETF
GALT	Galectin Therapeutics
GAMB	Gambling.com
GAN	Gan
GANX	Gain Therapeutics
GATO	Gatos Silver
GATX	Gatx
GBCI	Glacier Bancorp
GBIL	Goldman Sachs Treasury Access 0-1 Year ETF
GBIO	Generation Bio
GBL	Gamco Investors
GBLI	Global Indemnity
GBT	Global Blood Therapeutics
GBX	Greenbrier Companies
GCBC	Greene County Bancorp
GCI	GANNETT CO.
GCO	Genesco
GCP	Gcp Applied Technologies
GD	General Dynamics
GDDY	GoDaddy
GDEN	Golden Entertainment
GDOT	Green Dot
GDRX	GoodRx
GDS	GDS
GDX	MARKET VECTORS GOLD MINERS ETF
GDXJ	MARKET VECTORS JUNIOR GOLD MINE
GDYN	Grid Dynamics
GE	General Electric
GEF	Greif
GEF.B	Greif
GEL	Genesis Energy
GEM	Goldman Sachs ActiveBeta Emerging Markets Equity ETF
GENC	Gencor Industries
GENI	Genius Sports
GEO	Geo
GERN	Geron
GES	Guess
GEVO	Gevo
GFF	Griffon
GFI	Gold Fields
GFL	GFL Environmental
GFS	GlobalFoundries Ordinary Shares
GGAL	Grupo Financiero Galicia S.A.
GGB	Gerdau S.A.
GGG	Graco
GGPI	Gores Guggenheim
GH	Guardant Health
GHC	GRAHAM HOLDINGS
GHL	Greenhill & Co
GHLD	Guild Co
GHM	Graham
GHRS	GH Research
GIB	CGI
GIC	Global Industrial
GIII	G-iii Apparel
GIL	Gildan Activewear
GILD	Gilead Sciences
GILT	Gilat Satellite Networks
GIS	General Mills
GIW	GigInternational1
GKOS	Glaukos
GL	GLOBE LIFE
GLBE	Global-E Online
GLD	SPDR GOLD SHARES
GLDD	Great Lakes Dredge & Dock
GLDM	SPDR Gold MiniShares Trust
GLNG	Golar Lng
GLOB	Globant S.A.
GLP	Global Partners LP
GLPI	Gaming and Leisure Properties
GLRE	Greenlight Capital Re
GLSI	Greenwich LifeSciences
GLT	Glatfelter
GLTO	Galecto
GLTR	Aberdeen Standard Physical Precious Metals Basket Shares ETF
GLUE	Monte Rosa Therapeutics
GLW	Corning
GLYC	Glycomimetics o
GM	General Motors
GMAB	Genmab A/S
GME	GAMESTOP CORPORATION
GMED	Globus Medical
GMF	SPDR S&P Emerging Asia Pacific ETF
GMRE	Global Medical Reit Commo
GMS	Gms
GNE	Genie Energy
GNK	Genco Shipping & Trading
GNL	Global Net Lease
GNLN	Greenlane
GNOG	Golden Nugget Online Gaming
GNPX	Genprex
GNR	SPDR S&P Global Natural Resources ETF
GNRC	Generac
GNSS	Genasys
GNTX	Gentex
GNTY	Guaranty Bancshares
GNUS	Genius Brands International
GNW	GENWORTH FINANCIAL INC
GO	Grocery Outlet
GOCO	GoHealth
GOED	1847 Goedeker
GOEV	Canoo
GOGL	Golden Ocean
GOGO	Gogo
GOLD	Barrick Gold
GOLF	Acushnet
GOOD	Gladstone Commercial
GOOG	Alphabet Class C
GOOGL	Alphabet Class A
GOOS	Canada Goose
GORO	Gold Resource
GOSS	Gossamer Bio
GOVT	iShares US Treasury Bond
GP	Georgia-Pacific
GPC	Genuine Parts
GPI	Group 1 Automotive
GPK	Graphic Packaging
GPMT	Granite Point Mortgage Trust
GPN	Global Payments
GPRE	Green Plains
GPRK	GeoPark
GPRO	Gopro o
GPS	Gap
GQRE	FlexShares Global Quality Real Estate Index Fund
GRAY	GRAYBUG VISION
GRBK	Green Brick Energy
GRC	Gorman-rupp
GRCL	Gracell Biotechnologies
GREE	Greenidge Generation Class A Common
GRFS	Grifols S.A.
GRIL	Muscle Maker
GRMN	Garmin
GROY	Gold Royalty
GRPH	Graphite Bio
GRPN	Groupon
GRTS	Gritstone Oncology
GRTX	Galera Therapeutics
GRUB	Grubhub
GRVI	Grove
GRWG	Growgeneration .
GS	Goldman Sachs
GSAT	Globalstar
GSBC	Great Southern Bancorp
GSEV	Gores s VII
GSG	iShares S&P GSCI Commodity-Indexed Trust
GSHD	Goosehead Insurance
GSIE	Goldman Sachs ActiveBeta International Equity ETF
GSIT	Gsi Technology
GSKY	Greensky
GSL	Global Ship Lease
GSLC	Goldman Sachs ActiveBeta U.S. Large Cap Equity ETF
GSM	Ferroglobe
GSSC	Goldman Sachs ActiveBeta US Small Cap Equity ETF
GSY	Invesco Ultra Short Duration ETF
GT	THE GOODYEAR TIRE & RUBBER
GTES	Gates Industrial
GTH	Genetron
GTHX	G1 Therapeutics
GTLB	GitLab Class A
GTLS	Chart Industries
GTN	Gray Television
GTO	Invesco Total Return Bond ETF
GTPA	Gores Technology Partners
GTX	Garrett Technologies
GTY	Getty Realty
GTYH	Gty Technology
GUNR	FLEXSHARES GLOBAL UPSTREAM
GUSH	Direxion Daily S&P Oil & Gas Exploration & Production Bull 2x Shares
GVA	Granite Construction
GVI	iShares Intermediate Government/Credit Bond ETF
GWGH	Gwg
GWH	ESS Tech
GWRE	Guidewire Software
GWRS	Global Water Resources
GWW	Grainger (W.W.)
GWX	SPDR S&P International Small Cap ETF
GXC	SPDR S&P CHINA ETF
GXO	GXO Logistics
GYRO	Gyrodyne
H	Hyatt Hotels
HA	Hawaiian
HACK	ETFMG Prime Cyber Security ETF
HAE	Haemonetics
HAFC	Hanmi Financial
HAIL	SPDR S&P Smart Mobility ETF
HAIN	The Hain Celestial
HAL	Halliburton
HALO	Halozyme Therapeutics
HARP	Harpoon Therapeutics
HAS	Hasbro
HAUZ	Xtrackers International Real Estate ETF
HAYN	Haynes International
HAYW	Hayward
HBAN	Huntington Bancshares
HBB	Hamilton Beach Brands Hldg Co
HBCP	Home Bancorp
HBI	Hanesbrands
HBIO	Harvard Bioscience
HBM	Hudbay Minerals
HBNC	Horizon Bancorp
HBT	Hbt Financial
HCA	HCA
HCAT	Health Catalyst
HCC	Warrior Met Coal
HCCI	Heritage-crystal Clean
HCDI	Harbor Custom Development
HCI	Hci
HCKT	Hackett
HCSG	Healthcare Services
HCWB	HCW Biologics
HD	Home Depot
HDG	ProShares Hedge Replication
HDGE	AdvisorShares Ranger Equity Bear ETF
HDV	iShares Core High Dividend ETF
HE	Hawaiian Electric Industries
HEAR	Turtle Beach Commo
HEEM	iShares Currency Hedged MSCI Emerging Markets ETF
HEES	H&e Equipment Services
HEI	HEICO
HEI.A	Heico
HELE	Helen Of Troy
HEP	Holly Energy Partners
HEPS	D-MARKET Electronic Services & Trading
HERO	Global X Video Games & Esports ETF
HES	Hess
HESM	Hess Midstream LP
HEWJ	iShares Currency Hedged MSCI Japan ETF
HEXO	HEXO
HFC	HollyFrontier
HFFG	Hf Foods
HFWA	Heritage Financial
HGEN	Humanigen
HGV	Hilton Grand Vacations
HHC	The Howard Hughes
HI	Hillenbrand
HIBB	Hibbett Sports
HIFS	Hingham Institution
HIG	Hartford Financial Svc.Gp.
HII	Huntington Ingalls Industries
HIMS	Hims & Hers Health
HIPO	Hippo
HIVE	Hive Blockchain Technologies Common Shares
HIW	Highwoods Properties
HL	Hecla Mining
HLF	Herbalife Nutrition
HLI	Houlihan Lokey
HLIO	Helios Technologies
HLIT	Harmonic
HLMN	Hillman Solutions
HLNE	Hamilton Lane
HLT	Hilton Worldwide
HLTH	Cue Health
HLX	Helix Energy Solutions
HMHC	Houghton Mifflin Harcourt Comp
HMLP	HÃ¶egh LNG Partners LP
HMN	Horace Mann Educators
HMST	Homestreet
HMTV	Hemisphere Media c
HNDL	NASDAQ 7HANDL Index ETF
HNGR	Hanger
HNI	Hni
HNST	The Honest
HOFT	Hooker Furniture
HOG	Harley-Davidson
HOLI	Hollysys Automation Technologies
HOLX	Hologic
HOMB	Home Bancshares (Conway AR)
HON	Honeywell Int'l
HONE	Harborone Bancorp
HOOD	Robinhood Markets
HOOK	Hookipa Pharma
HOPE	Hope Bancorp
HOV	Hovnanian Enterprises
HOWL	Werewolf Therapeutics
HP	Helmerich & Payne
HPE	Hewlett Packard Enterprise
HPK	HighPeak Energy
HPP	Hudson Pacific Properties
HPQ	HP
HPX	HPX .
HQY	HealthEquity
HR	Healthcare Realty Trust
HRB	Block H&R
HRI	Herc
HRL	Hormel Foods
HRMY	Harmony Biosciences
HROW	Harrow Health
HRTG	Heritage Insurance In
HRTX	Heron Therapeutics Commo
HSC	Harsco
HSIC	Henry Schein
HSII	Heidrick & Struggles
HSKA	Heska
HST	Host Hotels & Resorts
HSTM	Healthstream
HSY	The Hershey
HT	Hersha Hospitality
HTA	Healthcare Trust of America
HTBI	Hometrust Bancshares
HTBK	Heritage Commerce
HTH	Hilltop
HTHT	Huazhu
HTLD	Heartland Express
HTLF	Heartland Financial Usa
HTUS	Hull Tactical US ETF
HTZ	Hertz Global
HUBB	Hubbell
HUBG	Hub
HUBS	HubSpot
HUDI	Huadi International
HUIZ	Huize
HUM	Humana
HUN	Huntsman
HURC	Hurco Companies
HURN	Huron Consulting
HUT	Hut 8 Mining
HUYA	HUYA
HVT	Haverty Furniture
HWBK	Hawthorn Bancshares
HWC	Hancock Whitney
HWKN	Hawkins
HWM	Howmet Aerospace
HXL	Hexcel
HY	Hyster-yale Materials
HYD	MARKET VECTORS HIGH YIELD
HYEM	VanEck Vectors Emerging Markets High Yield Bond ETF
HYFM	Hydrofarm
HYG	ISHARES IBOXX $ HIGH YIELD CORPORATE BD
HYLB	Xtrackers USD High Yield orate Bond
HYLN	Hyliion s
HYLS	First Trust Tactical High Yield ETF
HYMB	SPDR Nuveen S&P High Yield Municipal Bond ETF
HYS	PIMCO 0-5 YEAR HIGH YIELD CORP
HYW	Hywin
HYZN	Hyzon Motors Class A
HZNP	Horizon Therapeutics Public
HZO	Marinemax
IAA	IAA
IAC	IAC/INTERACTIVECORP
IAG	IAMGOLD
IAGG	iShares Core International Aggregate Bond Fund
IAI	iShares US Broker-Dealers ETF
IART	Integra LifeSciences
IAS	Integral Ad Science
IAT	iShares U.S. Regional Banks ETF
IAU	iShares Gold Trust
IBB	ISHARES NASDAQ BIOTECHNOLOGY IN
IBCP	Independent Bank (ionia
IBDN	iShares iBonds Dec 2022 orate ETF
IBDO	iShares iBonds Dec 2023 orate ETF
IBDP	iShares iBonds Dec 2024 orate ETF
IBER	Ibere Pharmaceuticals
IBEX	IBEX
IBIO	Ibio
IBKR	Interactive Brokers
IBM	International Business Machines
IBN	ICICI Bank
IBOC	International Bancshares
IBP	Installed Building Products
IBRX	ImmunityBio
IBTX	Independent Bank Co
IBUY	Amplify Online Retail ETF
ICAD	Icad
ICE	Intercontinental Exchange
ICF	ISHARES COHEN & STEERS REALTY M
ICFI	Icf International
ICHR	Ichor
ICL	ICL
ICLN	iShares Global Clean Energy ETF
ICLR	ICON Public
ICPT	Intercept Pharmaceuticals
ICSH	iShares Ultra Short-Term Bond ETF
ICUI	Icu Medical
ICVT	iShares Convertible Bond ETF
ICVX	Icosavax
IDA	IDACORP
IDCC	Interdigital
IDEV	iShares Core MSCI International Developed Markets ETF
IDEX	Ideanomics
IDN	Intellicheck
IDT	Idt
IDU	ISHARES DOW JONES U.S. UTILITIE
IDV	ISHARES DOW JONES EPAC SELECT D
IDXX	IDEXX Laboratories
IDYA	Ideaya Biosciences
IECS	iShares Evolved U.S. Consumer Staples ETF
IEF	ISHARES LEHMAN 7-10 YEAR TREASU
IEFA	iShares Core MSCI EAFE
IEI	ISHARES LEHMAN 3-7 YEAR TREASUR
IEMG	iShares Core MSCI Emerging Markets
IEP	Icahn Enterprises L.P.
IESC	Ies
IETC	iShares Evolved U.S. Technology ETF
IEUR	iShares Core MSCI Europe ETF
IEUS	iShares MSCI Europe Small-Cap ETF
IEV	iShares Europe ETF
IEX	IDEX CORPORATION
IFBD	Infobird
IFF	Intl Flavors & Fragrances
IFGL	iShares International Developed Real Estate ETF
IFS	Intercorp Financial Services
IGE	iShares North American Natural Resources ETF
IGIB	ISHARES LEHMAN INTERMEDIATE
IGLB	iShares Long-Term orate Bond ETF
IGM	iShares Expanded Tech Sector ETF
IGMS	Igm Biosciences
IGSB	iShares Short-Term orate Bond ETF
IGT	INTERNATIONAL GAME TECHNOLOGY
IGV	iShares Expanded Tech-Software Sector ET
IH	iHuman
IHDG	WisdomTree International Hedged Dividend Growth Fund
IHE	iShares U.S. Pharmaceuticals ETF
IHF	iShares U.S. Healthcare Providers ETF
IHG	InterContinental Hotels
IHI	iShares U.S. Medical Devices ETF
IHRT	iHeartMedia
IHS	IHS
III	Information Services Grp
IIIN	Insteel Industries
IIIV	I3 Verticals
IIN	Intricon
IINN	Inspira Technologies Oxy
IIPR	Innovative Industrial Properties
IIVI	II-VI
IJH	ISHARES S&P MIDCAP 400 INDEX FU
IJJ	ISHARES S&P MIDCAP 400 VALUE INDEX
IJK	ISHARES S&P MIDCAP 400 GROWTH
IJR	ISHARES S&P SMALLCAP 600 INDEX
IJS	iShares S&P SmallCap 600 Value ETF
IJT	iShares S&P SmallCap 600 Growth ETF
IKNA	Ikena Oncology
IKT	Inhibikase Therapeutics
ILF	ISHARES LATIN AMERICA 40 INDEX
ILMN	Illumina
ILPT	Industrial Logistics Properties Trust Common Share
IMAB	I-Mab
IMAX	Imax
IMCR	Immunocore
IMGN	Immunogen
IMGO	Imago BioSciences
IMKTA	Ingles Markets
IMMR	Immersion
IMNM	Immunome
IMO	Imperial Oil
IMPL	Impel NeuroPharma
IMRA	Imara
IMRX	Immuneering
IMUX	Immunic
IMVT	Immunovant
IMXI	International Money Express
INAB	IN8bio
INBK	First Internet Bancorp Common
INBX	Inhibrx
INCO	Columbia India Consumer ETF
INCY	Incyte
INDA	iShares MSCI India
INDB	Independent Bank
INDI	indie Semiconductor
INDS	Pacer Benchmark Industrial Real Estate SCTR ETF
INDT	INDUS Realty Trust
INDY	iShares India 50 ETF
INFA	Informatica
INFN	Infinera
INFU	Infusystems
INFY	Infosys ADR
ING	ING Groep N.V.
INGN	Inogen
INGR	Ingredion
INMD	InMode
INN	Summit Hotel Properties
INNV	InnovAge
INO	Inovio Pharmaceuticals
INSG	Inseego .
INSM	Insmed
INSP	Inspire Medical Systems
INST	Instructure
INSW	International Seaways
INT	World Fuel Services
INTA	Intapp
INTC	Intel
INTU	Intuit
INTZ	Intrusion
INVA	Innoviva
INVH	Invitation Homes
INVZ	Innoviz Technologies Ordinary shares
INZY	Inozyme Pharma
IONQ	IonQ
IONS	Ionis Pharmaceuticals
IOO	iShares Global 100 ETF
IOSP	Innospec
IOVA	Iovance Biotherapeutics
IP	International Paper
IPAC	iShares Core MSCI Pacific ETF
IPAR	Inter Parfums
IPAY	ETFMG Prime Mobile Payments ETF
IPG	Interpublic
IPGP	IPG Photonics
IPI	Intrepid Potash
IPSC	Century Therapeutics
IPW	iPower
IQ	iQIYI
IQDE	FlexShares International Quality Dividend Defensive Index Fund
IQDF	FlexShares International Quality Dividend Index Fund
IQDG	WisdomTree International Quality Dividend Growth Fund
IQLT	iShares MSCI Intl Quality Factor ETF
IQV	IQVIA
IR	Ingersoll-Rand
IRBO	iShares Robotics and Artificial Intelligence ETF
IRBT	iRobot
IRDM	Iridium Communications
IRM	Iron Mountain
IRMD	Iradimed
IRT	Independence Realty Trust
IRTC	iRhythm Technologies
IRWD	Ironwood Pharmaceuticals
IS	ironSource
ISBC	Investors Bancorp
ISEE	Iveric Bio
ISMD	Inspire Small/Mid Cap Impact ETF
ISPC	iSpecimen
ISRA	VanEck Vectors Israel ETF
ISRG	Intuitive Surgical
ISTB	iShares Core 1-5 Year USD Bond ETF
ISTR	Investar
IT	Gartner
ITA	iShares U.S. Aerospace & Defense ETF
ITB	ISHARES DOW JONES U.S. HOME
ITCI	Intra-cellular Therapies
ITEQ	BlueStar Israel Technology ETF
ITGR	Integer
ITI	Iteris
ITIC	Investors Title
ITM	VanEck Vectors AMT-Free Intermediate Municipal Index ETF
ITOS	iTeos Therapeutics
ITOT	iShares Core S&P Total U.S. Stock Market ETF
ITRI	Itron
ITRN	Ituran Location and Control
ITT	ITT
ITUB	Itau Unibanco S.A.
ITW	Illinois Tool Works
IUSB	iShares Core Total USD Bond Market ETF
IUSG	iShares Core S&P U.S. Growth ETF
IUSV	iShares Core S&P U.S. Value ETF
IVA	Inventiva
IVAC	Intevac
IVC	Invacare
IVE	ISHARES S&P 500 VALUE INDEX FUN
IVES	ETFMG Drone Economy Strategy ETF
IVLU	iShares Edge MSCI Intl Value Factor ETF
IVOL	Quadratic Interest Rate Volatility and Inflation Hedge ETF
IVOO	Vanguard S&P Mid-Cap 400 ETF
IVR	Invesco Mortgage Capital
IVT	InvenTrust Properties .
IVV	ISHARES S&P 500 INDEX
IVW	ISHARES S&P 500 GROWTH INDEX FUND
IVZ	Invesco
IWB	ISHARES RUSSELL 1000
IWC	iShares Microcap ETF
IWD	ISHARES RUSSELL 1000 VALUE
IWF	ISHARES RUSSELL 1000 GROWTH INDEX
IWL	iShares Russell Top 200 ETF
IWM	ISHARES RUSSELL 2000 INDEX
IWN	ISHARES RUSSELL 2000 VALUE
IWO	ISHARES RUSSELL 2000 GROWTH INDEX
IWP	ISHARES RUSSELL MIDCAP GROWTH
IWR	ISHARES RUSSELL MIDCAP INDEX FUND
IWS	ISHARES RUSSELL MIDCAP VALUE INDEX
IWV	ISHARES RUSSELL 3000
IWX	iShares Russell Top 200 Value ETF
IWY	iShares Russell Top 200 Growth ETF
IXC	iShares Global Energy ETF
IXG	iShares Global Financials ETF
IXJ	iShares Global Healthcare ETF
IXN	iShares Global Tech ETF
IXP	iShares Global Telecom ETF
IXUS	iShares Core MSCI Total International Stock ETF
IYC	iShares US Consumer Services ETF
IYE	ISHARES DOW JONES U.S. ENERGY
IYF	ISHARES DOW JONES U.S. FINANCIA
IYG	iShares U.S. Financial Services ETF
IYH	iShares U.S. Healthcare ETF
IYJ	ISHARES DOW JONES U.S. INDUSTRIAL
IYK	iShares U.S. Consumer Goods ETF
IYLD	iShares Morningstar Multi-Asset ome Index ETF
IYM	iShares U.S. Basic Materials ETF
IYR	ISHARES DOW JONES US REAL ESTATE
IYT	ISHARES DOW JONES TRANSPORTATION
IYW	ISHARES DOW JONES U.S. TECH
IYY	iShares Dow Jones U.S. ETF
IYZ	iShares U.S. Telecommunications ETF
IZRL	ARK Israel Innovative Technology ETF
J	JACOBS ENGINEERING GROUP
JAAA	Janus Henderson AAA CLO ETF
JACK	Jack In The Box
JAMF	JAMF
JANX	Janux Therapeutics
JAZZ	Jazz Pharmaceuticals
JBGS	JBG SMITH Properties
JBHT	J. B. Hunt Transport Services
JBI	Janus International
JBL	JABIL
JBLU	JetBlue Airways
JBSS	John B Sanfilippo
JBT	John Bean Technologies
JCI	Johnson Controls International
JD	JD.Com
JDST	Direxion Daily Junior Gold Miners Index Bear 2X
JEF	Jefferies Financial
JELD	Jeld-wen
JEPI	JPMorgan Equity Premium ome ETF
JETS	US Global Jets
JHEM	John Hancock Multifactor Emerging Markets ETF
JHG	Janus Henderson
JHMD	John Hancock Multifactor Developed International ETF
JHMM	John Hancock Multi-Factor Mid Cap ETF
JHMS	John Hancock Multifactor Consumer Staples ETF
JHMT	John Hancock Multi-Factor Technology ETF
JJSF	J&j Snack Foods
JKHY	Jack Henry & Associates
JLL	Jones Lang Lasalle
JNCE	Jounce Therapeutics
JNJ	Johnson & Johnson
JNK	SPDR BARCLAYS CAPITAL HIGH YIELD BOND
JNPR	Juniper Networks
JNUG	Direxion Daily Junior Gold Miners Index Bull 2x Shares
JOAN	JOANN
JOBY	Joby Aviation
JOE	St. Joe
JOUT	Johnson Outdoors
JP	Jefferson-Pilot
JPHY	JPMorgan High Yield Research Enhanced
JPIN	JPMorgan Diversified Return International Equity ETF
JPM	JPMorgan Chase &
JPSE	JPMorgan Diversified Return U.S. Small Cap Equity ETF
JPST	JPMorgan Ultra-Short ome ETF
JRVR	James River Lt
JSML	Janus Henderson Small Cap Growth Alpha ETF
JUPW	Jupiter Wellness
JW.A	John Wiley & Sons
JWEL	Jowell Global
JWN	Nordstrom
JXN	Jackson Financial
JYNT	The Joint .
JZXN	Jiuzi
K	Kellogg
KAI	Kadant
KALA	Kala Pharmaceuticals
KALU	Kaiser Aluminum
KALV	Kalvista Pharmaceuticals
KAMN	Kaman
KAR	KAR Auction Services
KARO	Karooooo
KBA	KraneShares Bosera MSCI China A Share ETF
KBAL	Kimball International Cl
KBE	SPDR S&P BANK ETF
KBH	KB HOME
KBNT	Kubient
KBR	KBR
KBWB	Invesco KBW Bank ETF
KBWD	Invesco KBW High Dividend Yield Financial ETF
KBWP	Invesco KBW Property & Casualty Insurance ETF
KBWY	Invesco KBW Premium Yield Equity REIT ETF
KC	Kingsoft Cloud
KD	Kyndryl
KDNY	Chinook Therapeutics
KDP	Keurig Dr Pepper
KE	Kimball Electronics Comm
KELYA	Kelly Services
KEMQ	KraneShares Emerging Markets Consumer Technology ETF
KEN	Kenon s
KERN	Akerna
KEX	Kirby
KEY	KeyCorp
KEYS	Keysight Technologies
KFRC	Kforce
KFY	Korn Ferry
KGC	Kinross Gold
KHC	Kraft Heinz Co
KIDS	Orthopediatrics .
KIE	SPDR S&P Insurance ETF
KIM	Kimco Realty
KIND	Nextdoor
KINZ	Kins Technology
KKR	KKR &
KLAC	KLA-Tencor
KLDO	Kaleido Biosciences
KLIC	Kulicke and Soffa Industries
KLTR	Kaltura
KLXE	KLX Energy Services
KMB	Kimberly-Clark
KMI	Kinder Morgan
KMLM	KFA Mount Lucas Index Strategy ETF
KMPR	Kemper
KMT	Kennametal
KMX	Carmax
KN	Knowles
KNOP	KNOT Offshore Partners LP
KNSA	Kiniksa Pharmaceuticals
KNSL	Kinsale Capital
KNTE	Kinnate Biopharma
KNX	Knight-Swift Transportation
KO	Coca-Cola
KOD	Kodiak Sciences
KODK	Eastman Kodak
KOIN	Innovation Shares NextGen Protocol ETF
KOMP	SPDR S&P Kensho New Economies Composite ETF
KOP	Koppers
KOPN	Kopin
KOS	Kosmos Energy
KPLT	Katapult
KPTI	Karyopharm Therapeutics c
KR	Kroger
KRA	Kraton
KRBN	KraneShares Global Carbon ETF
KRBP	Kiromic BioPharma
KRC	Kilroy Realty
KRE	SPDR S&P REGIONAL BANKING ETF
KREF	Kkr Real Estate Finance Trust
KRG	Kite Realty Trust
KRMD	Repro Med Systems
KRNL	Kernel
KRNT	Kornit Digital
KRNY	Kearny Financial
KRO	Kronos Worldwide
KRON	Kronos Bio
KROS	Keros Therapeutics
KRP	Kimbell Royalty Partners
KRT	Karat Packaging
KRTX	Karuna Therapeutics
KRUS	Kura Sushi Usa
KRYS	Krystal Biotech
KSS	Kohl's
KT	KT
KTB	Kontoor Brands
KTOS	Kratos Defense & Security
KUKE	Kuke Music
KULR	KULR Technology
KURA	Kura Oncology
KVHI	Kvh Industries
KW	Kennedy-Wilson
KWEB	KraneShares CSI China Internet ETF
KWR	Quaker Chemical
KXI	iShares Global Consumer Staples ETF
KYMR	Kymera Therapeutics
KZR	Kezar Life Sciences
L	Loews
LABD	Direxion Daily S&P Biotech Bear 3X Shares
LABP	Landos Biopharma
LABU	Direxion Daily S&P Biotech Bull 3x Shares
LAC	Lithium Americas
LAD	Lithia Motors
LADR	Ladder Capital
LAKE	Lakeland Industries
LAMR	Lamar Advertising (REIT)
LANC	Lancaster Colony
LAND	Gladstone Land Com
LARK	Landmark Bancorp
LASR	Nlight
LAUR	Laureate Education
LAW	CS Disco
LAWS	Lawson Products
LAZ	Lazard
LAZR	Luminar Technologies
LBAI	Lakeland Bancorp
LBC	Luther Burbank
LBPH	Longboard Pharmaceuticals
LBRDA	Liberty Broadband
LBRDK	Liberty Broadband
LBRT	Liberty Oilfield Services
LBTYA	Liberty Global
LBTYB	Liberty Global
LBTYK	Liberty Global Class C
LC	Lendingclub
LCI	Lannett
LCID	Lucid
LCII	Lci Industries
LCNB	Lcnb
LCUT	Lifetime Brands
LDEM	iShares ESG MSCI EM Leaders ETF
LDI	loanDepot
LDOS	LEIDOS HOLDINGS
LDUR	PIMCO Enhanced Low Duration Active Exchange-Traded Fund
LE	Lands End
LEA	Lear
LEAP	Ribbit LEAP
LECO	Lincoln Electric
LEG	Leggett & Platt
LEGH	Legacy Housing
LEGN	Legend Biotech
LEGR	First Trust Indxx Innovative Transaction & Process ETF
LEN	Lennar
LEN.B	Lennar
LESL	Leslie's
LEV	The Lion Electric
LEVI	Levi Strauss &
LEVL	Level One Bancorp
LFG	Archaea Energy
LFST	Lifestance Health
LFUS	Littelfuse
LFVN	Lifevantage
LGF.A	Lions Gate Entertainment
LGF.B	Lions Gate Entertainment
LGIH	LGI Homes
LGND	Ligand Pharmaceuticals
LGVN	Longeveron
LH	Laboratory of America
LHCG	LHC
LHDX	Lucira Health
LHX	L3HARRIS TECHNOLOGIES
LI	Li Auto
LICY	Li-Cycle .
LIFE	ATYR PHARMA
LII	Lennox International
LILA	Liberty Latin America
LILAK	Liberty Latin America
LILM	Lilium N.V. Class A Ordinary Shares
LIN	Linde
LIND	Lindblad Expeditions
LIT	Global X Lithium ETF
LITE	Lumentum
LIVN	LivaNova
LJPC	La Jolla Pharmaceutical Compan
LKFN	Lakeland Financial
LKQ	LKQ
LL	Lumber Liquidators
LLNW	Limelight Networks
LLY	Lilly (Eli) &
LMAT	Lemaitre Vascular
LMBS	First Trust Low Duration Mortgage Opportunities ETF
LMND	Lemonade
LMNR	Limoneira Co
LMST	Limestone Bancorp
LMT	Lockheed Martin
LNC	Lincoln National
LNDC	Landec
LNG	Cheniere Energy
LNN	Lindsay
LNSR	LENSAR
LNT	Alliant Energy
LNTH	Lantheus
LOB	Live Oak Bancshares
LOCO	El Pollo Loco c
LOGC	Logicbio Therapeutics
LOGI	Logitech International SA
LOPE	Grand Canyon Education
LOUP	Innovator Loup Frontier Tech ETF
LOVE	The Lovesac
LOW	Lowe's Cos.
LPG	Dorian Lpg
LPI	Laredo Petroleum
LPL	LG Display
LPLA	LPL Financial
LPRO	Open Lending Class A
LPSN	LivePerson
LPX	Louisiana-Pacific
LQD	ISHARES IBOXX $ INVEST GRADE CORP BOND
LQDA	Liquidia
LQDH	iShares Interest Rate Hedged orate Bond ETF
LQDT	Liquidity Service
LRCX	Lam Research
LRN	Stride
LRNZ	TrueMark AI & Deep Learning ETF
LSCC	Lattice Semiconductor
LSF	Laird Superfood
LSI	LIFE STORAGE
LSPD	Lightspeed Commerce
LSTR	Landstar System
LSXMA	The Liberty SiriusXM
LSXMB	The Liberty SiriusXM
LSXMK	The Liberty SiriusXM
LTC	Ltc Properties
LTCH	Latch
LTH	Life Time
LTHM	Livent
LTL	Ultra Telecommunications ProShares
LTPZ	PIMCO 15+ Year U.S. TIPS Index Exchange-Traded Fund
LTRN	Lantern Pharma
LTRPA	Liberty Tripadvisor
LTRPB	Liberty TripAdvisor s
LU	Lucent Technology
LULU	Lululemon Athletica
LUMN	Lumen Technologies
LUNA	Luna Innovations
LUNG	Pulmonx
LUV	Southwest Airlines
LVOX	LiveVox
LVRA	Levere
LVS	Las Vegas Sands
LVTX	LAVA Therapeutics NV
LW	Lamb Weston
LWLG	Lightwave Logic
LXEH	Lixiang Education
LXFR	Luxfer
LXP	Lexington Realty Trust
LXRX	Lexicon Pharmaceuticals
LYB	LyondellBasell
LYEL	Lyell Immunopharma
LYFT	Lyft
LYG	Lloyds Banking
LYLT	Loyalty Ventures
LYRA	Lyra Therapeutics
LYTS	Lsi Industries
LYV	Live Nation Entertainment
LZ	LegalZoom.com
LZB	La-z-boy
M	Macy's
MA	Mastercard
MAA	Mid-America Apartments
MAC	Macerich
MACC	Mission Advancement
MAG	MAG Silver
MAIN	Main Street Capital
MAN	ManpowerGroup
MANH	Manhattan Associates
MANT	Mantech International
MANU	Manchester United
MAPS	WM Technology
MAR	Marriott Int'l.
MARA	Marathon Digital s
MARB	First Trust Merger Arbitrage ETF
MAS	Masco
MASI	Masimo
MASS	908 Devices
MAT	Mattel
MATW	Matthews International
MATX	Matson
MAX	MediaAlpha
MAXN	Maxeon Solar
MAXR	Maxar Technologies
MBB	ISHARES BARCLAYS MBS BOND FUND
MBCN	Middlefield Banc Cmn
MBI	MBIA
MBII	Marrone Bio Innovations
MBIN	Merchants Bancorp
MBIO	Mustang Bio
MBT	Mobile TeleSystems Public Joint Stock
MBUU	Malibu Boats Com
MBWM	Mercantile Bank
MC	Moelis &
MCB	Metropolitan Bank Hldg
MCBC	Macatawa Bank
MCBS	Metrocity Bankshares
MCD	McDonald's
MCFT	Mastercraft Boat
MCG	Membership Collective
MCHI	iShares MSCI China
MCHP	Microchip Technology
MCK	McKesson
MCO	Moody's
MCRB	Seres Therapeutics
MCRI	Monarch Casino & Resort
MCS	Marcus
MCW	Mister Car Wash
MCY	Mercury General
MD	Mednax
MDB	MongoDB
MDC	M.D.C.
MDGL	Madrigal Pharmaceuticals
MDIV	First Trust Multi-Asset Diversified ome Index Fund
MDLZ	Mondelez International
MDRX	Allscripts-misys
MDT	Medtronic
MDU	MDU Resources
MDWT	Midwest
MDXG	MiMedx
MDY	SPDR S&P MIDCAP 400
MDYG	SPDR S&P 400 Mid Cap Growth ETF
MDYV	SPDR S&P 400 Mid Cap Value ETF
ME	23andMe
MEC	Mayville Engineering Co
MED	Medifast
MEDP	Medpace
MEDS	Trxade
MEG	Montrose Environmental
MEI	Methode Electronics
MEIP	Mei Pharma
MELI	Mercadolibre
MEOH	Methanex
MERC	Mercer International
MESA	Mesa Air
MET	MetLife
METC	Ramaco Resources
METX	Meten Ordinary Shares
MF	Missfresh
MFA	Mfa Financial
MFC	Manulife Financial
MFEM	PIMCO RAFI Dynamic Multi-Factor Emerging Market Equity ETF
MFG	Mizuho Financial
MFGP	softwareinto Micro Focus
MG	Mistras
MGA	Magna International
MGC	Vanguard Mega Cap ETF
MGEE	Mge Energy
MGI	Moneygram International
MGIC	Magic Software Enterprises
MGK	Vanguard Mega Cap Growth ETF
MGM	MGM Resorts International
MGNI	Magnite
MGNX	Macrogenics
MGP	MGM Growth Properties LLC
MGPI	Mgp Ingredients
MGRC	Mcgrath Rentcorp
MGTA	Magenta Therapeutics
MGTX	Meiragtx
MGV	Vanguard Mega Cap Value ETF
MGY	Magnolia Oil & Gas
MHH	Mastech Digital
MHK	Mohawk Industries
MHO	M/i Homes
MIC	Macquarie Infrastructure
MIDD	The Middleby
MILE	Metromile
MIME	Mimecast
MINT	PIMCO ENH SHRT MATURITY
MIR	Mirion Technologies
MIRM	Mirum Pharmaceuticals
MIRO	Miromatrix Medical
MITK	Mitek Systems
MITQ	Moving iMage Technologies
MJ	ETFMG Alternative Harvest ETF
MKC	McCormick &
MKL	Markel
MKSI	Mks Instruments
MKTW	MarketWise
MKTX	MARKETAXESS HOLDINGS
ML	MoneyLion
MLAB	Mesa Laboratories
MLCO	Melco Resorts & Entertainment
MLI	Mueller Industries
MLKN	MillerKnoll
MLM	Martin Marietta Materials
MLNK	MeridianLink
MLP	Maui Land & Pineapple
MLPA	Global X MLP ETF
MLPX	Global X MLP & Energy Infrastructure ETF
MLR	Miller Industries
MLSS	Milestone Scientific
MMAT	Meta Materials
MMC	Marsh & McLennan
MMI	MARCUS & MILLICHAP
MMM	3M
MMP	Magellan Midstream Partners L.P.
MMS	MAXIMUS
MMSI	Merit Medical Systems
MMX	Maverix Metals
MMYT	MakeMyTrip
MNA	IQ Merger Arbitrage ETF
MNDT	Mandiant
MNDY	monday.com
MNKD	Mannkind
MNMD	Mind Medicine (MindMed)
MNOV	Medicinova
MNRL	Brigham Minerals
MNRO	Monro
MNSB	Mainstreet Bancshares
MNSO	MINISO
MNST	Monster Beverage
MNTK	Montauk Renewables
MNTS	Momentus
MNTV	Momentive Global
MO	Altria
MOAT	VanEck Vectors Morningstar Wide Moat ETF
MOD	Modine Manufacturing
MODN	Model N
MODV	ModivCare
MOFG	Midwestone Financial
MOG.A	Moog
MOGO	Mogo
MOH	Molina Healthcare
MOLN	Molecular Partners
MOMO	Momo
MOO	MARKET VECTORS AGRIBUSINESS ETF
MOON	Direxion Moonshot Innovators ETF
MORF	Morphic
MORN	Morningstar
MORT	VanEck Vectors Mortgage REIT ome ETF
MOS	The Mosaic
MOV	Movado
MOVE	Movano
MP	MP Materials
MPAA	Motorcar Parts Of America
MPB	Mid Penn Bancorp
MPC	Marathon Petroleum
MPLN	MultiPlan
MPLX	MPLX LP
MPW	Medical Properties Trust
MPWR	Monolithic Power Systems
MPX	Marine Products
MQ	Marqeta
MRAI	Marpai
MRBK	Meridian Bank
MRC	Mrc Global
MRCY	Mercury Systems
MRGR	Proshares Merger ETF
MRK	Merck &
MRKR	Marker Therapeutics
MRM	Medirom Healthcare
MRNA	Moderna
MRNS	Marinus Pharmaceuticals
MRO	Marathon Oil
MRSN	Mersana Therapeutics
MRTN	Marten Transport
MRTX	Mirati Therapeutics
MRUS	Merus N.V.
MRVI	Maravai Lifesciences
MRVL	Marvell Technology
MS	Morgan Stanley
MSA	MSA Safety
MSBI	Midland States Bancorp
MSCI	MSCI
MSEX	Middlesex Water
MSFT	Microsoft
MSGE	Madison Square Garden Entertainment
MSGM	Motorsport Games
MSGS	Madison Square Garden Sports
MSI	Motorola Solutions
MSM	MSC Industrial Direct
MSOS	AdvisorShares Pure US Cannabis ETF
MSP	DATTO HOLDING CORP.
MSTR	Microstrategy
MSVX	LHA Market State Alpha Seeker ETF
MT	ArcelorMittal
MTB	M&T Bank
MTCH	Match
MTCR	Metacrine
MTD	Mettler Toledo
MTDR	Matador Resources
MTEM	Molecular Templates
MTG	MGIC Investment
MTH	Meritage Homes
MTN	Vail Resorts
MTOR	Meritor
MTRN	Materion
MTRX	Matrix Service
MTSI	MACOM Technology Solutions
MTTR	Matterport
MTUM	iShares Edge MSCI USA Momentum Factor ETF
MTW	MANITOWOC (THE)
MTX	Minerals Technologies
MTZ	MasTec
MU	Micron Technology
MUB	ISHARES S&P NATIONAL MUNICIPAL
MULN	Mullen Automotive
MUR	MURPHY OIL CORPORATION
MUSA	Murphy USA
MVBF	Mvb Financial .
MVIS	MicroVision
MVST	Microvast
MVV	PROSHARES ULTRA MIDCAP400
MWA	Mueller Water Products
MX	Magnachip Semiconductor
MXCT	MaxCyte
MXI	iShares Global Materials ETF
MXL	Maxlinear
MYE	Myers Industries
MYFW	First Western Financial
MYGN	Myriad Genetics
MYOV	Myovant Sciences
MYPS	PLAYSTUDIOS
MYRG	Myr
MYTE	MYT Netherlands Parent B.V.
NABL	N-able
NAD	Nuveen Quality Municipal ome Fund
NAIL	Direxion Daily Homebuilders & Supplies Bull 3X Shares
NANR	SPDR S&P North American Natural Resources ETF
NAPA	Duckhorn Portfolio
NARI	Inari Medical
NAT	Nordic American Tankers
NATH	Nathans Famous
NATI	National Instruments
NATR	Natures Sunshine Products
NAUT	Nautilus Biotechnology
NAVI	NAVIENT CORPORATION
NBEV	Newage
NBHC	National Bank
NBIX	Neurocrine Biosciences
NBN	Northeast Bancorp
NBR	NABORS INDUSTRIES LTD.
NBSE	Neubase Therapeutics
NBTB	Nbt Bancorp
NBTX	Nanobiotix
NC	Nacco Industries
NCBS	Nicolet Bankshares
NCLH	Norwegian Cruise Line
NCMI	National Cinemedia
NCNO	Ncino
NCR	NCR
NDAQ	Nasdaq
NDLS	Noodles & Co
NDSN	Nordson
NE	NOBLE CORPORATION
NEA	Nuveen AMT-Free Quality Municipal ome Fund
NEAR	iShares Short Maturity Bond ETF
NEE	NextEra Energy
NEGG	Newegg Commerce
NEM	Newmont Mining
NEO	NeoGenomics
NEOG	Neogen
NEP	NextEra Energy Partners LP
NERD	Roundhill BITKRAFT Esports & Digital Entertainment ETF
NERV	Minerva Neurosciences Com
NESR	National Energy Services Reunited . Ordinary s
NET	Cloudflare
NEU	Newmarket
NEWR	New Relic
NEWT	Newtek Business Services
NEX	Nextier Oilfield Solutions
NEXA	Nexa Resources S.A.
NEXI	NexImmune
NEXT	Nextdecade
NFBK	Northfield Bancorp
NFG	National Fuel Gas
NFLX	Netflix
NG	NovaGold Resources
NGD	New Gold
NGG	National Grid
NGM	Ngm Biopharmaceuticals
NGMS	NeoGames S.A.
NGVC	Natural Grocers
NGVT	Ingevity
NH	Nanthealth
NHC	National Healthcare
NHI	National Health Investors
NI	NiSource
NIO	NIO
NISN	NiSun International Enterprise Development
NJR	New Jersey Resources
NKE	Nike
NKLA	Nikola
NKSH	National Bankshares
NKTR	Nektar Therapeutics
NKTX	Nkarta
NL	Nl Industries
NLOK	NORTONLIFELOCK
NLS	Nautilus
NLSN	Nielsen
NLSP	NLS Pharmaceutics
NLTX	Neoleukin Therapeutics
NLY	Annaly Capital Management
NMFC	New Mountain Finance
NMIH	Nmi Comm
NMM	Navios Maritime Partners L.P.
NMR	Nomura
NMRD	Nemaura Medical
NMRK	Newmark
NNBR	Nn
NNI	Nelnet
NNN	National Retail Properties
NNOX	Nano-X Imaging
NOA	North American Construction
NOBL	ProShares S&P 500 Aristocrats
NOC	Northrop Grumman
NODK	Ni
NOG	Northern Oil and Gas
NOK	Nokia
NOMD	Nomad Foods
NOV	National Oilwell Varco
NOVA	Sunnova Energy International
NOVT	Novanta
NOW	ServiceNow
NP	Neenah
NPK	National Presto
NPO	Enpro Industries
NPTN	Neophotonics
NR	Newpark Resources
NRBO	Neurobo Pharmaceuticals
NRC	National Research
NRDS	NerdWallet
NREF	NexPoint Real Estate Finance
NRG	NRG Energy
NRGU	MicroSectors US Big Oil Index 3X Leveraged
NRIM	Northrim Bancorp
NRIX	Nurix Therapeutics
NRZ	New Residential Investment
NS	NuStar Energy L.P.
NSA	National Storage Affiliates Trust
NSC	Norfolk Southern
NSIT	Insight Enterprises
NSP	Insperity
NSSC	Napco Security Systems
NSTG	Nanostring Technologies
NTAP	NetApp
NTB	The Bank Of N.T. Butterfield & Son
NTCO	Natura &Co S.A.
NTCT	Netscout Systems
NTES	NetEase
NTGR	Netgear
NTLA	Intellia Therapeutics
NTNX	Nutanix
NTP	Nam Tai Property
NTR	Nutrien
NTRA	Natera
NTRS	Northern Trust
NTST	NetSTREIT
NTSX	WisdomTree 90/60 U.S. Balanced Fund
NTUS	Natus Medical
NUAN	Nuance Communications
NUE	Nucor
NUGO	Nuveen Growth Opportunities ETF
NUGT	DIREXION DAILY GOLD MINERS BULL
NUS	Nu Skin Enterprises
NUSI	Nationwide Risk-Managed ome ETF
NUV	Nuveen Municipal Value Fund
NUVA	NuVasive
NUVB	Nuvation Bio
NUVL	Nuvalent
NVAX	Novavax
NVCR	NovoCure
NVDA	Nvidia
NVEC	Nve
NVEE	Nv5 Global
NVG	Nuveen AMT-Free Municipal Credit ome Fund
NVGS	Navigator s
NVMI	Nova
NVR	NVR
NVRO	Nevro
NVST	Envista
NVT	nVent Electric
NVTS	Navitas Semiconductor
NWBI	Northwest Bancshares
NWE	NorthWestern
NWFL	Norwood Financial
NWL	Newell Brands
NWLI	National Western Life
NWN	Northwest Natural Co
NWPX	Northwest Pipe Co
NWS	News Class B
NWSA	News Class A
NX	Quanex Building Products
NXE	NexGen Energy
NXGN	Nextgen Healthcare
NXPI	NXP Semiconductors NV
NXRT	Nexpoint Residential Trust In
NXST	Nexstar Media
NXTC	Nextcure
NXTG	First Trust Indxx NextG ETF
NYCB	New York Community Bancorp
NYMT	New York Mortgage Trust
NYMX	Nymox Pharmaceutical
NYT	NEW YORK TIMES (THE)
NYXH	Nyxoah
O	Realty ome
OAS	Oasis Petroleum
OB	Outbrain
OBNK	Origin Bancorp
OC	Owens Corning
OCDX	Ortho Clinical Diagnostics
OCFC	Oceanfirst Financial
OCFT	OneConnect Financial Technology
OCG	Oriental Culture
OCGN	Ocugen
OCSL	Oaktree Specialty Lending
OCUL	Ocular Therapeutix Commo
OCX	Oncocyte
ODC	Oil-dri Of America
ODFL	Old Dominion Freight Line
ODP	OFFICE DEPOT
OEC	Orion Engineered Carbons S.A.
OEF	ISHARES S&P 100 INDEX FUND
OESX	Orion Energy Systems
OFC	Corporate Office Properties Trust
OFED	Oconee Federal Financial
OFG	Ofg Bancorp
OFIX	Orthofix International
OFLX	Omega Flex
OG	Onion Global
OGCP	Empire State Realty OP
OGE	OGE Energy
OGI	OrganiGram s
OGN	Organon
OGS	One Gas
OHI	Omega Healthcare Investors
OI	O-I GLASS
OIH	MARKET VECTORS OIL SERVICES ETF
OII	Oceaneering International
OIS	Oil States International
OKE	ONEOK
OKTA	Okta
OLED	Universal Display
OLK	Olink AB (publ)
OLLI	Ollie's Bargain Outlet
OLMA	Olema Pharmaceuticals
OLO	Olo
OLP	One Liberty Properties
OLPX	Olaplex
OM	Outset Medical
OMC	Omnicom
OMCL	Omnicell
OMER	Omeros
OMF	OneMain
OMFL	Invesco Russell 1000 Dynamic Multifactor ETF
OMFS	Invesco Russell 2000 Dynamic Multifactor ETF
OMGA	Omega Therapeutics
OMI	Owens & Minor
OMIC	Singular Genomics Systems
ON	ON Semiconductor
ONB	Old National Bancorp
ONCR	Oncorus
ONE	Bank One
ONEM	1Life Healthcare
ONEQ	Fidelity NASDAQ Composite Index Track
ONEW	Onewater Marine
ONLN	ProShares Online Retail ETF
ONON	On AG
ONTF	ON24 INC.
ONTO	Onto Innovation
OOMA	Ooma
OPBK	Op Bancorp
OPCH	Option Care Health
OPEN	Opendoor Technologies
OPI	Office Properties ome Trust Common Shares Of Be
OPK	Opko Health
OPRT	Oportun Financial
OPRX	Optimizerx
OPT	Opthea
OPTN	Optinose
OPY	Oppenheimer
OR	Osisko Gold Royalties
ORA	Ormat Technologies
ORAN	Orange S.A.
ORC	Orchid Island Capital
ORCC	Owl Rock Capital
ORCL	Oracle
ORGN	Origin Materials
ORGO	Organogenesis
ORGS	Orgenesis
ORI	Old Republic International
ORIC	Oric Pharmaceuticals
ORLA	Orla Mining
ORLY	O'Reilly Automotive
ORPH	Orphazyme A/S
ORRF	Orrstown Financial
OSBC	Old Second Bancorp
OSCR	Oscar Health
OSG	Overseas Shipholding
OSH	Oak Street Health
OSIS	Osi Systems
OSK	Oshkosh
OSPN	Onespan
OSTK	Overstock.Com
OSUR	Orasure Technologies
OSW	Onespaworld
OTEX	Open Text
OTIS	Otis Worldwide
OTLY	Oatly
OTTR	Otter Tail
OUNZ	Van Eck Merk Gold Trust
OUST	Ouster
OUT	Outfront Media (REIT)
OVBC	Ohio Valley Banc
OVID	Ovid Therapeutics
OVLY	Oak Valley Bancorp
OVV	Ovintiv
OXM	Oxford Industries
OXY	Occidental Petroleum
OYST	Oyster Point Pharma
OZK	Bank OZK
PAA	Plains All American Pipeline L.P.
PAAS	Pan American Silver
PACB	Pacific Biosciences
PACK	Ranpak
PACW	PacWest Bancorp
PAG	Penske Automotive
PAGS	PagSeguro Digital
PAHC	Phibro Animal Health
PALL	Aberdeen Standard Physical Palladium Shares ETF
PANW	Palo Alto Networks
PAPR	Innovator S&P 500 Power Buffer ETF
PAR	PAR Technology
PARR	Par Pacific
PASG	Passage Bio
PATK	Patrick Industries
PAVE	Global X US Infrastructure Development ETF
PAVM	Pavmed
PAX	Patria Investments
PAY	Paymentus
PAYA	Paya s
PAYC	Paycom
PAYO	Payoneer Global
PAYS	Paysign
PAYX	Paychex
PB	Prosperity Bancshares
PBA	Pembina Pipeline
PBCT	People's United Financial
PBD	Invesco Global Clean Energy ETF
PBE	Invesco Dynamic Biotechnology & Genome ETF
PBF	Pbf Energy
PBFS	Pioneer Bancorp
PBFX	PBF Logistics LP
PBH	Prestige Consumer Healthcare
PBI	PITNEY BOWES
PBIP	Prudential Bancorp Commo
PBJ	Invesco Dynamic Food & Beverage ETF
PBS	Invesco Dynamic Media ETF
PBTS	Powerbridge Technologies
PBW	Invesco WilderHill Clean Energy ETF
PBYI	Puma Biotechnology
PCAR	PACCAR
PCB	Pcb Bancorp
PCEF	Invesco CEF ome Composite ETF
PCG	PACIFIC GAS & ELECTRIC CO.
PCH	PotlatchDeltic
PCOR	Procore Technologies
PCRX	Pacira Biosciences
PCSB	Pcsb Financial
PCT	PureCycle Technologies
PCTI	Pctel
PCTY	Paylocity
PCVX	Vaxcyte
PCY	POWERSHARES EMERGING MARKETS SO
PCYG	Park City
PCYO	Pure Cycle
PD	PagerDuty
PDBC	Invesco Optimum Yield Diversified Commodity Strategy No K-1 ETF
PDCE	Pdc Energy
PDCO	PATTERSON COMPANIES
PDD	Pinduoduo
PDFS	Pdf Solutions
PDLB	Pdl Community Bancorp
PDM	Piedmont Office Realty Trust
PDN	Invesco FTSE RAFI Developed Markets ex-U.S. Small-Mid ETF
PDP	Invesco DWA Momentum ETF
PDS	Precision Drilling
PEAK	HEALTHPEAK PROPERTIES
PEB	Pebblebrook Hotel Trust
PEBK	Peoples Bancorp Of n
PEBO	Peoples Bancorp
PECO	Phillips Edison &
PEG	Public Serv. Enterprise
PEGA	Pegasystems
PEJ	Invesco Dynamic Leisure and Entertainment ETF
PEN	Penumbra
PENN	Penn National Gaming
PEP	PepsiCo
PERI	Perion Network
PETQ	Petiq
PETS	Petmed Express
PETV	PetVivo
PETZ	TDH
PEZ	Invesco DWA Consumer Cyclicals Momentum ETF
PFBC	Preferred Bank
PFC	Premier Financial
PFE	Pfizer
PFEB	Innovator S&P 500 Power Buffer ETF February Series
PFF	ISHARES S&P US PREFERRED STOCK
PFFA	Virtus InfraCap U.S. Preferred Stock ETF
PFFD	Global X U.S. Preferred ETF
PFFR	InfraCap REIT Preferred ETF
PFFV	Global X Variable Rate Preferred ETF
PFG	Principal Financial
PFGC	Performance Food
PFHD	Professional Hldg
PFIS	Peoples Financial Services Cor
PFS	Provident Financial
PFSI	PennyMac Financial Services
PFSW	Pfsweb
PFXF	VanEck Vectors Preferred Securities ex Financials ETF
PG	Procter & Gamble
PGC	Peapack Gladstone
PGEN	Precigen
PGF	Invesco Financial Preferred ETF
PGNY	Progyny
PGR	Progressive
PGRE	Paramount
PGTI	Pgt
PGX	POWERSHARES PREFERRED PORTFOLIO
PH	Parker-Hannifin
PHAR	Pharming N.V.
PHAS	Phasebio Pharmaceuticals
PHAT	Phathom Pharmaceuticals
PHG	Koninklijke Philips N.V.
PHM	Pulte Homes
PHO	Invesco Water Resources ETF
PHR	Phreesia
PHVS	Pharvaris N.V.
PHYS	Sprott Physical Gold Trust
PI	Impinj
PICK	iShares MSCI Global Select Metals & Mining Producers ETF
PID	Invesco International Dividend Achievers™ ETF
PII	Polaris
PINC	Premier
PINE	Alpine ome Property Trust
PING	Ping Identity
PINS	Pinterest
PIO	Invesco Global Water ETF
PIPR	Piper Sandler Companies
PIRS	Pieris Pharmaceuticals
PJAN	Innovator S&P 500 Power Buffer ETF
PJP	Invesco Dynamic Pharmaceuticals ETF
PJT	Pjt Partners
PJUL	Innovator S&P 500 Power Buffer ETF
PK	Park Hotels & Resorts
PKBK	Parke Bancorp
PKE	Park Aerospace
PKG	Packaging of America
PKI	PerkinElmer
PKOH	Park-ohio
PLAB	Photronics
PLAN	Anaplan
PLAY	Dave & Busters Entertainment
PLBC	Plumas Bancorp
PLBY	PLBY
PLCE	Childrens Place
PLD	Prologis
PLL	PIEDMONT LITHIUM LIMITED
PLMR	Palomar
PLNT	Planet Fitness
PLOW	Douglas Dynamics
PLPC	Preformed Line Products
PLRX	Pliant Therapeutics
PLSE	Pulse Biosciences
PLTK	Playtika
PLTR	Palantir Technologies
PLUG	Plug Power
PLUS	Eplus
PLXS	Plexus
PLYA	Playa Hotels & Resorts N.V.
PLYM	Plymouth Industrial Reit
PM	Philip Morris International
PMT	Pennymac Mortgage
PMVP	PMV Pharmaceuticals
PNC	PNC Financial Services
PNFP	Pinnacle Financial Partners
PNM	PNM Resources
PNR	Pentair
PNRG	Primeenergy Resources
PNTG	Pennant
PNTM	Pontem
PNW	Pinnacle West Capital
POCT	Innovator S&P 500 Power Buffer ETF
PODD	Insulet
POLY	Plantronics
POND	Angel Pond
POOL	Pool
POR	Portland General Electric
POSH	Poshmark
POST	Post
POW	Powered Brands
POWI	Power Integrations
POWL	Powell Industries
POWW	AMMO
PPA	Invesco Aerospace & Defense ETF
PPBI	Pacific Premier Bancorp
PPC	Pilgrim's Pride
PPG	PPG Industries
PPGH	Poema Global
PPL	PPL
PPLT	Aberdeen Standard Platinum Shares ETF
PPTY	PPTY - U.S. Diversified Real Estate ETF
PRA	Proassurance
PRAA	Pra
PRAX	Praxis Precision Medicines
PRCH	Porch
PRDO	Perdoceo Education
PREF	Principal Spectrum Preferred Securities Active ETF
PRF	Invesco FTSE RAFI US 1000 ETF
PRFT	Perficient
PRFX	PainReform
PRFZ	Invesco FTSE RAFI US 1500 Small-Mid ETF
PRG	PROG
PRGO	Perrigo
PRGS	Progress Software
PRI	Primerica
PRIM	Primoris Services
PRK	Park National
PRLB	Proto Labs
PRLD	Prelude Therapeutics
PRN	Invesco DWA Industrials Momentum ETF
PRNT	3D Printing ETF
PRO	Pros
PROG	Progenity
PROV	Provident Financial
PRPL	Purple Innovation
PRTA	Prothena Ordin
PRTH	Priority Technology
PRTK	Paratek Pharmaceuticals
PRTS	Carparts
PRTY	Party City Holdco
PRU	Prudential Financial
PRVA	Privia Health
PRVB	Provention Bio
PSA	Public Storage
PSB	Ps Business Parks
PSC	Principal U.S. Small Cap Multi-Factor ETF
PSCC	Invesco S&P SmallCap Consumer Staples ETF
PSCH	Invesco S&P SmallCap Health Care ETF
PSCI	Invesco S&P SmallCap Industrials ETF
PSCT	Invesco S&P SmallCap Information Technology ETF
PSFE	Paysafe
PSK	SPDR Wells Fargo Preferred Stock ETF
PSL	Invesco DWA Consumer Staples Momentum ETF
PSMT	Pricesmart
PSN	Parsons
PSNL	Personalis
PSO	Pearson
PSP	Invesco Global Listed Private Equity ETF
PSPC	Post Partnering
PSQ	PROSHARES SHORT QQQ
PSTG	Pure Storage
PSTH	Pershing Square Tontine s
PSTX	Poseida Therapeutics
PSX	Phillips 66
PSXP	Phillips 66 Partners LP
PTC	PTC
PTCT	PTC Therapeutics
PTEN	Patterson-uti Energy
PTF	Invesco DWA Technology Momentum ETF
PTGX	Protagonist Therapeutics
PTH	Invesco DWA Healthcare Momentum ETF
PTLC	Pacer Trendpilot US Large Cap ETF
PTLO	Portillo's Class A
PTMC	Pacer Trendpilot US Mid Cap ETF
PTNQ	Pacer Trendpilot 100 ETF
PTON	Peloton Interactive
PTRA	Proterra
PTSI	P.A.M. Transportation
PTVE	Pactiv Evergreen
PUBM	PubMatic
PUK	Prudential
PUMP	Propetro
PVBC	Provident Bancorp
PVG	Pretium Resources
PVH	PVH
PWFL	Powerfleet
PWOD	Penns Woods Bancorp
PWP	Perella Weinberg Partners
PWR	Quanta Services
PWSC	PowerSchool
PXD	Pioneer Natural Resources
PXF	Invesco FTSE RAFI Developed Markets ex-U.S. ETF
PXH	Invesco FTSE RAFI Emerging Markets ETF
PXLW	Pixelworks
PYCR	Paycor
PYPD	PolyPid
PYPL	PayPal
PYR	PyroGenesis Canada
PZA	Invesco National AMT-Free Municipal Bond ETF
PZN	Pzena Investment
PZZA	Papa John's International
QABA	First Trust NASDAQ ABA Community Bank Index
QAI	IQ Hedge Multi-Strategy Tracker ETF
QCLN	First Trust NASDAQ Clean Edge Green Energy Index Fund
QCOM	QUALCOMM
QCRH	Qcr
QDEL	Quidel
QDF	FlexShares Quality Dividend Index Fund
QED	IQ Hedge Event-Driven Tracker ETF
QEMM	SPDR MSCI Emerging Markets StrategicFactors ETF
QGEN	Qiagen NV
QH	Quhuo
QID	PROSHARES ULTRASHORT QQQ
QLD	PROSHARES ULTRA QQQ
QLS	IQ Hedge Long/Short Tracker ETF
QLTA	iShares Aaa – A Rated orate Bond ETF
QLYS	Qualys
QMCO	Quantum
QMN	IQ Hedge Market Neutral Tracker ETF
QNST	Quinstreet
QQEW	First Trust NASDAQ 100 Equal Weight
QQQ	POWERSHARES QQQ
QQQJ	Invesco NASDAQ Next Gen 100 ETF
QQQM	Invesco NASDAQ 100 ETF
QRTEA	Qurate Retail
QRTEB	Qurate Retail
QRVO	Qorvo
QS	QuantumScape
QSR	Restaurant Brands International
QTEC	First Trust NASDAQ 100 Technology
QTNT	Quotient r
QTRX	Quanterix
QTUM	Defiance Quantum ETF
QTWO	Q2
QUAD	Quad/graphics
QUAL	iShares Edge MSCI USA Quality Factor ETF
QUOT	Quotient Technology
QURE	uniQure N.V.
QYLD	Global X NASDAQ 100 Covered Call ETF
R	RYDER SYSTEM
RAAS	Cloopen
RACE	Ferrari N.V.
RAD	Rite Aid
RADI	Radius Global Infrastructure
RAIN	Rain Therapeutics
RAMP	LiveRamp
RANI	Rani Therapeutics
RAPT	Rapt Therapeutics
RARE	Ultragenyx Pharmaceutical
RBA	Ritchie Bros. Auctioneers orporated
RBB	Rbb Bancorp
RBBN	Ribbon Communication
RBCAA	Republic Bancorp
RBLX	Roblox
RBOT	Vicarious Surgical
RC	Ready Capital
RCAT	Red Cat
RCD	Invesco S&P 500® Equal Weight Consumer Discretionary ETF
RCI	Rogers Communications
RCII	Rent-a-center
RCKT	Rocket Pharmaceuticals
RCKY	Rocky Brands
RCL	Royal Caribbean Cruises
RCM	R1 Rcm
RCUS	Arcus Biosciences
RDFN	Redfin
RDN	Radian
RDNT	Radnet
RDUS	Radius Health o
RDVT	Red Violet
RDVY	First Trust NASDAQ Rising Dividend Achievers ETF
RDWR	Radware
RE	Everest Re
REAL	The Realreal
REET	iShares Global REIT ETF
REFR	Research Frontiers
REG	Regency Centers
REGI	Renewable Energy
REGN	Regeneron
RELY	Remitly Global
REM	iShares Mortgage Real Estate ETF
REMX	VanEck Vectors Rare Earth/Strategic Metals ETF
RENT	Rent the Runway
REPH	Recro Pharma
REPL	Replimune
REPX	Riley Exploration Permian
RES	Rpc
RESN	Resonant
RETA	Reata Pharmaceuticals
RETL	Direxion Daily Retail Bull 3X Shares
REV	Revlon
REVG	Rev
REW	ProShares Ultra Short Technology
REX	Rex American Resources
REXR	Rexford Industrial Realty
REYN	Reynolds Consumer Products
REZ	iShares Residential Real Estate ETF
REZI	Resideo Technologies
RF	Regions Financial
RFL	Rafael
RFP	Resolute Forest Products
RGA	Reinsurance
RGC	Regencell Bioscience
RGCO	Rgc Resources
RGEN	Repligen
RGI	Invesco S&P 500® Equal Weight Industrials ETF
RGLD	Royal Gold
RGNX	Regenxbio
RGP	Resources Connection
RGR	Sturm Ruger
RGS	Regis
RH	RH
RHI	Robert Half International
RHS	Invesco S&P 500® Equal Weight Consumer Staples ETF
RICK	Rci Hospitality
RIDE	Lordstown Motors
RIG	TRANSOCEAN LTD.
RIGL	Rigel Pharmaceuticals
RILY	B. Riley Financial
RING	iShares MSCI Global Gold Miners ETF
RIOT	Riot Blockchain
RIVN	Rivian Automotive
RJF	Raymond James Financial
RKLB	Rocket Lab USA
RKT	Rocket Cos
RL	Polo Ralph Lauren
RLAY	Relay Therapeutics
RLGT	Radiant Logistics
RLGY	Realogy
RLI	Rli
RLJ	Rlj Lodging
RLMD	Relmada Therapeutics
RLX	RLX Technology
RLY	SPDR SSgA Multi-Asset Real Return ETF
RLYB	Rallybio
RM	Regional Management
RMAX	Re/max
RMBI	Richmond Mutual Bancorporation
RMBS	Rambus
RMD	ResMed
RMNI	Rimini Street
RMO	Romeo Power
RMR	The Rmr
RMTI	Rockwell Medical
RNA	Avidity Biosciences
RNAZ	TransCode Therapeutics
RNG	RingCentral
RNLX	Renalytix AI
RNR	RenaissanceRe
RNST	Renasant
RNW	ReNew Energy Global Class A Ordinary Shares
RNXT	RenovoRx
ROAD	Construction Partners
ROBO	ROBO Global Robotics and Automation Index ETF
ROCC	Ranger Oil Class A
ROCK	Gibraltar Industries
ROG	Rogers
ROIC	Retail Opportunity
ROIV	Roivant Sciences Common Shares
ROK	Rockwell Automation
ROKU	Roku
ROL	Rollins
ROLL	Rbc Bearings
ROM	ProShares Ultra Technology
ROOT	Root Insurance
ROP	Roper Technologies
ROST	Ross Stores
ROVR	Rover
RPAR	RPAR Risk Parity ETF
RPAY	Repay
RPD	Rapid7
RPG	Invesco S&P 500® Pure Growth ETF
RPHM	Reneo Pharmaceuticals
RPID	Rapid Micro Biosystems
RPM	RPM International
RPRX	Royalty Pharma
RPT	Rpt Realty
RPTX	Repare Therapeutics
RPV	Invesco S&P 500® Pure Value ETF
RRBI	Red River Bancshares
RRC	RANGE RESOURCES CORPORATION
RRGB	Red Robin Gourmet Burgers
RRR	Red Rock Resorts
RRX	Regal Rexnord
RS	Reliance Steel
RSG	Republic Services
RSI	Rush Street Interactive
RSKD	Riskified
RSP	RYDEX S&P EQUAL WEIGHT ETF
RSX	MARKET VECTORS RUSSIA ETF
RSXJ	VanEck Vectors Russia Small-Cap ETF
RTH	VanEck Vectors Retail ETF
RTM	Invesco S&P 500® Equal Weight Materials ETF
RTX	Raytheon
RUBY	Rubius Therapeutics
RUN	Sunrun
RUSHA	Rush Enterprises
RUSHB	Rush Enterprises
RUSL	Direxion Daily Russia Bull 2x Shares
RUTH	Ruth's Hospitality
RVI	Retail Value
RVLV	Revolve
RVMD	Revolution Medicines
RVNC	Revance Therapeutics Com
RVP	Retractable Technologies
RVSB	Riverview Bancorp
RWJ	Invesco S&P SmallCap 600 Revenue ETF
RWM	PROSHARES SHORT RUSSELL2000
RWO	SPDR DJ Wilshire Global Real Estate ETF
RWR	SPDR DJ WILSHIRE REIT ETF
RWT	Redwood Trust
RWX	SPDR DJ WILSHIRE INTL REAL ESTA
RXDX	Prometheus Biosciences
RXI	iShares Global Consumer Discretionary ETF
RXRX	RECURSION PHARMACEUTICALS
RXST	RxSight
RXT	Rackspace Technologies
RY	Royal Bank of Canada
RYAAY	Ryanair
RYAM	Rayonier Advanced Materials In
RYAN	Ryan Specialty
RYF	Invesco S&P 500® Equal Weight Financials ETF
RYH	Invesco S&P 500® Equal Weight Health Care ETF
RYI	Ryerson
RYLD	Global X Russell 2000 Covered Call ETF
RYN	Rayonier
RYT	Invesco S&P 500® Equal Weight Technology ETF
RYTM	Rhythm Pharmaceuticals
RZG	Invesco S&P SmallCap 600® Pure Growth ETF
RZV	Invesco S&P SmallCap 600® Pure Value ETF
S	SPRINT CORPORATION
SA	Seabridge Gold
SABR	Sabre
SAFE	Safehold
SAFM	Sanderson Farms
SAFT	Safety Insurance
SAGE	Sage Therapeutics
SAH	Sonic Automotive
SAIA	Saia
SAIC	SCIENCE APPLICATIONS INTERNATIONAL CORPORATION
SAL	Salisbury Bancorp Common
SAM	Boston Beer
SAMG	Silvercrest Asset Management g
SAN	Banco Santander S.A.
SANA	Sana Biotechnology
SAND	Sandstorm Gold
SANM	Sanmina
SASR	Sandy Spring Bancorp
SATS	Echostar
SAVA	Cassava Sciences
SAVE	Spirit Airlines
SB	Safe Bulkers
SBAC	SBA Communications
SBCF	Seacoast Banking o
SBFG	Sb Financial Commo
SBGI	Sinclair Broadcast
SBH	Sally Beauty
SBII	Sandbridge X2
SBIO	ALPS Medical Breakthroughs ETF
SBLK	Star Bulk Carriers
SBNY	Signature Bank
SBRA	Sabra Health Care REIT
SBS	Companhia de Saneamento Basico do Estado de Sao Paulo - SABESP
SBSI	Southside Bancshares
SBSW	Sibanye Stillwater
SBT	Sterling Bancorp
SBTX	Silverback Therapeutics
SBUX	Starbucks
SCCO	Southern Copper
SCHA	Schwab U.S. Small-Cap ETF
SCHB	SCHW US BRD MKT ETF
SCHC	Schwab International Small-Cap Equity ETF
SCHE	Schwab Emerging Markets Equity ETF
SCHF	Schwab International Equity
SCHG	Schwab U.S. Large-Cap Growth ETF
SCHH	Schwab US REIT ETF
SCHK	Schwab 1000 Index ETF
SCHL	Scholastic
SCHM	Schwab US Mid-Cap ETF
SCHN	Schnitzer Steel
SCHO	Schwab Short-Term U.S. Treasury ETF
SCHP	Schwab U.S. TIPS ETF
SCHR	Schwab Intermediate-Term U.S. Treasury ETF
SCHV	Schwab U.S. Large-Cap Value ETF
SCHW	Charles Schwab
SCHZ	Schwab U.S. Aggregate Bond ETF
SCI	Service International
SCL	Stepan
SCO	PROSHARES ULTRASHORT DJ-AIG CRU
SCOR	Comscore
SCPH	Scpharmaceuticals
SCPL	SciPlay
SCPS	Scopus Biopharma
SCS	Steelcase
SCSC	Scansource
SCU	Sculptor Capital Management
SCVL	Shoe Carnival
SCVX	SCVX .
SCWX	Secureworks .
SCZ	iShares MSCI EAFE Small-Cap ETF
SD	SandRidge Energy
SDC	SmileDirectClub
SDGR	Schrodinger
SDOW	UltraPro Short Dow30
SDS	SDS DUMMY
SDY	SPDR S&P DIVIDEND ETF
SE	SEA LIMITED
SEAC	Seachange International
SEAS	Seaworld Entertainment
SEAT	Vivid Seats Class A
SEB	Seaboard
SEDG	SolarEdge Technologies
SEE	Sealed Air
SEER	Seer
SEIC	SEI Investments
SELB	Selecta Biosciences
SEM	Select Medical
SEMR	SEMrush
SENEA	Seneca Foods
SENS	Senseonics s
SERA	Sera Prognostics
SF	Stifel Financial
SFBS	Servisfirst Bancshares
SFE	Safeguard Scientifics
SFIX	Stitch Fix
SFL	Sfl
SFM	Sprouts Farmers Market
SFNC	Simmons First National
SFST	Southern First Bancshares
SFT	Shift Technologies
SGA	Saga Communications
SGC	Superior Uniform
SGEN	Seattle Genetics
SGFY	Signify Health
SGH	Smart Global
SGHT	Sight Sciences
SGMO	Sangamo Therapeutics
SGMS	Scientific Games
SGOL	Aberdeen Standard Physical Gold Shares ETF
SGRY	Surgery Partners
SGTX	Sigilon Therapeutics
SH	PROSHARES SHORT S&P500
SHAK	Shake Shack
SHBI	Shore Bancshares
SHC	Sotera Health Co
SHCR	Sharecare
SHEN	Shenandoah Telecommunicat
SHLS	Shoals Technologies
SHLX	Shell Midstream Partners L.P.
SHM	SPDR LEHMAN SHORT TERM MUNICIPA
SHOO	Steven Madden
SHOP	Shopify
SHV	ISHARES LEHMAN SHORT TREASURY B
SHW	Sherwin-Williams
SHY	iShares 1-3 Year Treasury Bond
SHYF	Shyft
SHYG	iShares 0-5 Year High Yield orate Bond ETF
SI	Silvergate Capital
SIBN	Si-bone
SID	Companhia Siderurgica Nacional
SIEB	Siebert Financial
SIEN	Sientra
SIG	SIGNET JEWELERS LIMITED
SIGA	Siga Technologies
SIGI	Selective Insurance
SIL	Global X Silver Miners ETF
SILJ	ETFMG Prime Junior Silver ETF
SILK	Silk Road Medical
SILV	Silvercrest Metals
SIRI	Sirius XM
SITC	SITE Centers
SITE	SiteOne Landscape Supply
SITM	Sitime .
SIVB	SVB Financial
SIVR	Aberdeen Standard Physical Silver Shares ETF
SIX	Six Flags Entertainment
SJI	South Jersey Industries
SJM	JM Smucker
SJNK	SPDR Bloomberg Barclays Short Term High Yield Bond
SJR	Shaw Communications
SJW	Sjw
SKIN	The Beauty Health Class A
SKLZ	Skillz
SKM	SK Telecom
SKT	Tanger Factory Outlet
SKX	Skechers U.S.A.
SKY	Skyline
SKYT	SkyWater Technology
SKYW	Skywest
SKYY	First Trust ISE Cloud Computing Index Fund
SLAB	Silicon Laboratories
SLAM	Slam .
SLB	Schlumberger
SLCA	Us Silica
SLDB	Solid Biosciences
SLF	Sun Life Financial
SLG	SL Green Realty
SLGN	Silgan
SLI	Standard Lithium
SLM	SLM CORPORATION
SLNO	Soleno Therapeutics
SLP	Simulations Plus
SLQD	iShares 0-5 Year Investment Grade orate Bond ETF
SLQT	SelectQuote
SLRC	SLR Investment
SLV	ISHARES SILVER TRUST
SLVM	Sylvamo
SLVP	iShares MSCI Global Silver Miners ETF
SLY	SPDR S&P 600 Small Cap ETF
SLYG	SPDR S&P 600 Small Cap Growth ETF
SLYV	SPDR S&P 600 Small Cap Value ETF
SM	Sm Energy
SMAR	Smartsheet
SMB	VanEck Vectors AMT-Free Short Municipal Index ETF
SMBC	Southern Missouri Bancorp
SMBK	Smartfinancial
SMCI	Super Micro Computer
SMCP	AlphaMark Actively Managed Small Cap ETF
SMDD	ProShares UltraPro Short MidCap400
SMDV	ProShares Russell 2000 Dividend Growers ETF
SMDY	Syntax Stratified MidCap ETF
SMED	Sharps Compliance
SMFG	Sumitomo Mitsui Financial
SMG	The Scotts Miracle-Gro
SMH	MARKET VECTORS SEMICONDUCTOR ETF
SMHI	Seacor Marine
SMLF	iShares Edge MSCI Multifactor USA Small-Cap ETF
SMLR	Semler Scientific
SMLV	SPDR SSGA US Small Cap Low Volatility Index ETF
SMMD	iShares Russell 2500 ETF
SMMF	Summit Financial
SMMT	Summit Therapeutics
SMMU	Pimco Short Term Municipal Bond Fund
SMMV	iShares Edge MSCI Min Vol USA Small-Cap ETF
SMN	ProShares Ultra Short Basic Materials
SMP	Standard Motor Products
SMPL	The Simply Good Foods
SMSI	Smith Micro Software
SMTC	Semtech
SMWB	Similarweb
SNA	Snap-on
SNAP	Snap
SNBR	Sleep Number
SNCR	Synchronoss
SNCY	Sun Country Airlines
SNDL	Sundial Growers
SNDR	Schneider National
SNDX	Syndax Pharmaceuticals
SNEX	StoneX
SNFCA	Security Natl Financial
SNN	Smith & Nephew
SNOW	Snowflake
SNPE	Xtrackers S&P 500 ESG ETF
SNPO	Snap One
SNPS	Synopsys
SNSE	Sensei Biotherapeutics
SNSR	Global X Funds Global X Internet of Things ETF
SNTG	Sentage
SNUG	Merlyn.AI Tactical Growth and ome ETF
SNV	Synovus Financial
SNX	SYNNEX
SO	Southern
SOFI	SoFi Technologies
SOI	Solaris Oilfield Infrastructure
SON	Sonoco Products
SONO	Sonos
SOPA	Society Pass
SOPH	SOPHiA GENETICS
SOXL	Direxion Daily Semiconductor Bull 3x Shares
SOXS	Direxion Daily Semiconductor Bear 3x Shares
SOXX	iShares PHLX Semiconductor ETF
SP	Sp Plus o
SPAB	SPDR Portfolio Aggregate Bond ETF
SPB	Spectrum Brands
SPCE	Virgin Galactic
SPDW	SPDR Portfolio World ex-US ETF
SPEM	SPDR Portfolio Emerging Markets ETF
SPFF	Global X SuperIncome Preferred ETF
SPFI	South Plains Financial
SPG	Simon Property
SPGI	S&P Global
SPH	Suburban Propane Partners
SPHB	Invesco S&P 500 High Beta ETF
SPHD	Invesco S&P 500® High Dividend Low Volatility ETF
SPHQ	Invesco S&P 500® Quality ETF
SPIB	SPDR Portfolio Intermediate Term orate Bond ETF
SPIP	SPDR Portfolio TIPS ETF
SPLB	SPDR Portfolio Long Term orate Bond ETF
SPLK	Splunk
SPLP	Steel Partners s L.P.
SPLV	POWERSHARES S&P 500 LOW VOLATIL
SPMB	SPDR Portfolio Mortgage Backed Bond ETF
SPMD	SPDR Portfolio S&P 400 Mid Cap ETF
SPNE	Seaspine
SPNS	Sapiens International
SPNT	SiriusPoint
SPOK	Spok o
SPOT	Spotify Technology S.A.
SPPI	Spectrum Pharmaceuticals
SPR	Spirit AeroSystems
SPRB	SPRUCE BIOSCIENCES
SPRO	Spero Therapeutics
SPSB	SPDR Portfolio Short Term orate Bond ETF
SPSC	Sps Commerce
SPSM	SPDR Portfolio S&P 600 Small Cap ETF
SPT	Sprout Social
SPTI	SPDR Portfolio Intermediate Term Treasury
SPTL	SPDR Portfolio Long Term Treasury ETF
SPTM	SPDR Portfolio S&P 1500 Composite Stock Market ETF
SPTN	Spartannash o
SPTS	SPDR Portfolio Short Term Treasury ETF
SPWH	Sportsmans Warehouse
SPWR	Sunpower
SPXC	Spx
SPXL	Direxion Daily S&P 500 Bull 3X
SPXS	Direxion Daily S&P 500 Bear 3X
SPXU	PROSHARES ULTRAPRO SHORT S&P500
SPY	SPDR S&P 500
SPYD	SPDR Portfolio S&P 500 High Dividend ETF
SPYG	SPDR Portfolio S&P 500 Growth ETF
SPYV	SPDR Portfolio S&P 500 Value ETF
SQ	Square
SQFT	Presidio Property Trust
SQL	SeqLL
SQM	Sociedad Quimica y Minera de Chile S.A.
SQQQ	PROSHARES ULTRAPRO SHORT QQQ
SQSP	Squarespace
SQZ	SQZ Biotechnologies Co
SR	Spire
SRC	Spirit Realty Capital
SRCE	1st Source
SRCL	STERICYCLE
SRDX	Surmodics
SRE	Sempra Energy
SRET	Global X SuperDividend REIT ETF
SREV	Servicesource
SRG	Seritage Growth Properties
SRI	Stoneridge
SRLN	SPDR Blackstone/GSO Senior Loan
SRNE	Sorrento Therapeutics c
SRPT	Sarepta Therapeutics
SRRK	Scholar Rock
SRT	Startek
SRTY	ProShareUltraPro Short Russell 2000
SRVR	Pacer Benchmark Data & Infrastructure Real Estate SCTR ETF
SSB	South State
SSBK	Southern States Bancshares
SSD	Simpson Manufacturing
SSL	Sasol
SSNC	SS&C Technologies
SSO	PROSHARES ULTRA S&P500
SSP	E. W. Scripps Co
SSRM	SSR Mining
SSTI	Shotspotter
SSTK	Shutterstock
SSYS	Stratasys (ISRAEL)
ST	Sensata Technologies
STAA	STAAR Surgical
STAG	STAG Industrial
STAR	Istar
STBA	S&t Bancorp
STC	Stewart Information Svcs
STE	STERIS
STEM	Stem
STEP	StepStone
STER	Sterling Check .
STIP	iShares 0-5 Year TIPS Bond ETF
STKL	SunOpta
STLA	Stellantis N.V.
STLD	Steel Dynamics
STM	STMicroelectronics N.V.
STNE	StoneCo
STNG	Scorpio Tankers
STOK	Stoke Therapeutics
STOR	STORE Capital
STPZ	PIMCO 1-5 Year U.S. TIPS Index Exchange-Traded Fund
STRA	Strayer Education
STRL	Sterling Construction Co
STRN	86260J102 NASDAQ PILLAR Stran &
STRO	Sutro Biopharma
STRS	Stratus Properties
STSA	Satsuma Pharmaceuticals
STT	State Street
STTK	Shattuck Labs
STVN	Stevanato
STWD	Starwood Property Trust
STX	Seagate Technology
STXB	Spirit Of Texas Bancshares
STXS	Stereotaxis
STZ	Constellation Brands
SU	Suncor Energy
SUB	iShares S&P Short Term National AMT-Free Bond ETF
SUI	Sun Communities
SUM	Summit Materials
SUMO	Sumo Logic
SUN	SUNOCO LP
SUPN	Supernus Pharmaceuticals
SUPV	Grupo Supervielle S.A.
SUSA	iShares MSCI USA ESG Select ETF
SUSL	iShares ESG MSCI USA Leaders ETF
SUZ	Suzano S.A.
SVC	Service Properties Trust
SVM	Silvercorp Metals
SVRA	Savara
SVXY	PROSHARES SHORT VIX SHORT TERM
SWAN	Amplify BlackSwan Growth & Treasury Core ETF
SWAV	Shockwave Medical
SWBI	Smith & Wesson Brands
SWCH	Switch
SWI	Solar Winds
SWIM	Latham
SWIR	Sierra Wireless
SWK	Stanley Black & Decker
SWKH	Swk
SWKS	Skyworks Solutions
SWM	Schweitzer-mauduit Intl
SWN	SOUTHWESTERN ENERGY
SWTX	Springworks Therapeutics
SWX	Southwest Gas
SXC	Suncoke Energy
SXI	Standex International
SXT	Sensient Technologies
SYBT	Stock Yards Bancorp Comm
SYF	Synchrony Financial
SYK	Stryker
SYNA	Synaptics
SYNH	Syneos Health
SYRS	Syros Pharmaceuticals
SYTA	Siyata Mobile
SYY	Sysco
SZK	ProShares Ultra Short Consumer Goods
T	AT&T
TA	TravelCenters of America
TAC	TransAlta
TACO	Levy Acquisition .
TAIL	Cambria Tail Risk ETF
TAL	TAL Education
TALO	Talos Energy
TALS	Talaris Therapeutics
TAN	Invesco Solar ETF
TAP	Molson Coors Brewing
TARA	Protara Therapeutics
TARO	Taro Pharmaceutical Industries
TARS	Tarsus Pharmaceuticals
TASK	TaskUs
TAST	Carrols Restaurant
TBBK	Bancorp
TBF	PROSHARES SHORT 20+ YEAR TREASU
TBI	Trueblue
TBK	Triumph Bancorp
TBNK	Territorial Bancorp
TBPH	Theravance Biopharma
TBT	PROSHARES ULTRASHORT LEHMAN 20+
TCBI	Texas Capital Bancshares
TCBK	Trico Bancshares
TCBX	Third Coast Bancshares
TCDA	Tricida
TCFC	The Community Financial or
TCI	Transcontinental Realty
TCMD	Tactile Systems Technology
TCOM	Trip.com
TCRR	Tcr2 Therapeutics
TCRX	TScan Therapeutics
TCS	The Container Store
TCX	Tucows o
TD	The Toronto-Dominion Bank
TDC	TERADATA CORPORATION
TDG	TransDigm
TDIV	First Trust NASDAQ Technology Dividend Index Fund
TDOC	Teladoc Health
TDS	Telephone and Data Systems
TDSB	Cabana Target Drawdown 7 ETF
TDSC	Cabana Target Drawdown 10 ETF
TDTT	FlexShares iBoxx 3-Year Target Duration TIPS Index Fund
TDUP	ThredUp
TDV	ProShares S&P Technology Dividend Aristocrats ETF
TDW	Tidewater
TDY	Teledyne Technologies
TEAM	Atlassian
TECH	Bio-Techne
TECK	Teck Resources
TECL	Direxion Daily Technology Bull 3X Shares
TECS	Direxion Daily Technology Bear 3X Shares
TEF	Telefonica S.A.
TEL	TE Connectivity
TELA	Tela Bio
TELL	Tellurian
TEN	Tenneco
TENB	Tenable
TER	TERADYNE
TERN	Terns Pharmaceuticals
TETC	Tech & Energy Transition
TEVA	Teva Pharmaceutical Industries
TEX	Terex
TFC	TRUIST FINANCIAL CORPORATION
TFI	SPDR Barclays Capital Municipal Bond ETF
TFII	TFI International
TFSL	Tfs Financial
TFX	Teleflex
TG	Tredegar
TGB	Taseko Mines
TGH	Textainer
TGI	Triumph
TGLS	Tecnoglass
TGNA	TEGNA
TGT	Target
TGTX	TG Therapeutics
TH	Target Hospitality .
THC	TENET HEALTHCARE CORPORATION
THCX	Cannabis ETF
THD	ISHARES MSCI THAILAND INDEX FUND
THFF	First Financial
THG	The Hanover Insurance
THNQ	ROBO Global Artificial Intelligence ETF
THO	Thor Industries
THR	Thermon
THRM	Gentherm
THRY	Thryv
THS	TreeHouse Foods
TIG	Trean Insurance
TIGO	Millicom International Cellular S.A.
TIGR	UP Fintech American Depositary Share representing fifteen Class A Ordinary Shares
TIL	Instil Bio
TILE	Interface
TILT	FlexShares Morningstar US Market Factor Tilt Index Fund
TIOA	Tio Tech A
TIP	ISHARES BARCLAYS TIPS BOND FUND
TIPT	Tiptree
TIPX	SPDR Bloomberg 1-10 Year TIPS ETF
TIRX	Tian Ruixiang
TISI	Team
TITN	Titan Machinery
TIXT	TELUS International (Cda)
TJX	TJX Companies
TKNO	Alpha Teknova
TKR	The Timken
TLH	iShares 10-20 Year Treasury Bond ETF
TLIS	Talis Biomedical
TLK	Perusahaan Perseroan (Persero) PT Telekomunikasi Indonesia Tbk
TLMD	SOC Telemed
TLRY	Tilray
TLS	TELOS CORP
TLT	ISHARES BARCLAYS 20+ YEAR TREAS BOND
TLTD	FlexShares Morningstar Developed Markets ex-US Market Factor Tilt Index ETF
TLYS	Tilly's
TMCI	Treace Medical Concepts
TMDX	Transmedics
TME	Tencent Music Entertainment
TMF	Direxion Daily 20+ Year Treasury Bull 3X Shares (based on the NYSE 20 Year Plus Treasury Bond Index;
TMHC	Taylor Morrison Home
TMO	Thermo Fisher Scientific
TMP	Tompkins Financial
TMST	Timkensteel
TMUS	T-MOBILE US
TMV	DIREXION DAILY 30-YR TREASURY B
TMX	Terminix Global s
TNA	DIREXION DAILY SMALL CAP BULL 3X SHARES
TNC	Tennant
TNDM	Tandem Diabetes Care
TNET	TriNet
TNL	Travel + Leisure
TNYA	Tenaya Therapeutics
TOL	Toll Brothers
TOST	Toast
TOTL	SPDR DoubleLine Total Return Tactical ETF
TOWN	Towne Bank
TPAY	Ecofin Digital Payments Infrastructure Fund
TPB	Turning Point Brands
TPC	Tutor Perini
TPGY	TPG Pace Beneficial Finance
TPH	Tri Pointe Homes
TPIC	Tpi Composites
TPL	Texas Pacific Land
TPR	Tapestry
TPTX	Turning Point Therapeutics
TPX	Tempur Sealy International
TPYP	Tortoise North American Pipeline Fund
TQQQ	PROSHARES ULTRAPRO QQQ
TR	Tootsie Roll Industries
TRC	Tejon Ranch
TREC	Trecora Resources
TREE	Lendingtree
TREX	Trex
TRGP	Targa Resources
TRHC	Tabula Rasa Healthcare
TRI	Thomson Reuters
TRIP	TripAdvisor
TRKA	Troika Media
TRMB	Trimble
TRMD	TORM
TRMK	Trustmark
TRN	Trinity Industries
TRNO	Terreno Realty
TRNS	Transcat
TROW	T. Rowe Price
TROX	Tronox
TRP	TC Energy
TRQ	Turquoise Hill Resources
TRS	Trimas
TRST	Trustco Bank n
TRTN	Triton International
TRTX	Tpg Re Finance Trust
TRU	TransUnion
TRUE	Truecar
TRUP	Trupanion
TRV	The Travelers Companies
TS	Tenaris S.A.
TSBK	Timberland Bancorp
TSC	Tristate Capital
TSCO	Tractor Supply
TSE	Trinseo Sa
TSEM	Tower Semiconductor
TSHA	Taysha Gene Therapies
TSLA	Tesla
TSLX	Sixth Street Specialty Lending
TSN	Tyson Foods
TSP	TuSimple
TSVT	2seventy bio
TT	Trane Technologies
TTC	The Toro
TTCF	Tattooed Chef
TTD	The Trade Desk
TTEC	Teletech
TTEK	Tetra Tech
TTGT	Techtarget
TTM	Tata Motors
TTMI	Ttm Technologies
TTWO	Take-Two Interactive
TU	TELUS
TUGC	TradeUP Global
TUP	Tupperware Brands
TUR	ISHARES MSCI TURKEY INVESTABLE
TUYA	Tuya
TV	Grupo Televisa S.A.B.
TVTX	Travere Therapeutics
TVTY	Tivity Health
TW	Tradeweb Markets
TWKS	Thoughtworks
TWLO	Twilio
TWM	PROSHARES ULTRASHORT RUSSELL200
TWNK	Hostess Brands
TWO	Two Harbors Investment
TWOA	two
TWST	Twist Bioscience
TWTR	Twitter
TX	Ternium S.A.
TXG	10x Genomics
TXMD	Therapeuticsmd
TXN	Texas Instruments
TXRH	Texas Roadhouse
TXT	Textron
TYL	Tyler Technologies
TYME	Tyme Technologies
TZA	DIREXION DAILY SMALL CAP BEAR 3X SHARES
U	Unity Software
UA	Under Armour Class C
UAA	Under Armour Class A
UAL	United Continental
UAN	CVR Partners
UAPR	Innovator S&P 500 Ultra Buffer ETF
UBA	Urstadt Biddle Properties
UBER	Uber Technologies
UBFO	United Security
UBP	Urstadt Biddle Properties
UBS	UBS AG
UBSI	United Bankshares
UBX	Unity Biotechnology
UCBI	United Community Banks
UCL	Unocal Co
UCO	PROSHARES ULTRA DJ-AIG CRUDE OIL
UCTT	Ultra Clean
UDMY	Udemy
UDOW	PROSHARES ULTRAPRO DOW30
UDR	UDR
UE	Urban Edge Properties
UEC	Uranium Energy
UEIC	Universal Electronics
UEVM	VictoryShares USAA MSCI Emerging Markets Value Momentum ETF
UFCS	United Fire
UFEB	Innovator S&P 500 Ultra Buffer ETF February Series
UFI	Unifi
UFPI	UFP Industries
UFPT	Ufp Technologies
UGE	ProShares Ultra Consumer Goods
UGI	UGI
UGP	Ultrapar Participacoes S.A.
UHAL	Amerco
UHS	Universal Health Services
UHT	Universal Health Realty
UI	Ubiquiti Networks
UIHC	United Insurance
UIS	Unisys
UJAN	Innovator S&P 500 Ultra Buffer ETF
UJUL	Innovator S&P 500 Ultra Buffer ETF
ULBI	Ultralife Batteries
ULCC	Frontier
ULH	Universal Logistics
ULST	SPDR SSgA Ultra Short Term Bond ETF of SSgA Active Trust
ULTA	Ulta Beauty
UMBF	UMB Financial
UMC	United Microelectronics
UMH	Umh Properties
UMPQ	Umpqua
UNCY	Unicycive Therapeutics
UNF	Unifirst
UNFI	United Natural Foods
UNG	UNITED STATES NATURAL GAS FUND
UNH	United Health
UNIT	Unit
UNM	Unum
UNP	Union Pacific
UNTY	Unity Bancorp
UNVR	Univar Solutions
UOCT	Innovator S&P 500 Ultra Buffer ETF
UP	Wheels Up Experience
UPC	Union Planters
UPH	UpHealth
UPLD	Upland Software
UPRO	PROSHARES ULTRAPRO S&P500
UPS	United Parcel Service
UPST	Upstart
UPWK	Upwork
URA	Global X Uranium ETF
URBN	URBAN OUTFITTERS
URGN	Urogen Pharma
URI	United Rentals
URNM	NorthShore Global Uranium Mining ETF
URTY	PROSHARES ULTRAPRO RUSSELL2000
USAC	USA Compression Partners
USB	U.S. Bancorp
USFD	US Foods
USHY	iShares Broad USD High Yield orate Bond
USIG	iShares Broad USD Investment Grade orate Bond ETF
USLM	United States Lime
USM	United States Cellular
USMC	Principal U.S. Mega-Cap ETF
USMV	iShares Edge MSCI Min Vol USA
USNA	Usana Health Sciences
USO	UNITED STATES OIL
USPH	U.S. Physical Therapy
USRT	iShares Core U.S. REIT ETF
USSG	MSCI USA ESG Leaders Equity ETF
USX	Us Xpress Enterprises
UTF	Cohen & Steers Infrastructure Fund
UTHR	United Therapeutics
UTI	Universal Technical Inst
UTL	Unitil
UTMD	Utah Medical Products
UTME	UTime
UTZ	Utz Brands
UUP	POWERSHARES DB USD INDEX BULLIS
UUUU	Energy Fuels
UVE	Universal Insurance
UVSP	Univest Of Pa
UVV	Universal
UVXY	PROSHARES TRUST ULTRA VIX SHORT
UWM	PROSHARES ULTRA RUSSELL2000
UWMC	UWM s
UYG	PROSHARES ULTRA FINANCIALS
V	Visa
VAC	Marriott Vacations Worldwide
VACC	Vaccitech
VAL	Ensco
VALN	Valneva
VALU	Value Line
VAMO	Cambria Value and Momentum ETF
VAPO	Vapotherm
VAW	Vanguard Materials ETF
VB	VANGUARD SMALL-CAP ETF
VBIV	Vbi Vaccines
VBK	Vanguard Small Cap Growth ETF
VBR	Vanguard Small Cap Value ETF
VBTX	Veritex Common
VC	Visteon
VCEL	Aastrom Biosciences Comm
VCIT	VANGUARD IT CORP BOND ETF
VCLT	Vanguard Long-term orate Bonds
VCR	Vanguard Consumer Discretionary ETF
VCSH	VANGUARD ST CORP BOND ETF
VCYT	Veracyte
VDC	Vanguard Consumer Staples ETF
VDE	Vanguard Energy ETF
VEA	VANGUARD EUROPE PACIFIC ETF
VEC	Vectrus
VECO	Veeco Instruments
VECT	VectivBio
VEEV	Veeva Systems
VEL	Velocity Financial
VEON	VEON
VERA	Vera Therapeutics
VERI	Veritone
VERO	Venus Concept
VERU	Veru
VERV	Verve Therapeutics
VERX	Vertex
VERY	Vericity
VET	Vermilion Energy
VEU	VANGUARD FTSE AW EX-US ETF
VFC	V.F.
VFF	Village Farms International
VFH	Vanguard Financials ETF
VFQY	Vanguard U.S. Quality Factor ETF
VG	Vonage
VGIT	Vanguard Intermediate-Term Treasury ETF
VGK	VANGUARD EUROPEAN ETF
VGLT	Vanguard Long-Term Treasury ETF
VGR	Vector
VGSH	Vanguard Short-Term Treasury ETF
VGT	VANGUARD INFORMATION TECH ETF
VHC	Virnetx
VHI	Valhi
VHT	Vanguard Healthcare ETF
VIAO	VIA optronics AG
VIAV	VIAVI SOLUTIONS
VICI	VICI Properties
VICR	Vicor
VIDI	Vident International Equity Fund
VIG	VANGUARD DIVIDEND APPREC ETF
VIGI	Vanguard International Dividend Appreciation Index Fund
VII	7GC &
VINP	Vinci Partners Investments
VIOG	Vanguard S&P Small-Cap 600 Growth ETF
VIOO	Vanguard S&P Small-Cap 600 ETF
VIOV	Vanguard S&P Small-Cap 600 Value ETF
VIPS	Vipshop
VIR	Vir Biotechnology
VIRI	Virios Therapeutics
VIRT	Virtu Financial
VIS	Vanguard Industrial ETF
VITL	Vital Farms
VIV	Telefonica Brasil S.A.
VIVO	Meridian Bioscience
VIXM	ProShares VIX Mid-Term Futures ETF
VIXY	PROSHARES TRUST VIX SHORT-TERM
VKTX	Viking Therapeutics
VLD	Velo3D
VLDR	Velodyne Lidar
VLGEA	Village Super Market
VLO	Valero Energy
VLON	Vallon Pharmaceuticals
VLTA	Volta
VLUE	iShares Edge MSCI USA Value Factor ETF
VLY	Valley National Bancorp
VMAR	Vision Marine Technologies
VMBS	Vanguard Mortgage-Backed Securities ETF
VMC	Vulcan Materials
VMD	Viemed Healthcare
VMEO	Vimeo
VMI	Valmont Industries
VMW	Vmware
VNDA	Vanda Pharmaceuticals
VNE	Veoneer
VNLA	Janus Henderson Short Duration ome ETF
VNO	Vornado Realty Trust
VNOM	Viper Energy Partners LP
VNQ	VANGUARD REIT ETF
VNQI	Vanguard Global ex-US Real Estate
VNRX	Volitionrx
VNT	Vontier
VNTR	Venator Materials
VO	VANGUARD MID-CAP ETF
VOD	Vodafone Plc
VOE	Vanguard Mid-Cap Value ETF
VONE	Vanguard Russell 1000 ETF
VONG	Vanguard Russell 1000 Growth ETF
VONV	Vanguard Russell 1000 Value ETF
VOO	VANGUARD S&P 500 ETF
VOOG	Vanguard S&P 500 Growth ETF
VOOV	Vanguard S&P 500 Value ETF
VOR	Vor Biopharma
VOT	Vanguard Mid-Cap Growth ETF
VOX	Vanguard Communication Services ETF
VOXX	Voxx International
VOYA	Voya Financial
VPG	Vishay Precision
VPL	VANGUARD PACIFIC ETF
VPOP	Simplify Volt Pop Culture Disruption ETF
VPU	Vanguard Utilities ETF
VRA	Vera Bradley
VRAI	Virtus Real Asset ome ETF
VRAR	Glimpse
VRAY	Viewray
VRCA	Verrica Pharmaceuticals
VREX	Varex Imaging
VRM	Vroom
VRNS	Varonis Systems
VRNT	Verint Systems
VRP	Invesco Variable Rate Preferred ETF
VRPX	Virpax Pharmaceuticals
VRRM	Verra Mobility
VRS	Verso
VRSK	Verisk Analytics
VRSN	Verisign
VRT	Vertiv
VRTS	Veritas Software
VRTV	Veritiv
VRTX	Vertex Pharmaceuticals
VSAT	Viasat
VSCO	Victoria's Secret
VSEC	Vse
VSGX	Vanguard ESG International Stock ETF
VSH	Vishay Intertechnology
VSS	Vanguard FTSE All-World ex-US Small-Cap ETF
VST	Vistra Energy
VSTA	Vasta Platform
VSTM	Verastem
VSTO	Vista Outdoor
VT	VANGUARD TOTAL WORLD STOCK ETF
VTEB	Vanguard Tax-Exempt Bond Index ETF
VTEX	VTEX
VTGN	VistaGen Therapeutics
VTI	VANGUARD TOTAL STOCK MARKET ETF
VTIP	Vanguard Short-Term Inflation-Protected Securities ETF
VTOL	Bristow
VTR	Ventas
VTRS	Viatris
VTRU	Vitru
VTV	VANGUARD VALUE ETF
VTVT	Vtv Therapeutics
VTWO	Vanguard Russell 2000 ETF
VTYX	Ventyx Biosciences
VUG	VANGUARD GROWTH ETF
VUSB	Vanguard Ultra-Short Bond ETF
VUZI	Vuzix
VV	VANGUARD LARGE-CAP ETF
VVI	Viad
VVNT	Vivint Smart Home
VVOS	Vivos Therapeutics
VVV	Valvoline
VWO	VANGUARD MSCI EMERGING MARKETS ETF
VWOB	Vanguard Emerging Markets Government Bond ETF
VXF	Vanguard Extended Market VIPERs ETF
VXRT	Vaxart
VXUS	Vanguard Total International Stock Index
VXX	IPATH S&P 500 VIX SHORT-TERM FUTURES
VXZ	iPath Series B S&P 500® VIX Mid-Term Futures ETN
VYGG	Vy Global Growth
VYGR	Voyager Therapeutics
VYM	VANGUARD HIGH DIVIDEND YLD ETF
VYMI	Vanguard International High Dividend Yield Index Fund
VZ	Verizon Communications
VZIO	Vizio
W	Wayfair
WAB	Wabtec
WABC	Westamerica Bancorp
WAFD	Washington Federal
WAL	Western Alliance Bancorporation
WASH	Washington Trust
WAT	Waters
WAVE	Eco Wave Power Global
WB	Weibo
WBA	Walgreens Boots Alliance
WBIF	WBI Bull|Bear Value 1000 ETF
WBIL	WBI Bull|Bear Quality 1000 ETF
WBS	Webster Financial
WBT	Welbilt
WBX	Wallbox N.V.
WCC	Wesco International
WCLD	WisdomTree Cloud Computing Fund
WCN	Waste Connections
WD	Walker & Dunlop
WDAY	Workday
WDC	Western Digital
WDFC	Wd-40
WDH	Waterdrop
WE	WeWork
WEBR	Weber
WEC	Wec Energy
WELL	Welltower
WEN	The Wendy's
WERN	Werner Enterprises
WES	Western Midstream Partners LP
WETF	Wisdomtree Investments
WEX	WEX
WEYS	Weyco
WFC	Wells Fargo
WFG	West Fraser Timber
WFH	Direxion Work From Home ETF
WFRD	Weatherford International
WGO	Winnebago Industries
WH	Wyndham Hotels & Resorts
WHD	Cactus
WHG	Westwood
WHR	Whirlpool
WILC	G. Willi-Food International
WIMI	WiMi Hologram Cloud
WINA	Winmark
WING	Wingstop
WIP	SPDR DB INTERNATIONAL GOVERNMENT
WIRE	Encore Wire
WISH	ContextLogic
WIT	Wipro
WIX	Wix.com
WK	Workiva
WKHS	Workhorse
WKME	WalkMe
WLDN	Willdan
WLFC	Willis Lease Finance
WLK	Westlake Chemical
WLKP	Westlake Chemical Partners LP
WLL	Whiting Petroleum
WM	Waste Management
WMB	Williams Cos.
WMC	Western Asset Mortgage
WMG	Warner Music
WMK	Weis Markets
WMS	Advanced Drainage Systems
WMT	Walmart
WNC	Wabash National
WNEB	Western New England Bancorp
WNS	WNS (Holdings)
WNW	Wunong Net Technology
WOLF	Wolfspeed
WOOF	Petco Health & Wellness
WOR	Worthington Industries
WOW	Wideopenwest
WPC	W. P. Carey
WPM	Wheaton Precious Metals
WPP	WPP
WPRT	Westport Fuel Systems
WRB	W. R. Berkley
WRBY	Warby Parker
WRE	Washington Real Estate Investment Trust
WRK	WestRock
WRLD	World Acceptance
WSBC	Wesbanco
WSBF	Waterstone Financial Com
WSC	Willscot Mobile Mini
WSFS	Wsfs Financial
WSM	Williams-Sonoma
WSO	Watsco
WSR	Whitestone Reit
WST	West Pharmaceutical Services
WTBA	West Bancorporation
WTFC	Wintrust Financial
WTI	W&t Offshore
WTM	White Mountains Insurance
WTMF	WisdomTree Managed Futures Strategy Fund
WTRG	Essential Utilities
WTRH	Waitr
WTS	Watts Water Technologies
WTTR	Select Energy Services
WU	Western Union Co
WUGI	Esoterica NextG Economy ETF
WVE	Wave Life Sciences
WW	Ww International
WWD	Woodward
WWE	World Wrestling Entertainment
WWW	Wolverine World Wide
WY	Weyerhaeuser
WYNN	Wynn Resorts
X	UNITED STATES STEEL CORPORATION
XAIR	Beyond Air
XAR	SPDR S&P Aerospace & Defense ETF
XBI	SPDR SERIES TRUST SPDR S&P BIO
XBIT	Xbiotech
XCUR	Exicure
XEL	Xcel Energy
XENE	Xenon Pharmaceuticals
XENT	Intersect Ent o
XERS	Xeris Pharmaceuticals
XES	SPDR S&P Oil & Gas Equipment & Services ETF
XFOR	X4 Pharmaceuticals
XGN	Exagen
XHB	SPDR SERIES TRUST SPDR HOMEBUIL
XHE	SPDR S&P Health Care Equipment ETF
XHR	Xenia Hotels & Resorts
XITK	SPDR FactSet Innovative Technology ETF
XL	XL Fleet
XLB	MATERIALS SELECT SECTOR SPDR
XLC	Communication Services Select Sector SPDR Fund
XLE	ENERGY SELECT SECTOR SPDR
XLF	FINANCIAL SELECT SECTOR SPDR
XLG	Invesco S&P 500® Top 50 ETF
XLI	INDUSTRIAL SELECT SECTOR SPDR
XLK	TECHNOLOGY SELECT SECTOR SPDR
XLP	CONSUMER STAPLES SELECT SECTOR SPDR
XLRE	Real Estate Select Sector SPDR Fund
XLU	UTILITIES SELECT SECTOR SPDR
XLV	HEALTH CARE SELECT SECTOR SPDR
XLY	CONSUMER DISCRET SELECT SECTOR SPDR
XM	Qualtrics International
XME	SPDR S&P METALS & MINING
XMLV	Invesco S&P MidCap Low Volatility ETF
XMTR	Xometry
XNCR	Xencor
XNTK	NYSE Technology ETF
XOM	Exxon Mobil
XOMA	Xoma
XOP	SPDR S&P OIL&GAS EXPLORATION & PROD
XOS	Xos
XP	XP
XPEL	Xpel
XPER	Xperi
XPEV	XPENG INC.
XPH	SPDR S&P Pharmaceuticals ETF
XPO	XPO Logistics
XPOF	Xponential Fitness
XRAY	Dentsply Sirona
XRT	SPDR S&P RETAIL
XRX	Xerox
XSD	SPDR S&P Semiconductor ETF
XSLV	Invesco S&P SmallCap Low Volatility ETF
XSOE	WisdomTree Emerging Markets ex-State-Owned Enterprises Fund
XT	iShares Exponential Technologies ETF
XTL	SPDR S&P Telecom ETF
XTN	SPDR S&P Transportation ETF
XYL	Xylem
YALA	Yalla
YANG	Direxion Daily FTSE China Bear 3X Shares
YCS	PROSHARES ULTRASHORT YEN NEW
YELP	Yelp
YETI	YETI
YEXT	Yext
YGMZ	MingZhu Logistics
YINN	Direxion Daily FTSE China Bull 3X Shares
YLD	Principal Active ome ETF
YMAB	Y-mabs Therapeutics
YMM	Full Truck Alliance
YOLO	AdvisorShares Pure Cannabis ETF
YORW	York Water
YOU	Clear Secure
YPF	YPF Sociedad Anonima
YQ	17 Education & Technology
YSG	Yatsen
YTRA	Yatra Online
YUM	Yum! Brands
YUMC	Yum China
YY	JOYY
YYY	Amplify High ome ETF
Z	Zillow
ZBH	Zimmer Biomet
ZBRA	Zebra Technologies
ZCMD	Zhongchao
ZD	Ziff Davis
ZEN	Zendesk
ZENV	Zenvia
ZETA	Zeta Global
ZEUS	Olympic Steel
ZG	Zillow
ZGNX	Zogenix
ZH	Zhihu
ZI	ZoomInfo Technologies
ZIM	ZIM Integrated Shipping Services
ZION	Zions Bancorp
ZIP	ZipRecruiter
ZLAB	Zai Lab
ZM	Zoom
ZME	Zhangmen Education
ZNGA	Zynga
ZNTL	Zentalis Pharmaceuticals
ZOM	Zomedica
ZROZ	PIMCO 25+ Year Zero Coupon US Treasury
ZS	Zscaler
ZSL	PROSHARES ULTRASHORT SILVER
ZTO	ZTO Express (Cayman)
ZTS	Zoetis
ZUMZ	Zumiez
ZUO	Zuora
ZVIA	Zevia
ZWS	Zurn Water Solutions
ZYME	Zymeworks
ZYXI	Zynex
'''