'''
Change History
2022/12/13 - use the revised test_rl_macd_v2, default params to (3,7,19), 0, 0

Change History
2023/03/03  add cwh check
'''
from datetime import datetime
from datetime import timedelta

import json
import numpy as np
import traceback
from test_mongo import MongoExplorer
import pandas as pd
from pandas import json_normalize
from test_yahoo import QuoteExplorer
from test_g20_v2 import StockScore
from test_stockstats import StockStats
from openpyxl import load_workbook
from pathlib import Path
from copy import copy
from typing import Union, Optional
#import numpy as np
#import pandas as pd
import openpyxl
#from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
#import time
from test_cup_with_handle import Rule_Cup_with_Handle

class RLDailyTrader:
    def __init__(self, options):
        self.mongo = MongoExplorer()
        self.runDate = options['AAOD']
        self.aaod = options['AAOD']
        #self.symbol = options['symbol']
        #self.long = options['long']
        #self.short = options['short']
        #self.signal = options['signal']
        #self.macd_threshold = options['macd_threshold']
        #self.macd_min_len = options['macd_min_len']
        #self.today = datetime.now()

        self.list = ['MSFT', 'SPY', 'SHOP', 'TWLO', 'ROKU', 'OKTA', 'PAYC', 'GOOGL', 'LYFT', 'TTD', 'AYX', 'ISRG', 'CRM',
                     'ATVI', 'MDB', 'SQ', 'BABA', 'AAPL', 'TCEHY', 'META', 'AMZN', 'TSLA', 'ELV', 'COST', 'GS',
                     'GE', 'DELL', 'AZO', 'POOL', 'NOW', 'BKNG', 'ORLY', 'ALGN', 'HUM', 'LIN', 'PCTY',
                     'ETSY', 'DECK', 'FIVE', 'NEO', 'PODD', 'TXG', 'CABO', 'NET', 'TW', 'SNOW', 'CHK',
                     'ASML', 'TEAM', 'DDOG', 'PANW', 'CZR', 'CTSH', 'MTCH', 'SBSW', 'JD', 'IQV',
                     'OPEN', 'UPWK', 'EXPI', 'JBHT', 'CLH', 'TYL', 'MSCI', 'AXON', 'FND', 'GNRC', 'PEN',
                     'TPL', 'FANG', 'WK', 'CDAY', 'SMAR', 'WMG', 'AMBA', 'GLOB', 'EPAM', 'CACC', 'REXR', 'EXPO',
                     'ACHC', 'ANSS', 'ATKR', 'FOXF', 'IIPR', 'LKQ', 'QLYS', 'RCM', 'WDAY', 'ULTA', 'DT',
                     'MEDP', 'DPZ', 'TNET', 'UTHR', 'DKS','RPD', 'ASAN', 'SAIA', 'TREX', 'BJ',
                     'CYBR', 'ADBE', 'ESI', 'HLI', 'CHPT', 'KRNT', 'KEYS', 'ARES', 'UPST', 'ONTO', 'AMT', 'SBAC',
                     'MUSA', 'LNG', 'VAC', 'PLNT', 'AIRC', 'ALGM', 'HII', 'WSC', 'TDG', 'BYDDF', 'ABBV',
                     'NVO', 'CHK', 'UNH', 'REGN', 'EXR', 'ENPH', 'ELV', 'BYDDY', 'SCI', 'ARGX', 'BAH', 'PWR',
                     'KNSL', 'AZPN', 'DTM', 'ALNY', 'YZCAY', 'ON', 'WMB', 'RSG', 'WCN', 'ADP', 'CVS',
                     'VRTX', 'DOX', 'LLY', 'HUBB', 'FWONK', 'ALB', 'FSLR', 'GGB', 'RRX', 'HOMB', 'SYM',
                     'OZK', 'TECK', 'VLO', 'NUE', 'PXD', 'STLD', 'GDDY', 'NXST', 'BIIB', 'BMRN', 'HES',
                     'IRDM', 'WRB', 'GL', 'CQP', 'AMN', 'ELP', 'GLPI', 'MCK', 'MOH', 'CB', 'CW', 'GD', 'DUK',
                     'EIX', 'GMED', 'H', 'IRTC', 'LKFN', 'MMSI', 'NJR', 'ORA', 'PRTA', 'PTC', 'RHP', 'TH',
                     'URI', 'VIPS', 'VST', 'WING', 'XENE', 'FMX', 'PDD', 'FIZZ', 'RRC', 'RYI', 'PECO', 'PWSC', 'PXD',
                     'AEIS', 'WHD', 'INSW', 'OKE', 'RES', 'TRGP', 'A', 'AIN', 'AME', 'APOG', 'ANET', 'CWST',
                     'CAT', 'CDW', 'GIB', 'CHDN', 'CI', 'CR', 'DRI', 'EMR', 'FCFS', 'FLS', 'HSIC', 'HXL', 'INST',
                     'ITCI', 'OHI', 'PAYX', 'RBC', 'SPXC', 'VICI', 'WAB', 'WEX', 'DINO', 'NFE', 'BIPC', 'WFRD',
                     'ASIX', 'CRC', 'CVE', 'CSWI', 'ESTE', 'ENS', 'EQNR', 'OXY', 'PARR', 'SNDX', 'WLK', 'DEN',
                     'LNW', 'CHRD', 'DECK', 'ACHC', 'ALNY', 'ALB', 'FSLR', 'TECK', 'BMRN', 'GATX',
                     'GEF', 'KRYS', 'MLI', 'NE', 'OGE', 'SYBT', 'TGLS', 'TKR', 'VAL', 'TCOM', 'ASO', 'DG',
                     'JAZZ', 'JKS', 'MGY', 'MTDR', 'PDCE', 'DCP', 'SMCI', 'AES', 'AGYS', 'BMI', 'CALX',
                     'CASY', 'CBZ', 'WIRE', 'NSP', 'IHG', 'IRM', 'MCD', 'NXGN', 'TRTN', 'XLU', 'WEC', 'XEL', 'APD',
                     'ARLP', 'CCEP', 'CTRA', 'PSX', 'HWM', 'NVO', 'GLPI', 'RETA', 'TXRH', 'DKL', 'LOPE', 'USAC',
                     'CVX', 'SUN', 'SUPN', 'MGRC', 'HPK', 'ASAI', 'BOWL', 'SUBCY', 'ET', 'HQY', 'HUBG',
                     'LNTH', 'LMT', 'SPTN', 'SNEX', 'AMR', 'AFG', 'ACGL', 'CEIX', 'DMLP', 'ERIE', 'RLI', 'SAFT',
                     'SIGI', 'HCC', 'CBOE', 'PI', 'IPAR', 'LPLA', 'PJT', 'AGCO', 'ALG', 'NGLOY', 'NLY', 'AIT',
                     'BANF', 'DDS', 'EME', 'EXLS', 'FN', 'FIBK', 'PAC', 'ASR', 'IDA', 'IMKTA', 'ITA', 'KR',
                     'LNN', 'MSM', 'NOC', 'NVT', 'ROL', 'SJW', 'TEX', 'TRN', 'VMI', 'WCC', 'ATI', 'AMPH',
                     'APLS', 'ARCH', 'ASH', 'CF', 'CFR', 'THG', 'NGVT', 'IOSP', 'MTRN', 'RS', 'SCCO', 'RYAN',
                     'AXON', 'RJF', 'OMAB', 'FICO', 'HAE', 'OGS', 'ACA', 'ABG', 'BKE', 'CNP', 'CMS', 'FIX',
                     'JKHY', 'NNN', 'NVEE', 'OXM', 'SLGN', 'TTC', 'TWO', 'GWW', 'FMC', 'GLP', 'LXU', 'TAL',
                     'AIR', 'AMAL', 'BDC', 'ELF', 'GPC', 'IEP', 'EDU', 'SAIC', 'SRE', 'VC', 'RTX', 'IBN', 'MGPI',
                     'RPM', 'ADC', 'HEI', 'MRTN', 'PCAR', 'CRS', 'ACT', 'NUE', 'AEP', 'CLFD', 'IT', 'HOLX', 'NSSC',
                     'PAG', 'SPLP', 'TJX', 'PSN', 'FOUR', 'AON', 'UTZ', 'RSG', 'CNXN', 'BDX', 'BWXT', 'CTAS',
                     'DGII', 'GILD', 'MCRI', 'TRV', 'ANCTF', 'ATKR', 'LLY', 'GL', 'PBT', 'ADUS', 'IEX', 'LTC',
                     'ROST', 'AFL', 'MPC', 'SBR', 'CMI', 'HTHT', 'ENSG', 'UHS', 'MURGY', 'NI', 'WPM', 'SLVM', 'GGB',
                     'AVA', 'INCY', 'NVCR', 'COP', 'CSGP', 'DCI', 'PFGC', 'TWI', 'UOVEY', 'WPC', 'GMAB', 'MDRX', 'NDSN',
                     'SPSC', 'ABC', 'EDPFY', 'HALO', 'VRTV', 'GEO', 'UNF', 'AXSM', 'YUM', 'GES', 'YUMC', 'AMGN', 'OLN',
                     'EPD', 'KLAC', 'MMP', 'BURL', 'CWT', 'CSL', 'DLTR', 'DY', 'MA', 'MSI', 'NTCT', 'NEE', 'ORCL', 'SNA',
                     'SBUX', 'WYNN', 'XYL', 'AIG', 'CCI', 'CMC', 'RIO', 'STLD', 'AMP', 'ASX', 'LSCC', 'PLXS', 'COLD',
                     'APPF', 'BRKR', 'CROX', 'DSGX', 'AQUA', 'FCPT', 'GMS', 'HCA', 'NSIT', 'IBB', 'LMAT', 'NEP', 'OLLI',
                     'PTCT', 'TPR', 'RL', 'GEL', 'AIMC', 'EWZ', 'KNX', 'FWONA', 'MTD', 'COOP', 'ODFL', 'WAT',
                     'XPEL', 'MEG', 'CCJ', 'MEOH', 'PRI', 'PLMR', 'CEG', 'ADI', 'KLIC', 'MTSI', 'MCHP', 'MPWR', 'OCSL',
                     'TXN', 'ATRI', 'BCC', 'CRHKY', 'CCOI', 'CTS', 'DXCM', 'DFIN', 'FELE', 'LSTR', 'MBUU', 'O', 'RBA',
                     'SBRA', 'SGEN', 'SKX', 'SLYV', 'MDY', 'VHT', 'WM', 'PII', 'AVGO', 'FRHC', 'CAR', 'CACI', 'ETN', 'NPO',
                     'FDS', 'PPA', 'ROP', 'STRL', 'TMO', 'OWL', 'MKTX', 'POWI', 'ROAD', 'DHI', 'PLUS', 'GPI', 'ICFI',
                     'MEI', 'PRIM', 'R', 'TDY', 'XLE', 'FCX', 'PEP', 'HPE', 'ABB', 'AMX', 'EQIX', 'IBP', 'IYH',
                     'LECO', 'NRC', 'PKG', 'DGX', 'TSCO', 'VRSN', 'WSO', 'ZTO', 'UI', 'VOYA', 'CRUS', 'GOLF', 'ASND',
                     'EPRT', 'HGV', 'ITW', 'PFSI', 'SXI', 'WWD', 'IIJIY', 'PHR', 'WTRG', 'QSR', 'TPX', 'ASMIY', 'PLL',
                     'ATSG', 'NEU', 'CRDO', 'SPT', 'DORM', 'APG', 'HOMB', 'RWWI', 'DHCC', 'IE', 'ACDC', 'GPLB', 'AAMC',
                     'USLM', 'KNIT', 'DPSI', 'CHKR', 'UFPT', 'LNKB', 'VWFB', 'HLNE', 'OTIS', 'SSBK', 'BBSI', 'EVI', 'FC',
                     'FUTU', 'KRUS', 'KRMD', 'RGEN', 'TCBC', 'XPOF', 'JXN', 'ZGN', 'CCSI', 'LCFY', 'ATAT', 'PLTNU', 'RICK',
                     'IBEX', 'ESQ', 'TRNS', 'WIX', 'APHE', 'LKST', 'NVDA', 'BLD', 'RXMD', 'ALTR', 'CLMB', 'BWMN',
                     'SVMB', 'VCTR', 'AERG', 'AURX', 'ERLFF', 'ESAB', 'MLTX', 'CMG', 'GWOX', 'MMS', 'MXF', 'VSEC', 'FLUX',
                     'RCL', 'SAP', 'SIEGY', 'SYK', 'RRX', 'AAON', 'AYI', 'PLOW', 'HEES', 'JBT', 'NVEC', 'ELA', 'ANIK', 'BLX',
                     'JPM', 'MOS', 'GBFH', 'CCRD', 'HVRRY', 'EVR', 'EEFT', 'FL', 'ROK', 'BESIY', 'IFNNY', 'IDCC', 'FAST',
                     'EWD', 'ADRNY', 'KTB', 'BA', 'VATE', 'GFS', 'ITT', 'SDXAY', 'CASH', 'ARW', 'V', 'MBLY', 'HSTM',
                     'LVMUY', 'FSBW', 'NTES', 'URBN', 'EADSY', 'CDNS', 'CASS', 'LFUS', 'WT', 'IR', 'ARCAY', 'LRCX', 'FTNT',
                     'MANH', 'SCTH', 'DO', 'ARNC', 'NVR', 'INTA', 'BWA', 'TJBH', 'AGM', 'MTZ', 'COCO', 'CARR', 'PH',
                     'FTCO', 'GSHD', 'HUBS', 'LMPX', 'MRTI', 'NGLD', 'DKNG', 'NFLX', 'UHAL', 'CTVA', 'KIDS', 'RADI', 'RIOT',
                     'AMRX', 'IDXX', 'INSP', 'WST', 'CPRT', 'RMD']

        #self.list = ['SHOP', 'COUP', 'TWLO', 'ROKU', 'OKTA', 'PAYC', 'GOOGL', 'LYFT', 'TTD', 'AYX', 'ISRG', 'CRM',
        #             'ATVI', 'MDB', 'SQ', 'BABA', 'AAPL', 'TCEHY', 'FB', 'AMZN', 'TSLA', 'ANTM', 'COST', 'GS',
        #             'GE', 'CDLX', 'APPN', 'RNG', 'DELL', 'AZO', 'CHTR', 'FTNT', 'POOL', 'NOW', 'BKNG', 'ORLY',
        #             'ALGN', 'HUM', 'IAC', 'LIN', 'SNAP', 'PCTY', 'IDXX', 'NVCR', 'ETSY', 'EW', 'CNC', 'MASI', 'CMG',
        #             'TTNDY', 'ILMN', 'DECK', 'FIVE', 'ANET', 'HUBS', 'EQIX', 'MPWR']
        #   'SPSC', 'CHTR', 'UI', 'CTLT', 'XLRN','SPPT','MIME', 'MLYBY','SWCH', 'SRD','COUP','MYOV','EVOP','STOR', 'SJI',
        #self.list = ['BKNG'] 'WIX', 'BILI','MA', 'HCA', 'PRAH', 'NFLX', 'ZTS','VEEV', 'PYPL', 'AVGO', 'ZM', 'CRWD', 'CVNA', 'DOCU', 'PINS', 'V'
        self.exclusionList = ['PENN', 'JCOM', 'GKOS', 'EYE', 'BURL', 'W', 'GOLF', 'DG', 'TRU', 'SSNC', 'TSCO', 'FLT', 'PDD',
                              'ZS', 'ROST','DXCM', 'NTES', 'NIO', 'LI', 'ODFL', 'XPO', 'SAM','FRC']

    def copy_excel_cell_range(
            self,
            src_ws: openpyxl.worksheet.worksheet.Worksheet,
            min_row: int = None,
            max_row: int = None,
            min_col: int = None,
            max_col: int = None,
            tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
            tgt_min_row: int = 1,
            tgt_min_col: int = 1,
            with_style: bool = True
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        copies all cells from the source worksheet [src_ws] starting from [min_row] row
        and [min_col] column up to [max_row] row and [max_col] column
        to target worksheet [tgt_ws] starting from [tgt_min_row] row
        and [tgt_min_col] column.

        @param src_ws:  source worksheet
        @param min_row: smallest row index in the source worksheet (1-based index)
        @param max_row: largest row index in the source worksheet (1-based index)
        @param min_col: smallest column index in the source worksheet (1-based index)
        @param max_col: largest column index in the source worksheet (1-based index)
        @param tgt_ws:  target worksheet.
                        If None, then the copy will be done to the same (source) worksheet.
        @param tgt_min_row: target row index (1-based index)
        @param tgt_min_col: target column index (1-based index)
        @param with_style:  whether to copy cell style. Default: True

        @return: target worksheet object
        """
        if tgt_ws is None:
            tgt_ws = src_ws

        # https://stackoverflow.com/a/34838233/5741205
        for row in src_ws.iter_rows(min_row=min_row, max_row=max_row,
                                    min_col=min_col, max_col=max_col):
            for cell in row:
                tgt_cell = tgt_ws.cell(
                    row=cell.row + tgt_min_row - 1,
                    column=cell.col_idx + tgt_min_col - 1,
                    value=cell.value
                )
                if with_style and cell.has_style:
                    # tgt_cell._style = copy(cell._style)
                    tgt_cell.font = copy(cell.font)
                    tgt_cell.border = copy(cell.border)
                    tgt_cell.fill = copy(cell.fill)
                    tgt_cell.number_format = copy(cell.number_format)
                    tgt_cell.protection = copy(cell.protection)
                    tgt_cell.alignment = copy(cell.alignment)
        return tgt_ws


    def append_df_to_excel(
            self,
            filename: Union[str, Path],
            df: pd.DataFrame,
            sheet_name: str = 'Sheet1',
            startrow: Optional[int] = None,
            max_col_width: int = 30,
            autofilter: bool = False,
            fmt_int: str = "#,##0",
            fmt_float: str = "#,##0.00",
            fmt_date: str = "yyyy-mm-dd",
            fmt_datetime: str = "yyyy-mm-dd hh:mm",
            truncate_sheet: bool = False,
            storage_options: Optional[dict] = None,
            **to_excel_kwargs
    ) -> None:
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        @param filename: File path or existing ExcelWriter
                         (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                           (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                         Per default (startrow=None) calculate the last row
                         in the existing DF and write to the next row...
        @param max_col_width: maximum column width in Excel. Default: 40
        @param autofilter: boolean - whether add Excel autofilter or not. Default: False
        @param fmt_int: Excel format for integer numbers
        @param fmt_float: Excel format for float numbers
        @param fmt_date: Excel format for dates
        @param fmt_datetime: Excel format for datetime's
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                               before writing DataFrame to Excel file
        @param storage_options: dict, optional
            Extra options that make sense for a particular storage connection, e.g. host, port,
            username, password, etc., if using a URL that will be parsed by fsspec, e.g.,
            starting “s3://”, “gcs://”.
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None

        Usage examples:

        >>> append_df_to_excel('/tmp/test.xlsx', df, autofilter=True,
                               freeze_panes=(1,0))

        >>> append_df_to_excel('/tmp/test.xlsx', df, header=None, index=False)

        >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                               index=False)

        >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                               index=False, startrow=25)

        >>> append_df_to_excel('/tmp/test.xlsx', df, index=False,
                               fmt_datetime="dd.mm.yyyy hh:mm")

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        def set_column_format(ws, column_letter, fmt):
            for cell in ws[column_letter]:
                cell.number_format = fmt
        filename = Path(filename)
        file_exists = filename.is_file()
        # process parameters
        # calculate first column number
        # if the DF will be written using `index=True`, then `first_col = 2`, else `first_col = 1`
        first_col = int(to_excel_kwargs.get("index", True)) + 1
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')
        # save content of existing sheets
        if file_exists:
            wb = load_workbook(filename)
            sheet_names = wb.sheetnames
            sheet_exists = sheet_name in sheet_names
            sheets = {ws.title: ws for ws in wb.worksheets}

        with pd.ExcelWriter(
            filename.with_suffix(".xlsx"),
            engine="openpyxl",
            mode="a" if file_exists else "w",
            if_sheet_exists="new" if file_exists else None,
            date_format=fmt_date,
            datetime_format=fmt_datetime,
            storage_options=storage_options
        ) as writer:
            if file_exists:
                # try to open an existing workbook
                writer.book = wb
                # get the last row in the existing Excel sheet
                # if it was not specified explicitly
                if startrow is None and sheet_name in writer.book.sheetnames:
                    startrow = writer.book[sheet_name].max_row
                # truncate sheet
                if truncate_sheet and sheet_name in writer.book.sheetnames:
                    # index of [sheet_name] sheet
                    idx = writer.book.sheetnames.index(sheet_name)
                    # remove [sheet_name]
                    writer.book.remove(writer.book.worksheets[idx])
                    # create an empty sheet [sheet_name] using old index
                    writer.book.create_sheet(sheet_name, idx)
                # copy existing sheets
                writer.sheets = sheets
            else:
                # file doesn't exist, we are creating a new one
                startrow = 0

            # write out the DataFrame to an ExcelWriter
            df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
            worksheet = writer.sheets[sheet_name]

            if autofilter:
                worksheet.auto_filter.ref = worksheet.dimensions

            for xl_col_no, dtyp in enumerate(df.dtypes, first_col):
                col_no = xl_col_no - first_col
                width = max(df.iloc[:, col_no].astype(str).str.len().max(),
                            len(df.columns[col_no]) + 6)
                width = min(max_col_width, width)
                column_letter = get_column_letter(xl_col_no)
                worksheet.column_dimensions[column_letter].width = width
                if np.issubdtype(dtyp, np.integer):
                    set_column_format(worksheet, column_letter, fmt_int)
                if np.issubdtype(dtyp, np.floating):
                    set_column_format(worksheet, column_letter, fmt_float)

        if file_exists and sheet_exists:
            # move (append) rows from new worksheet to the `sheet_name` worksheet
            wb = load_workbook(filename)
            # retrieve generated worksheet name
            new_sheet_name = set(wb.sheetnames) - set(sheet_names)
            if new_sheet_name:
                new_sheet_name = list(new_sheet_name)[0]
            # copy rows written by `df.to_excel(...)` to
            self.copy_excel_cell_range(
                src_ws=wb[new_sheet_name],
                tgt_ws=wb[sheet_name],
                tgt_min_row=startrow + 1,
                with_style=True
            )
            # remove new (generated by Pandas) worksheet
            del wb[new_sheet_name]
            wb.save(filename)
            wb.close()

    #def run(self, ticker, always_run_rl=False):
    def run(self, ticker, run_rl=1, g20_threshold=20):
        result = None
        try:
            #print(self.aaod)
            q = QuoteExplorer()
            q.get_quotes(ticker, self.aaod)

            # default
            short = 3
            long = 7
            signal = 19
            macd_threshold = 0
            macd_min_len = 0

            mongo_rl_param = self.mongo.mongoDB['test_rl_macd_param']
            mongo_query_param = {"symbol": ticker}
            quote_param = mongo_rl_param.find(mongo_query_param)
            if mongo_rl_param.count_documents(mongo_query_param) > 0:
                short = quote_param[0]['short']
                long = quote_param[0]['long']
                signal = quote_param[0]['signal']
                macd_threshold = quote_param[0]['macd_threshold']
                macd_min_len = quote_param[0]['macd_min_len']

            #g20 = StockScore({'AAOD': self.aaod, 'symbol': ticker})
            #if g20.run_fundamentals() == True:
            ss = StockStats(ticker)
            #ss.macd(self.short, self.long, self.signal)
            #print(ss.stock)
            #self.c2 = ss.macd_crossing_by_threshold()
            #last_macd_date = self.c2.loc[len(self.c2)-1, 'date']
            #date_diff = self.today - last_macd_date
            #last_macd_len = self.c2.loc[len(self.c2)-1, 'len']
            m = ss.macd_by_date(self.aaod, short, long, signal)
            #print(m)

            #s = ss.stock
            #s = s.reset_index()
            #last_macd_date = s.iloc[-1]['date'].strftime("%Y-%m-%d")
            #last_macd_cross = s.iloc[-1]['macd_cross']
            #last_macd_r = s.iloc[-1]['r']
            #last_macd_len = s.iloc[-1]['len']
            #print(last_macd_date, last_macd_cross, last_macd_r)

            #if last_macd_date == self.aaod and (last_macd_cross == 1 or last_macd_cross == -1 or last_macd_r < 0.2) and last_macd_len >= 5:
            #if date_diff.days <= 2 and last_macd_len >= 5:
            #if (m['r'] < 0.2 and m['len'] >= 5) or (m['len'] == 1 and m['pre_len'] >= 5):
            #if (m['r'] < 0.2 and m['len'] >= 5) or m['len'] == 1:
            if (m['r'] < macd_threshold and m['len'] >= macd_min_len) or m['len'] == 1:
                print('check g20')
                g20 = StockScore({'AAOD': self.aaod, 'symbol': ticker})
                g20.run(save_rl=True, run_rl=run_rl, retrain_rl=True, g20_threshold=g20_threshold)
                if g20.result.get('rl_result') is not None:
                    #print(self.c2)
                    #print(json.dumps(g20.result, indent=4))
                    print(g20.result)
                    g20.save_g20()
                    result = g20.result
        except Exception as error:
            # print(error)
            print(traceback.format_exc())
            pass
        return result

    def run_all(self, g20_threshold=20):
        #symbol = 'PUMP'  # 'MTDR'     #'CDLX' #'APPN' #'EYE'  #'RNG'     #'GOOGL'     #'EVBG'     #'BAND'     #'NFLX' #'GRWG'     #'ACMR'     #'MDB'     #'RCM'     #'FB'     #'AAPL'     #'ANTM'     #'AMZN'     #'TSLA'   #'BAND'     #'ROKU'     #'SHOP'     #'TWLO'

        mongoDB = self.mongo.mongoDB
        mongo_col = mongoDB['etrade_companies']
        # mongo_query = {"Yahoo_Symbol": 'BRK-A'}
        # mongo_query = {"Yahoo_Symbol": 'APG'}
        # mongo_query = {"Yahoo_Symbol": 'FOE'}
        # mongo_query = {"Yahoo_Symbol": 'AMD'}
        # mongo_query = {"Yahoo_Symbol": 'TSLA'}
        # mongo_query = {"Yahoo_Symbol": 'BILL'}
        #mongo_query = {"Yahoo_Symbol": 'SHOP'}
        #mongo_query = {"Yahoo_Symbol": 'AAPL'}
        #mongo_query = {"Yahoo_Symbol": 'V'}
        #mongo_query = {"Yahoo_Symbol": 'TLND'}
        #mongo_query = {"Yahoo_Symbol": 'XME'}
        # mongo_query = {"Yahoo_Symbol": 'KOD'}
        #mongo_query = {"Yahoo_Symbol": 'NFLX'}
        #mongo_query = {"Yahoo_Symbol": 'NFLX'}
        #mongo_query = {"Yahoo_Symbol": 'AYX'}
        #mongo_query = {"Yahoo_Symbol": 'YUMC'}
        mongo_query = {'status': 'active'}
        com = mongo_col.find(mongo_query, no_cursor_timeout=True)

        #aaod = datetime.now().strftime("%Y-%m-%d")
        #today = datetime.now()

        index = 1
        restartIndex = 1        #5200        #964        #2282        #8074        #2118        #8057        #1568      # 3752
        stopIndex = 1000000  # 3753
        for i in com:
            print(str(index) + "	" + i['Yahoo_Symbol'])
            if index > stopIndex:
                break
            if index >= restartIndex:
                try:
                    if i['Yahoo_Symbol'] not in self.list:      # and i['Yahoo_Symbol'] not in self.exclusionList:
                        cwh = Rule_Cup_with_Handle(self.aaod)
                        cwh_result = cwh.trade_with_cwh(ticker=i['Yahoo_Symbol'], aaod=self.aaod, look_back=20, cwh_back=20, db_look_back=30)
                        #print(cwh_result)

                        self.run(ticker=i['Yahoo_Symbol'], run_rl=1, g20_threshold=g20_threshold)
                except Exception as error:
                    #print(error)
                    print(traceback.format_exc())
                    pass

            #time.sleep(1.00)
            index += 1

    def run_portfolio(self, g20_threshold=20):
        index = 1
        restartIndex = 76        #1568      # 3752
        stopIndex = 1000000  # 3753
        for ticker in self.list:
            print(str(index) + "	" + ticker)
            if index > stopIndex:
                break
            if index >= restartIndex:
                try:
                    cwh = Rule_Cup_with_Handle(self.aaod)
                    cwh_result = cwh.trade_with_cwh(ticker=ticker, aaod=self.aaod, look_back=30, cwh_back=20, db_look_back=30)
                    #print(cwh_result)

                    self.run(ticker=ticker, run_rl=2, g20_threshold=g20_threshold)
                except Exception as error:
                    #print(error)
                    print(traceback.format_exc())
                    pass
            index += 1

    def to_excel(self):
        mongo = MongoExplorer()
        mongo_col = mongo.mongoDB['stock_g_score']
        mongo_query = {"runDate": self.aaod, "$or": [{"Recommendation": ""}, {"symbol": {"$in": self.list}}]}
        g = mongo_col.find(mongo_query, no_cursor_timeout=True)

        #pd.set_option('display.max_rows', None)
        #pd.set_option('display.max_columns', None)
        #pd.set_option('display.width', None)
        #pd.set_option('display.max_colwidth', None)
        df = pd.DataFrame(list(g))

        macd = json_normalize(df.get('MACD'))
        df = df.join(macd)

        rl_result = json_normalize(df.get('rl_result'))
        df = df.join(rl_result)

        excel_file_name = 'test_rl_macd_daily_20221214.xlsx'
        self.append_df_to_excel(excel_file_name, df,
                startcol=17, header=False,
                columns=[
                    'symbol',
                    'Reason',
                    'Recommendation',
                    'model_run_date',
                    'start_date',
                    'end_date',
                    'duration',
                    'model_gain',
                    'model_perf',
                    'buy_and_hold_gain',
                    'buy_and_hold_perf',
                    'model_score',
                    'predict_date',
                    'predict_macd_accum',
                    'predict_macd_len',
                    'predict_action',
                    'predict_vol',
                    'AAOD',
                    'close',
                    'Score',
                    'CAP',
                    'CapScore',
                    'PEScore',
                    'AveragePE',
                    'PE',
                    'EPS',
                    'IndustryRank',
                    'Rev Growth',
                    'Gross Margin',
                    'G20',
                    'SR20',
                    'G20Close',
                    'G20Date',
                    'G20Total',
                    'G10',
                    'SR10',
                    'G10Close',
                    'G10Date',
                    'G10Total',
                    'G5',
                    'SR5',
                    'G5Close',
                    'G5Date',
                    'G5Total',
                    'G1',
                    'SR1',
                    'G1Close',
                    'G1Date',
                    'G1Total',
                    'runDate',
                    'macd_sign',
                    'peak',
                    'peak_date',
                    'accum',
                    'len',
                    'r',
                    'pre_macd_sign',
                    'pre_peak',
                    'pre_peak_date',
                    'pre_accum',
                    'pre_len'
                ])
        #get cwh results
        aaod_2 = (datetime(*(int(s) for s in self.aaod.split('-'))) + timedelta(days=(-1) * 5)).strftime("%Y-%m-%d")
        query1 = {'$and': [{'end_date': {'$lte': self.aaod}}, {'end_date': {'$gte': aaod_2}}]}
        #print(query1)
        cwh = mongo.mongoDB['stock_cwh_results'].find(query1, no_cursor_timeout=True)
        #pd.set_option('display.max_rows', None)
        #pd.set_option('display.max_columns', None)
        #pd.set_option('display.width', None)
        #pd.set_option('display.max_colwidth', None)
        cd = pd.DataFrame(list(cwh))

        cwh_df = pd.DataFrame(columns=['_id', 'symbol', 'cwh_sign', 'cwh_end', 'pearson', 'sigma', 'end_date', 'perf', 'perf_date'])

        for i, c in cd.iterrows():
            if c['cwh_sign'] == 1:
                try:
                    g20 = StockScore({'AAOD': self.aaod, 'symbol': c['symbol']})
                    g20.run(save_rl=True, run_rl=0)
                    if g20.result['Recommendation'] == '':
                        #print(c)
                        cwh_df = pd.concat([cwh_df, cd.iloc[[i]]])
                        #print(cwh_df)
                except:
                    pass
            elif c['cwh_sign'] == 2 and c['symbol'] in self.list:
                cwh_df = pd.concat([cwh_df, cd.iloc[[i]]])

        if not cwh_df.empty:
            cwh_df['Reason'] = cwh_df['cwh_sign']
            cwh_df['Recommendation'] = ''
            cwh_df['model_run_date'] = self.aaod
            cwh_df['start_date'] = ''
            #cwh_df['end_date'] = cwh_df['end-date']
            cwh_df['duration'] = cwh_df['sigma']
            cwh_df['model_gain'] = ''
            cwh_df['model_perf'] = ''
            cwh_df['buy_and_hold_gain'] = ''
            cwh_df['buy_and_hold_perf'] = ''
            cwh_df['model_score'] = cwh_df['pearson']
            #print(cwh_df)

            cwh_df = cwh_df.reset_index()

            self.append_df_to_excel(excel_file_name, cwh_df,
                    startcol=17, header=False,
                    columns=[
                        'symbol',
                        'Reason',
                        'Recommendation',
                        'model_run_date',
                        'start_date',
                        'end_date',
                        'duration',
                        'model_gain',
                        'model_perf',
                        'buy_and_hold_gain',
                        'buy_and_hold_perf',
                        'model_score'
                    ])

if __name__ == '__main__':
    startTime = datetime.now()

    aaod = datetime.now().strftime("%Y-%m-%d")
    #aaod = (datetime.now() + timedelta(days=-1)).strftime("%Y-%m-%d")
    mongo = MongoExplorer()
    mongo_col_q = mongo.mongoDB.get_collection('AMZN')
    qDates = list(mongo_col_q.find().sort("Date", -1))

    max_date = datetime(*(int(s) for s in qDates[0]['Date'].split('-')))
    date_diff = (datetime.now() - max_date).days
    #print(aaod, max_date, max_date.weekday(), date_diff)

    #if date_diff > 1 and max_date.weekday() < 4: aaod = (max_date + timedelta(days=1)).strftime("%Y-%m-%d")
    if date_diff >= 1 and date_diff <= 2 and max_date.weekday() == 4: aaod = max_date.strftime("%Y-%m-%d")
    #if date_diff >= 3 and max_date.weekday() == 4: aaod = (max_date + timedelta(days=3)).strftime("%Y-%m-%d")
    #aaod = '2023-05-05'

    print(aaod)

    #d = RLDailyTrader({'AAOD': aaod, 'short': 6, 'long': 13, 'signal': 9})
    d = RLDailyTrader({'AAOD': aaod})
    d.run_portfolio()
    endTime1 = datetime.now()
    runTime1 = endTime1 - startTime

    #d.to_excel()
    d.run_all()
    d.to_excel()

    endTime = datetime.now()
    runTime2 = endTime - endTime1
    runTime = endTime - startTime
    print('start:', startTime, 'portfolio end:', endTime1, 'portfolio run time: ', runTime1, 'all end: ', endTime, 'all run time: ', runTime2, 'run time: ', runTime)

'''
        excel_file_name = 'stock_g_score_' + aaod.replace('-', '') + '.xlsx'
        w = pd.ExcelWriter(excel_file_name, engine='xlsxwriter')
        df.to_excel(w, sheet_name='Sheet1',
                    columns=[
                        'symbol',
                        'Reason',
                        'Recommendation',
                        'model_run_date',
                        'start_date',
                        'end_date',
                        'duration',
                        'model_gain',
                        'model_perf',
                        'buy_and_hold_gain',
                        'buy_and_hold_perf',
                        'model_score',
                        'predict_date',
                        'predict_macd_accum',
                        'predict_macd_len',
                        'predict_action',
                        'predict_vol',
                        'AAOD',
                        'close',
                        'Score',
                        'CAP',
                        'CapScore',
                        'PEScore',
                        'AveragePE',
                        'PE',
                        'EPS',
                        'IndustryRank',
                        'Rev Growth',
                        'Gross Margin',
                        'G20',
                        'SR20',
                        'G20Close',
                        'G20Date',
                        'G20Total',
                        'G10',
                        'SR10',
                        'G10Close',
                        'G10Date',
                        'G10Total',
                        'G5',
                        'SR5',
                        'G5Close',
                        'G5Date',
                        'G5Total',
                        'G1',
                        'SR1',
                        'G1Close',
                        'G1Date',
                        'G1Total',
                        'runDate',
                        'macd_sign',
                        'peak',
                        'peak_date',
                        'accum',
                        'len',
                        'r',
                        'pre_macd_sign',
                        'pre_peak',
                        'pre_peak_date',
                        'pre_accum',
                        'pre_len'
                    ])
        w.save()

        with pd.ExcelWriter(excel_file_name, engine='openpyxl', mode='a', if_sheet_exists='replace')as w:
            w.book = load_workbook(excel_file_name)
            w.sheets = {ws.title: ws for ws in w.book.worksheets}
            #df.to_excel(w, sheet_name='Sheet1', startrow=w.book['Sheet1'].max_row, startcol=17, header=False,
            df.to_excel(w, sheet_name='Sheet1', startcol=17,
                columns = [
                            'symbol',
                            'Reason',
                            'Recommendation',
                            'model_run_date',
                            'start_date',
                            'end_date',
                            'duration',
                            'model_gain',
                            'model_perf',
                            'buy_and_hold_gain',
                            'buy_and_hold_perf',
                            'model_score',
                            'predict_date',
                            'predict_macd_accum',
                            'predict_macd_len',
                            'predict_action',
                            'predict_vol',
                            'AAOD',
                            'close',
                            'Score',
                            'CAP',
                            'CapScore',
                            'PEScore',
                            'AveragePE',
                            'PE',
                            'EPS',
                            'IndustryRank',
                            'Rev Growth',
                            'Gross Margin',
                            'G20',
                            'SR20',
                            'G20Close',
                            'G20Date',
                            'G20Total',
                            'G10',
                            'SR10',
                            'G10Close',
                            'G10Date',
                            'G10Total',
                            'G5',
                            'SR5',
                            'G5Close',
                            'G5Date',
                            'G5Total',
                            'G1',
                            'SR1',
                            'G1Close',
                            'G1Date',
                            'G1Total',
                            'runDate',
                            'macd_sign',
                            'peak',
                            'peak_date',
                            'accum',
                            'len',
                            'r',
                            'pre_macd_sign',
                            'pre_peak',
                            'pre_peak_date',
                            'pre_accum',
                            'pre_len'
                        ])


'''

#   NET, PCTY, HUM, NOW