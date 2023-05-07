from datetime import datetime
import json
import traceback
from test_mongo import MongoExplorer
import pandas as pd
from pandas import json_normalize
from test_yahoo import QuoteExplorer
from test_g20 import StockScore
from openpyxl import load_workbook

class RLTrader:
    def __init__(self, options):
        self.runDate = options['AAOD']
        self.AAOD = options['AAOD']
        self.symbol = options['symbol']

if __name__ == '__main__':
    aaod = datetime.now().strftime("%Y-%m-%d")
    #symbol = 'SHOP'
    symbol = 'NET'
    try:
        q = QuoteExplorer()
        q.get_quotes(symbol, aaod)

        g20 = StockScore({'AAOD': aaod, 'symbol': symbol})
        g20.run()
        print(json.dumps(g20.result, indent=4))
        g20.save_g20()

        mongo = MongoExplorer()
        mongo_col = mongo.mongoDB['stock_g_score']
        mongo_query = {"runDate": aaod, "Recommendation": ""}
        g = mongo_col.find(mongo_query, no_cursor_timeout=True)
        df = pd.DataFrame(list(g))

        macd = json_normalize(df['MACD'])
        df = df.join(macd)

        rl_result = json_normalize(df['rl_result'])
        df = df.join(rl_result)

        excel_file_name = 'test_rl_macd_trade.xlsx'
        w = pd.ExcelWriter(excel_file_name, engine='openpyxl', mode='a')
        w.book = load_workbook(excel_file_name)
        w.sheets = {ws.title: ws for ws in w.book.worksheets}
        df.to_excel(w, sheet_name='Sheet1', startrow=w.book['Sheet1'].max_row, startcol=8, header=False,
                    columns=[
                        'symbol',
                        'Reason',
                        'Recommendation',
                        'model_run_date',
                        'start_date',
                        'end_date',
                        'duration',
                        'model_perf',
                        'buy_and_hold',
                        'model_score',
                        'predict_date',
                        'predict_macd_accum',
                        'predict_macd_len',
                        'predict_action',
                        'predict_vol',
                        'AAOD',
                        'close',
                        'Score',
                        'CAP',
                        'CapScore',
                        'PEScore',
                        'AveragePE',
                        'PE',
                        'EPS',
                        'G20',
                        'SR20',
                        'G20Close',
                        'G20Date',
                        'G20Total',
                        'G10',
                        'SR10',
                        'G10Close',
                        'G10Date',
                        'G10Total',
                        'G5',
                        'SR5',
                        'G5Close',
                        'G5Date',
                        'G5Total',
                        'G1',
                        'SR1',
                        'G1Close',
                        'G1Date',
                        'G1Total',
                        'runDate',
                        'macd_sign',
                        'peak',
                        'peak_date',
                        'accum',
                        'len',
                        'r',
                        'pre_macd_sign',
                        'pre_peak',
                        'pre_peak_date',
                        'pre_accum',
                        'pre_len'
                    ])
        w.save()

    except Exception as error:
        # print(error)
        print(traceback.format_exc())
