from datetime import datetime
import json
import numpy as np
import traceback
from test_mongo import MongoExplorer
import pandas as pd
from pandas import json_normalize
from test_yahoo import QuoteExplorer
from test_g20_v2 import StockScore
from openpyxl import load_workbook

class NpEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        else:
            return super(NpEncoder, self).default(obj)

class RLTrader:
    def __init__(self, options):
        self.runDate = options['AAOD']
        self.AAOD = options['AAOD']
        self.symbol = options['symbol']

if __name__ == '__main__':
    aaod = datetime.now().strftime("%Y-%m-%d")
    #symbol = 'SHOP'         #'PENN'         #'BKNG'     #'ZM'       #'NFLX'     #'NTES'     #'ISRG'     #'FTNT'      #'GOLF'     #'MTDR'      #'W'    #'SHOP'     #'AYX'      #'TWLO'     #'PUMP'     #'MTDR'     #'CDLX' #'APPN' #'EYE'  #'RNG'     #'GOOGL'     #'EVBG'     #'BAND'     #'NFLX' #'GRWG'     #'ACMR'     #'MDB'     #'RCM'     #'FB'     #'AAPL'     #'ANTM'     #'AMZN'     #'TSLA'   #'BAND'     #'ROKU'     #'SHOP'     #'TWLO'
    #3symbol = 'SHOP'
    symbol = 'NET'
    try:
        q = QuoteExplorer()
        q.get_quotes(symbol, aaod)

        g20 = StockScore({'AAOD': aaod, 'symbol': symbol})
        g20.run(save_rl=True, always_run_rl=True)
        #print(g20.result)
        print(json.dumps(g20.result, indent=4, cls=NpEncoder))
        g20.save_g20()

        mongo = MongoExplorer()
        mongo_col = mongo.mongoDB['stock_g_score']
        #mongo_query = {"symbol": symbol, "runDate": aaod, "Recommendation": ""}
        mongo_query = {"symbol": symbol, "runDate": aaod}
        g = mongo_col.find(mongo_query, no_cursor_timeout=True)
        df = pd.DataFrame(list(g))

        macd = json_normalize(df['MACD'])
        df = df.join(macd)

        rl_result = json_normalize(df['rl_result'])
        df = df.join(rl_result)

        excel_file_name = 'test_rl_macd_trade_v2.xlsx'
        with pd.ExcelWriter(excel_file_name, engine='openpyxl', mode='a')as w:
            w.book = load_workbook(excel_file_name)
            w.sheets = {ws.title: ws for ws in w.book.worksheets}
            df.to_excel(w, sheet_name='Sheet1', startrow=w.book['Sheet1'].max_row, startcol=17, header=False,
                        columns=[
                            'symbol',
                            'Reason',
                            'Recommendation',
                            'model_run_date',
                            'start_date',
                            'end_date',
                            'duration',
                            'model_gain',
                            'model_perf',
                            'buy_and_hold_gain',
                            'buy_and_hold_perf',
                            'model_score',
                            'predict_date',
                            'predict_macd_accum',
                            'predict_macd_len',
                            'predict_action',
                            'predict_vol',
                            'AAOD',
                            'close',
                            'Score',
                            'CAP',
                            'CapScore',
                            'PEScore',
                            'AveragePE',
                            'PE',
                            'EPS',
                            'IndustryRank',
                            'Rev Growth',
                            'Gross Margin',
                            'G20',
                            'SR20',
                            'G20Close',
                            'G20Date',
                            'G20Total',
                            'G10',
                            'SR10',
                            'G10Close',
                            'G10Date',
                            'G10Total',
                            'G5',
                            'SR5',
                            'G5Close',
                            'G5Date',
                            'G5Total',
                            'G1',
                            'SR1',
                            'G1Close',
                            'G1Date',
                            'G1Total',
                            'runDate',
                            'macd_sign',
                            'peak',
                            'peak_date',
                            'accum',
                            'len',
                            'r',
                            'pre_macd_sign',
                            'pre_peak',
                            'pre_peak_date',
                            'pre_accum',
                            'pre_len'
                        ])
            w.save()
            w.close()
    except Exception as error:
        # print(error)
        print(traceback.format_exc())
